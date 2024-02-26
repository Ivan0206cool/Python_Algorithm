import pandas as pd
#from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles import DEFAULT_FONT
# from ChartGen import *
import re
import ntpath
import json, itertools
import os
from pathlib import Path
from datetime import datetime




def get_card_config_action(info):
    if "init" in info:
        card = ""
        config_type = ""
        card_action = info
        card_round = ""
    elif "after action" in info or "before action" in info or "Reget action" in info :
        if "after action" in info:
            action = "After"
        elif "before action" in info:
            action = "Before"
        elif "Reget action" in info:
            action = "Reget after"
        card_info = re.sub(r"(-record after action|-record before action|-Reget action)","", info)
        card = card_info.split('-')[1]
        config_type = card_info.split('-')[2]
        card_action = action + " " +card_info.split('-')[0]
        caround = card_info.split('-')[3]
        card_round = float(re.match(r"([a-zA-Z]+)(\d+)", caround).group(2))
    else:
        # card = info.split('-')[1]
        # config_type = info.split('-')[2]
        # card_action = info.split('-')[0]
        # caround = info.split('-')[3]
        # === clka
        card = info.split('-')[1]
        SSM_info = (info.split('-')[2])[3:]
        card_action = info.split('-')[0]
        caround = info.split('-')[4]
        FDL_info = (info.split('-')[3])[3:]
        # === clka
        card_round = float(re.match(r"([a-zA-Z]+)(\d+)", caround).group(2))
    # return card,config_type,card_action,card_round
    return card,SSM_info,FDL_info,card_action,card_round
def merge_row_title_for_sheet(name,sheet,row,col,end_row,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center',vertical='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_row-row >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=end_row,start_column=col,end_column=col)
def read_line_and_split_time(f_read):
    full_line = f_read.readline()
    if not full_line:
        return full_line, ''
    else:
        text_line = re.split(
            r'\d{4}\-\d{2}-\d{2} \d{2}:\d{2}:\d{2} : ', full_line)[1]
        time_stamp = re.findall(
            r'\d{4}\-\d{2}-\d{2} \d{2}:\d{2}:\d{2} : ', full_line)[0].strip(" : ")
        return time_stamp, text_line
def get_dut_info(input_report_path):
    f_read = open(input_report_path, 'r', encoding="utf-8")
    time_stamp, text_line = read_line_and_split_time(
        f_read)  # first line is 'begin'
    if 'begin' in text_line:
        time_stamp, text_line = read_line_and_split_time(f_read)
    dut_info = {}
    while "dut" in text_line.lower():
        if re.search(r'DUT[\d]{1} is', text_line):
            dut_num = text_line.split('is')[0].split('DUT')[1].strip()
            matches = re.findall(r'\b(LOOP|GE)([^\,]+)', text_line)
            for match in matches:
                dut_chassis = match[0] + ' ' + match[1].strip()
            print(dut_chassis)
            dut_slot_cnt = text_line.split(',slot#')[1].split(',')[
                0].split('\n')[0]
            dut_info[dut_num] = {
                "Chassis": dut_chassis,
                "Total slots": dut_slot_cnt,
                "Slot Info": []
            }
        elif "dut=" in text_line.lower():
            dut_num = text_line.split('=')[1].split(' ')[0]
            dut_slot = text_line.split('=')[1].split(' ')[1]
            if "CTRL" in dut_slot:
                ctrl_hw_ver = text_line.split('HW=')[1].split(' ')[0]
                ctrl_sw_ver = text_line.split('HW=')[0].split(' ')[2].split('=')[1]
                ctrl_serial_ver = text_line.split('HW=')[0].split(' ')[3].split('=')[1]
                print(text_line)
                print(ctrl_hw_ver)
                print(ctrl_sw_ver)
                if text_line.split('=')[2].split(" ")[0].split("L")[1] == "1":  # ctrl1
                    print(text_line)
                    # ctrl_fpga1_ver = text_line.split('FPGA1=')[1].split(" ")[0]
                    # ctrl_fpga2_ver = text_line.split('FPGA2=')[1].split(" ")[0]
                    # ctrl_cpld_ver = text_line.split('CPLD=')[1].split(" ")[0]
                    dut_info[dut_num]["Slot Info"].append(
                        {
                            "Name": "Ctrl 1",
                            "HW Version": ctrl_hw_ver,
                            "SW Version": ctrl_sw_ver,
                            "Serial Version": ctrl_serial_ver,
                            # "FPGA1 Version": ctrl_fpga1_ver,
                            # "FPGA2 Version": ctrl_fpga2_ver,
                            # "CPLD Version": ctrl_cpld_ver
                        })
                else:
                    dut_info[dut_num]["Slot Info"].append(
                        {
                            "Name": "Ctrl 2",
                            "HW Version": ctrl_hw_ver,
                            "SW Version": ctrl_sw_ver,
                            "Serial Version": ctrl_serial_ver
                        })
            elif "slot" in dut_slot:
                slot_info = re.findall(r'slot:(.*)\s+ver', text_line)
                print(slot_info)
                slot_num = str(slot_info).split()[1].strip()
                slot_card_type = str(slot_info).split()[2].strip()
                slot_ver = text_line.split('ver:')[1].split('"')[1]
                dut_info[dut_num]["Slot Info"].append(
                    {
                        "Name": F"Slot {slot_num}",
                        "Card Type": slot_card_type,
                        "Version": slot_ver
                    })

        time_stamp, text_line = read_line_and_split_time(f_read)
    f_read.close
    with open(F'{os.path.dirname(input_report_path)}/dut_info.json', 'w') as fw:
        fw.write(json.dumps(dut_info, indent=4))
    return dut_info
def change_font_for_sheet(sheet, font):
    max_row = sheet.max_row
    max_column  = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            sheet.cell(i,j).font = Font(name=font)
def adjust_column_width(sheet):
    new_column_length = 0
    for column_cells in sheet.columns:
        break_line_count = max(str(cell.value).count('\n') for cell in column_cells)
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        if break_line_count > 0 :
            new_column_length /= break_line_count
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            sheet.column_dimensions[new_column_letter].width = new_column_length*1.23
def add_border_in_excel(sheet):
    thin = Side(border_style="thin", color="000000")
    max_row = sheet.max_row
    max_column  = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
def set_title_for_sheet(name,sheet,row,col,end_col,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_col-col >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=row,start_column=col,end_column=end_col)
def add_item_in_test_summary(test_summary, item, sheet_link_name):
    test_summary[item] = {
        "pass": 0,
        "fail": 0,
        'warning': 0,
        "sheet link": sheet_link_name
    }
def set_info_title_for_sheet(sheet,first_row):
    # set_title_for_sheet('Test Card',sheet,first_row,1,1,"5cb800")
    # set_title_for_sheet('Config type',sheet,first_row,2,2,"5cb800")
    # set_title_for_sheet('Action',sheet,first_row,3,3,"5cb800")
    # set_title_for_sheet('Round',sheet,first_row,4,4,"5cb800")
    # set_title_for_sheet('Pass',sheet,first_row,5,5,"5cb800")
    # set_title_for_sheet('Fail',sheet,first_row,6,6,"5cb800")
    # set_title_for_sheet('Warning',sheet,first_row,7,7,"5cb800")
    # === clka
    set_title_for_sheet('Test Card',sheet,first_row,1,1,"5cb800")
    set_title_for_sheet('SSM Sabit',sheet,first_row,2,2,"5cb800")
    set_title_for_sheet('FDL Sabit',sheet,first_row,3,3,"5cb800")
    set_title_for_sheet('Action',sheet,first_row,4,4,"5cb800")
    set_title_for_sheet('Round',sheet,first_row,5,5,"5cb800")
    set_title_for_sheet('Pass',sheet,first_row,6,6,"5cb800")
    set_title_for_sheet('Fail',sheet,first_row,7,7,"5cb800")
    set_title_for_sheet('Warning',sheet,first_row,8,8,"5cb800")
    for i in range(8):
        sheet.merge_cells(start_row=1,end_row=2,start_column=i+1,end_column=i+1)
def set_info_title_for_fail_sheet(sheet,first_row):
    set_title_for_sheet('Test Card',sheet,first_row,1,1,"5cb800")
    set_title_for_sheet('Config type',sheet,first_row,2,2,"5cb800")
    set_title_for_sheet('Action',sheet,first_row,3,3,"5cb800")
    set_title_for_sheet('Round',sheet,first_row,4,4,"5cb800")
    for i in range(4):
        sheet.merge_cells(start_row=1,end_row=2,start_column=i+1,end_column=i+1)
def generate_test_case(case_mapping_table):
    case_translation_table = {}
    test_case_length = {}
    test_counter = {}
    for trans_name, ini_name in case_mapping_table.items():
        if ini_name['first_case'] not in case_translation_table:
            case_translation_table[ini_name['first_case']] = trans_name
        if ini_name['second_case'] not in case_translation_table:
            case_translation_table[ini_name['second_case']] = trans_name
        if ini_name['first_case'] not in test_case_length:
            test_case_length[ini_name['first_case']] = 0
        if ini_name['second_case'] not in test_case_length:
            test_case_length[ini_name['second_case']] = 0
        if trans_name not in test_counter:
            test_counter[trans_name] = 0
    return case_translation_table, test_case_length, test_counter
def generate_fail_item_set(test_case,case_translation_table1):
    fail_item = {}
    if test_case in case_translation_table1:
        fail_item_set_name = case_translation_table1.get(test_case)
        fail_item[fail_item_set_name] = set()
    elif test_case in bert_case:
        fail_item[test_case] = set()
    return fail_item
# === parser function === #
def slot_number_transform(value):
    slot_name = {
        1 : 'A',
        2 : 'B',
        3 : 'C',
        4 : 'D',
        5 : '1',
        6 : '2',
        7 : '3',
        8 : '4',
        9 : '5',
        10 : '6',
        11 : '7',
        12 : '8',
        13 : '9',
        14 : '10',
        15 : '11',
        16 : '12',
    }
    if value in slot_name:
        corresponding_string = slot_name[value]
        return corresponding_string
    else:
        return f"None"    
def RS485_Status_parse(info,test_case,case_translation_table1):
    data = {}
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    result = 'ok'
    if test_case == 'Rs485SlotStatistEntry':
        data['RS485 info by slot'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['RS485 info by slot']['DUT'] = int(val)
            elif key == 'Slot':
                data['RS485 info by slot']['Slot'] = slot_number_transform(int(val))
            elif 'Txcount' in key:
                data['RS485 info by slot']['Tx count'] = int(val)
            elif 'Rxcount' in key:
                data['RS485 info by slot']['Rx count'] = int(val)
            elif 'Norxcount' in key:
                data['RS485 info by slot']['No RX count'] = int(val)
            elif 'Timeoutcount' in key:
                data['RS485 info by slot']['Timeout count'] = int(val)
            elif 'Alivecount' in key:
                data['RS485 info by slot']['Alive count'] = int(val)
            elif 'Nonalivecount' in key:
                data['RS485 info by slot']['Non-alive count'] = int(val)
            elif 'Oversizecount' in key:
                data['RS485 info by slot']['Oversize count'] = int(val)
            elif 'Initcount' in key:
                data['RS485 info by slot']['Init count'] = int(val)
            elif 'Badcmdcount' in key:
                data['RS485 info by slot']['Bad CMD count'] = int(val)
            elif 'Badcmdnum' in key:
                data['RS485 info by slot']['Bad CMD num'] = int(val)
        data['RS485 info by slot']['Result'] = 'ok'
        if data['RS485 info by slot']['No RX count'] > 0:
            data['RS485 info by slot']['Result'] = 'warning'
            fail_item['RS485 info by slot'].add('No RX count')
        if data['RS485 info by slot']['Timeout count'] > 0:
            data['RS485 info by slot']['Result'] = 'warning'
            fail_item['RS485 info by slot'].add('Timeout count')
        if data['RS485 info by slot']['Non-alive count'] > 0:
            data['RS485 info by slot']['Result'] = 'warning'
            fail_item['RS485 info by slot'].add('Non-alive count')
        if data['RS485 info by slot']['Result'] == 'fail':
            fail_item['RS485 info by slot'].add('Result')
            test_summary['RS485 Status']["fail"] += 1
        elif data['RS485 info by slot']['Result'] == 'warning':
            fail_item['RS485 info by slot'].add('Result')
            test_summary['RS485 Status']["warning"] += 1
        else:
            test_summary['RS485 Status']["pass"] += 1
    else:
        data['RS485 info'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['RS485 info']['DUT'] = int(val)
            elif 'Rs485Buffererror' in key:
                data['RS485 info']['Buffer error'] = int(val)
            elif 'Rs485Sliperror' in key:
                data['RS485 info']['Slip error'] = int(val)
            elif 'Rs485FCSerror' in key:
                data['RS485 info']['FCS error'] = int(val)
        data['RS485 info']['Result'] = 'ok'
        if data['RS485 info']['Buffer error'] > 0:
            data['RS485 info']['Result'] = 'warning'
            fail_item['RS485 info'].add('Buffer error')
        if data['RS485 info']['FCS error'] > 0:
            data['RS485 info']['Result'] = 'fail'
            fail_item['RS485 info'].add('FCS error')
        if data['RS485 info']['Slip error'] > 0:
            data['RS485 info']['Result'] = 'warning'
            fail_item['RS485 info'].add('Slip error')
        if data['RS485 info']['Result'] == 'fail':
            fail_item['RS485 info'].add('Result')
            test_summary['RS485 Status']["fail"] += 1
        elif data['RS485 info']['Result'] == 'warning':
            fail_item['RS485 info'].add('Result')
            test_summary['RS485 Status']["warning"] += 1
        else:
            test_summary['RS485 Status']["pass"] += 1
    return data, fail_item
def PW_Status_parse(info,test_case,case_translation_table1):
    data = {}
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    result = 'ok'
    if test_case == 'PwBundleStaticsEntry':
        data['Working PW Status'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Working PW Status']['DUT'] = int(val)
            elif key == 'CTRL':
                data['Working PW Status']['CTRL'] = int(val)
            elif 'PwStaticsPort' in key:
                data['Working PW Status']['Bundle ID'] = int(val)
            elif 'PwPath' in key:
                data['Working PW Status']['Primary/Secondary'] = val
            elif 'PwJitBufUnderRun' in key:
                data['Working PW Status']['Jitter buffer Under run'] = float(val)
            elif 'PwJitBufOverRun' in key:
                data['Working PW Status']['Jitter buffer Over run'] = float(val)
            elif 'PwJitBufDepth' in key:
                data['Working PW Status']['Jitter buffer Depth'] = float(val)
            elif 'PwJitBufMin' in key:
                data['Working PW Status']['Jitter buffer MIN'] = float(val)
            elif 'PwJitBufMax' in key:
                data['Working PW Status']['Jitter buffer MAX'] = float(val)
            elif 'PwLBitCount' in key:
                data['Working PW Status']['L bit count'] = int(val)
            elif 'PwRBitCount' in key:
                data['Working PW Status']['R bit count'] = int(val)
            elif 'PwTxGoodCount' in key:
                data['Working PW Status']['TX Good count'] = int(val)
            elif 'PwRxGoodCount' in key:
                data['Working PW Status']['RX Good count'] = int(val)
            elif 'PwRxLostCount' in key:
                data['Working PW Status']['RX lost count'] = int(val)
            elif 'PwLink' in key: 
                data['Working PW Status']['Link'] = val
        data['Working PW Status']['Result'] = 'ok'
        if data['Working PW Status']['Jitter buffer Under run'] != 0.0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('Jitter buffer Under run')
        if data['Working PW Status']['Jitter buffer Over run'] != 0.0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('Jitter buffer Over run')
        if data['Working PW Status']['Jitter buffer Depth'] > 12.0 or data['Working PW Status']['Jitter buffer Depth'] < 4.0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('Jitter buffer Depth')
        if data['Working PW Status']['Jitter buffer MIN'] < 4.0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('Jitter buffer MIN')
        if data['Working PW Status']['Jitter buffer MAX'] > 20.0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('Jitter buffer MAX')
        if data['Working PW Status']['L bit count'] != 0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('L bit count')
        if data['Working PW Status']['R bit count'] != 0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('R bit count')
        if data['Working PW Status']['RX lost count'] > 0:
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('RX lost count')
        if data['Working PW Status']['Link'] != "Up":
            data['Working PW Status']['Result'] = 'fail'
            fail_item['Working PW Status'].add('Link')        
        if data['Working PW Status']['Result'] != 'ok':
            fail_item['Working PW Status'].add('Result')
            test_summary['PW Status']["fail"] += 1
        else:
            test_summary['PW Status']["pass"] += 1
    elif test_case == 'PwStandbyStaticsEntry':
        data['Standby PW Status'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Standby PW Status']['DUT'] = int(val)
            elif key == 'CTRL':
                data['Standby PW Status']['CTRL'] = int(val)
            elif 'PwStandbyStaticsPort' in key:
                data['Standby PW Status']['Bundle ID'] = int(val)
            elif 'PwStandbyPath' in key:
                data['Standby PW Status']['Primary/Secondary'] = val
            elif 'PwStandbyJitBufUnderRun' in key:
                data['Standby PW Status']['Jitter buffer Under run'] = float(val)
            elif 'PwStandbyJitBufOverRun' in key:
                data['Standby PW Status']['Jitter buffer Over run'] = float(val)
            elif 'PwStandbyJitBufDepth' in key:
                data['Standby PW Status']['Jitter buffer Depth'] = float(val)
            elif 'PwStandbyJitBufMin' in key:
                data['Standby PW Status']['Jitter buffer MIN'] = float(val)
            elif 'PwStandbyJitBufMax' in key:
                data['Standby PW Status']['Jitter buffer MAX'] = float(val)
            elif 'PwStandbyLBitCount' in key:
                data['Standby PW Status']['L bit count'] = int(val)
            elif 'PwStandbyRBitCount' in key:
                data['Standby PW Status']['R bit count'] = int(val)
            elif 'PwStandbyTxGoodCount' in key:
                data['Standby PW Status']['TX Good count'] = int(val)
            elif 'PwStandbyRxGoodCount' in key:
                data['Standby PW Status']['RX Good count'] = int(val)
            elif 'PwStandbyRxLostCount' in key:
                data['Standby PW Status']['RX lost count'] = int(val)
            elif 'PwStandbyLink' in key: 
                data['Standby PW Status']['Link'] = val
        data['Standby PW Status']['Result'] = 'ok'
        if data['Standby PW Status']['Jitter buffer Under run'] != 0.0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('Jitter buffer Under run')
        if data['Standby PW Status']['Jitter buffer Over run'] != 0.0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('Jitter buffer Over run')
        if data['Standby PW Status']['Jitter buffer Depth'] > 12.0 or data['Standby PW Status']['Jitter buffer Depth'] < 4.0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('Jitter buffer Depth')
        if data['Standby PW Status']['Jitter buffer MIN'] < 4.0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('Jitter buffer MIN')
        if data['Standby PW Status']['Jitter buffer MAX'] > 20.0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('Jitter buffer MAX')
        if data['Standby PW Status']['L bit count'] != 0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('L bit count')
        if data['Standby PW Status']['R bit count'] != 0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('R bit count')
        if data['Standby PW Status']['RX lost count'] > 0:
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('RX lost count')
        if data['Standby PW Status']['Link'] != "Up":
            data['Standby PW Status']['Result'] = 'fail'
            fail_item['Standby PW Status'].add('Link')        
        if data['Standby PW Status']['Result'] != 'ok':
            fail_item['Standby PW Status'].add('Result')
            test_summary['PW Status']["fail"] += 1
        else:
            test_summary['PW Status']["pass"] += 1
    return data, fail_item
def PW_Statist_parse(info,test_case,case_translation_table1):
    data = {}
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    result = 'ok'
    if test_case == 'PwPriProtectStatistEntry':
        data['Working PW Statist'] = {}
        data['Working PW Statist']['Primary/Secondary'] = 'Pri'
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Working PW Statist']['DUT'] = int(val)
            elif 'PwPriProtectIndex' in key:
                data['Working PW Statist']['Bundle ID'] = int(val)
            # elif '' in key:
            #     data['Working PW Statist']['bid_index'] = float(val)
            elif 'PwPriEth1Linkstat' in key:
                data['Working PW Statist']['Eth1LinkStat'] = val
            elif 'PwPriEth1RxSeqNum' in key:
                data['Working PW Statist']['Eth1RxSeqNum'] = float(val)
            elif 'PwPriEth1RxSeqLos' in key:
                data['Working PW Statist']['Eth1RxSeqLos'] = float(val)
            elif 'PwPriEth1RxJBLos' in key:
                data['Working PW Statist']['Eth1RxJBLos'] = int(val)
            elif 'PwPriDiffDelay' in key:
                data['Working PW Statist']['DiffDelay'] = int(val)
            # elif 'PwLink' in key: 
            #     data['Working PW Statist']['Link'] = val
        data['Working PW Statist']['Result'] = 'ok'
        if data['Working PW Statist']['Eth1RxJBLos'] > 0:
            data['Working PW Statist']['Result'] = 'fail'
            fail_item['Working PW Statist'].add('Eth1RxJBLos')
        if data['Working PW Statist']['Result'] != 'ok':
            fail_item['Working PW Statist'].add('Result')
            test_summary['PW Statist']["fail"] += 1
        else:
            test_summary['PW Statist']["pass"] += 1
    elif test_case == 'PwSecProtectStatistEntry':
        data['Standby PW Statist'] = {}
        data['Standby PW Statist']['Primary/Secondary'] = 'Sec'
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Standby PW Statist']['DUT'] = int(val)
            elif 'PwSecProtectIndex' in key:
                data['Standby PW Statist']['Bundle ID'] = int(val)
            # elif '' in key:
            #     data['Standby PW Statist']['bid_index'] = float(val)
            elif 'PwSecEth1Linkstat' in key:
                data['Standby PW Statist']['Eth1LinkStat'] = val
            elif 'PwSecEth1RxSeqNum' in key:
                data['Standby PW Statist']['Eth1RxSeqNum'] = float(val)
            elif 'PwSecEth1RxSeqLos' in key:
                data['Standby PW Statist']['Eth1RxSeqLos'] = float(val)
            elif 'PwSecEth1RxJBLos' in key:
                data['Standby PW Statist']['Eth1RxJBLos'] = int(val)
            elif 'PwSecDiffDelay' in key:
                data['Standby PW Statist']['DiffDelay'] = val
            # elif 'PwLink' in key: 
            #     data['Standby PW Statist']['Link'] = val
        data['Standby PW Statist']['Result'] = 'ok'
        if data['Standby PW Statist']['Eth1RxJBLos'] > 0:
            data['Standby PW Statist']['Result'] = 'fail'
            fail_item['Standby PW Statist'].add('Eth1RxJBLos')
        if data['Standby PW Statist']['Result'] != 'ok':
            fail_item['Standby PW Statist'].add('Result')
            test_summary['PW Statist']["fail"] += 1
        else:
            test_summary['PW Statist']["pass"] += 1
    return data, fail_item
def Card_Port_Status_parse(info,test_case,case_translation_table1):
    data = {}
    result = 'ok'
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    if test_case == 'Qe1PortStatusEntry':
        data['QE1'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['QE1']['DUT'] = int(val)
            elif 'Qe1PSSlotIndex' in key:
                data['QE1']['Slot'] = int(val)
            elif 'Qe1PSPortIndex' in key:
                data['QE1']['Port'] = int(val)
            elif 'Qe1PortLOS' in key:
                data['QE1']['Qe1PortLOS'] = val
            elif 'Qe1PortLOF' in key:
                data['QE1']['Qe1PortLOF'] = val
            elif 'Qe1PortRcvAIS' in key:
                data['QE1']['Qe1PortRcvAIS'] = val
            elif 'Qe1PortRcvRAI' in key:
                data['QE1']['Qe1PortRcvRAI'] = val
            elif 'Qe1PortXmtRAI' in key:
                data['QE1']['Qe1PortXmtRAI'] = val
            elif 'Qe1PortXmtAIS' in key:
                data['QE1']['Qe1PortXmtAIS'] = val
            elif 'Qe1PortBPVcount' in key:
                data['QE1']['Qe1PortBPVcount'] = int(val)
            elif 'Qe1PortEScount' in key:
                data['QE1']['Qe1PortEScount'] = int(val)
        data['QE1']['Result'] = 'ok'
        if data['QE1']['Qe1PortLOS'] == 'Yes':
            data['QE1']['Result'] = 'fail'
            fail_item['QE1'].add('Qe1PortLOS')
        if data['QE1']['Qe1PortLOF'] == 'Yes':
            data['QE1']['Result'] = 'fail'
            fail_item['QE1'].add('Qe1PortLOF')        
        if data['QE1']['Result'] != 'ok':
            fail_item['QE1'].add('Result')
            test_summary['Card Port Status']["fail"] += 1
        else:
            test_summary['Card Port Status']["pass"] += 1
    elif test_case == 'FomStatusEntry':
        data['FOM'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['FOM']['DUT'] = int(val)
            elif 'Qe1PSSlotIndex' in key:
                data['FOM']['Slot'] = int(val)
            elif 'Qe1PSPortIndex' in key:
                data['FOM']['Port'] = int(val)
            elif 'FomType' in key:
                data['FOM']['FomType'] = val
            elif 'FomOpticalLOS' in key:
                data['FOM']['FomOpticalLOS'] = val
            elif 'FomOpticalLOF' in key:
                data['FOM']['FomOpticalLOF'] = val
            elif 'FomOpticalFE' in key:
                data['FOM']['FomOpticalFE'] = int(val)
            elif 'FomOpticalFCS' in key:
                data['FOM']['FomOpticalFCS'] = int(val)
            elif 'FomOpticalEOC' in key:
                data['FOM']['FomOpticalEOC'] = int(val)
            elif 'FomE1LOF' in key:
                data['FOM']['FomE1LOF'] = val
            elif 'FomE1CRC' in key:
                data['FOM']['FomE1CRC'] = int(val)
            elif 'FomE1RxAIS' in key:
                data['FOM']['FomE1RxAIS'] = val
            elif 'FomE1RxRAI' in key:
                data['FOM']['FomE1RxRAI'] = val
            elif 'FomE1TxAIS' in key:
                data['FOM']['FomE1TxAIS'] = val
            elif 'FomE1TxRAI' in key:
                data['FOM']['FomE1TxRAI'] = val
            elif 'FomE1LOS' in key:
                data['FOM']['FomE1LOS'] = val
            elif 'FomE1BPVCRC' in key:
                data['FOM']['FomE1BPVCRC'] = int(val)
        data['FOM']['Result'] = 'ok'
        if data['FOM']['FomOpticalLOS'] == 'Yes':
            data['FOM']['Result'] = 'fail'
            fail_item['FOM'].add('FomOpticalLOS')
        if data['FOM']['FomOpticalLOF'] == 'Yes':
            data['FOM']['Result'] = 'fail'
            fail_item['FOM'].add('FomOpticalLOF')        
        if data['FOM']['FomE1LOF'] == 'Yes':
            data['FOM']['Result'] = 'fail'
            fail_item['FOM'].add('FomE1LOF')
        if data['FOM']['FomE1LOS'] == 'Yes':
            data['FOM']['Result'] = 'fail'
            fail_item['FOM'].add('FomE1LOS')    
        if data['FOM']['Result'] != 'ok':
            fail_item['FOM'].add('Result')
            test_summary['Card Port Status']["fail"] += 1
        else:
            test_summary['Card Port Status']["pass"] += 1
    return data, fail_item
def DPLL_Status_parse(info,test_case,case_translation_table1):
    data = {}
    result = 'ok'
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    if test_case == 'DpllPriStatusTable':
        data['Working'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Working']['DUT'] = int(val)
            elif 'DpllInput1Status' in key:
                data['Working']['Input Status'] = val
            elif 'DpllLockStatus' in key:
                data['Working']['Lock Status'] = val
            elif 'DpllInputSticky' in key:
                data['Working']['Input Sticky'] = val
            elif 'DpllLockSticky' in key:
                data['Working']['Lock Sticky'] = val
        data['Working']['Result'] = 'ok'
        if data['Working']['Input Status'] != 'Normal':
            data['Working']['Result'] = 'fail'
            fail_item['Working'].add('Input Status')      
        if data['Working']['Lock Status'] == 'Unlock':
            data['Working']['Result'] = 'fail'
            fail_item['Working'].add('Lock Status')   
        if data['Working']['Input Sticky'] != 'Normal':
            data['Working']['Result'] = 'fail'
            fail_item['Working'].add('Input Sticky')   
        if data['Working']['Lock Sticky'] == 'Unlock':
            data['Working']['Result'] = 'fail'
            fail_item['Working'].add('Lock Sticky')   
        if data['Working']['Result'] != 'ok':
            fail_item['Working'].add('Result')
            test_summary['DPLL Status']["fail"] += 1
        else:
            test_summary['DPLL Status']["pass"] += 1
    elif test_case == 'DpllStandbyStatusTable':
        data['Standby'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Standby']['DUT'] = int(val)
            elif 'StandbydpllInput1Status' in key:
                data['Standby']['Input Status'] = val
            elif 'StandbydpllLockStatus' in key:
                data['Standby']['Lock Status'] = val
            elif 'StandbydpllInputSticky' in key:
                data['Standby']['Input Sticky'] = val
            elif 'StandbydpllLockSticky' in key:
                data['Standby']['Lock Sticky'] = val
        data['Standby']['Result'] = 'ok'
        if data['Standby']['Input Status'] != 'Normal':
            data['Standby']['Result'] = 'fail'
            fail_item['Standby'].add('Input Status')      
        if data['Standby']['Lock Status'] == 'Unlock':
            data['Standby']['Result'] = 'fail'
            fail_item['Standby'].add('Lock Status')   
        if data['Standby']['Input Sticky'] != 'Normal':
            data['Standby']['Result'] = 'fail'
            fail_item['Standby'].add('Input Sticky')   
        if data['Standby']['Lock Sticky'] == 'Unlock':
            data['Standby']['Result'] = 'fail'
            fail_item['Standby'].add('Lock Sticky')   
        if data['Standby']['Result'] != 'ok':
            fail_item['Standby'].add('Result')
            test_summary['DPLL Status']["fail"] += 1
        else:
            test_summary['DPLL Status']["pass"] += 1
    return data, fail_item
def Long_Time_Bert_parse(info,test_case,case_translation_table1):
    data = {}
    result = 'pass'
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    if test_case == 'Qe1LineDiagControlEntry':
        data['QE1 Bert'] = {}
        for x,(key,val) in enumerate(info.items()):
            if 'Qe1PRBSStatus' in key:
                data['QE1 Bert']['PRBS Status'] = val
            elif 'Qe1PRBSErrorCounts' in key:
                data['QE1 Bert']['Error Counts'] = int(val)
            elif 'Qe1PRBSErrorSeconds' in key:
                data['QE1 Bert']['Error Seconds'] = int(val)
        data['QE1 Bert']['Result'] = 'pass'
        if data['QE1 Bert']['Error Counts'] != 0:
            data['QE1 Bert']['Result'] = 'fail'
            fail_item['QE1 Bert'].add('Error Counts')      
        if data['QE1 Bert']['Error Seconds'] != 0:
            data['QE1 Bert']['Result'] = 'fail'
            fail_item['QE1 Bert'].add('Error Seconds')   
        if data['QE1 Bert']['Result'] != 'pass':
            fail_item['QE1 Bert'].add('Result')
            test_summary['Long Time Bert']["fail"] += 1
        else:
            test_summary['Long Time Bert']["pass"] += 1
    return data, fail_item
def HDLC_Status_parse(info,test_case,case_translation_table1):
    data = {}
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    result = 'ok'
    if test_case == 'HdlcPriStatusTable':
        data['Primary'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Primary']['DUT'] = int(val)
            elif 'PriFpgaRxFCSerr' in key:
                data['Primary']['FpgaRxFCSerr'] = int(val)
        data['Primary']['Result'] = 'ok'
        if data['Primary']['FpgaRxFCSerr'] != 0:
            data['Primary']['Result'] = 'fail'
            fail_item['Primary'].add('FpgaRxFCSerr')
        if data['Primary']['Result'] != 'ok':
            fail_item['Primary'].add('Result')
            test_summary['HDLC Status']["fail"] += 1
        else:
            test_summary['HDLC Status']["pass"] += 1
    if test_case == 'HdlcSecStatusTable':
        data['Secondary'] = {}
        for x,(key,val) in enumerate(info.items()):
            if key == 'Dut':
                data['Secondary']['DUT'] = int(val)
            elif 'SecFpgaRxFCSerr' in key:
                data['Secondary']['FpgaRxFCSerr'] = int(val)
        data['Secondary']['Result'] = 'ok'
        if data['Secondary']['FpgaRxFCSerr'] != 0:
            data['Secondary']['Result'] = 'fail'
            fail_item['Secondary'].add('FpgaRxFCSerr')
        if data['Secondary']['Result'] != 'ok':
            fail_item['Secondary'].add('Result')
            test_summary['HDLC Status']["fail"] += 1
        else:
            test_summary['HDLC Status']["pass"] += 1
    return data, fail_item
def TSI_in_out_parse(info,test_case,case_translation_table1):
    data = {}
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    result = 'ok'
    if test_case == 'TsiDataCasInOutTable':
        data['TSI in out data'] = {}
        for x,(key,val) in enumerate(info.items()):
            if 'TS' in key:
                data['TSI in out data']['TS'] = int(val)
            elif "set outData" in key:
                data['TSI in out data']['Set OutData'] = val
            elif 'TsiDataCasInOutget' in key:
                data['TSI in out data']['Get InData'] = val[val.find("i_val:") + len("i_val:")+2:val.find("i_val:") + len("i_val:")+4]
            elif 'Result' in key:
                data['TSI in out data']['Result'] = val
        # if data['TSI in out data']['Set OutData'][0] != data['TSI in out data']['Get InData'][0] and \
        #  data['TSI in out data']['Set OutData'][1] != data['TSI in out data']['Get InData'][1] and \
        #  data['TSI in out data']['Get InData'] == 'FF':
        #     data['TSI in out data']['Result'] = 'fail'
        #     fail_item['TSI in out data'].add('Set OutData')
        #     fail_item['TSI in out data'].add('Get InData')
        if data['TSI in out data']['Result'] == 'fail':
            fail_item['TSI in out data'].add('Result')
            test_summary['TSI in out']["fail"] += 1
        elif data['TSI in out data']['Result'] == 'fail,retry':
            pass
        else:
            test_summary['TSI in out']["pass"] += 1
    return data, fail_item
def PWACR_Status_parse(info,test_case,case_translation_table1):
    data = {}
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    result = 'ok'
    data['PWACR Status'] = {}
    for x,(key,val) in enumerate(info.items()):
        if key == 'Dut':
            data['PWACR Status']['DUT'] = int(val)
        elif 'PwACRIndex' in key:
            data['PWACR Status']['Index'] = int(val)
        elif 'PwACRStatusIndex' in key:
            data['PWACR Status']['Status Index'] = val
        elif 'PwACRStatusPortID' in key:
            data['PWACR Status']['Status Port ID'] = int(val)
        elif 'PwACRStatusEnable' in key:
            data['PWACR Status']['Status Enable'] = val
        elif 'PwACRStatusStatus' in key:
            data['PWACR Status']['Status'] = val
        elif 'PwACRDisconCnt' in key:
            data['PWACR Status']['Discon Cnt'] = int(val)
        elif 'PwACRRestartCnt' in key:
            data['PWACR Status']['Restart Cnt'] = int(val)
        elif 'PwACRCovariance' in key:
            data['PWACR Status']['Covariance'] = float(val)
        elif 'PwACRFreq' in key:
            data['PWACR Status']['Freq'] = float(val)
    data['PWACR Status']['Result'] = 'ok'
    if data['PWACR Status']['Status'] == 'acquisition':
        data['PWACR Status']['Result'] = 'fail'
        fail_item['PWACR Status'].add('Status')
    if data['PWACR Status']['Result'] != 'ok':
        fail_item['PWACR Status'].add('Result')
        test_summary['PWACR Status']["fail"] += 1
    else:
        test_summary['PWACR Status']["pass"] += 1
    return data, fail_item
def CTRL_Bert_parse(info,test_case,bert_case,case_translation_table1):
    data = {}
    result = 'pass'
    fail_item = generate_fail_item_set(bert_case,case_translation_table1)
    if test_case == 'CtlBertTable':
        data[bert_case] = {}
        for x,(key,val) in enumerate(info.items()):
            if 'CtlBertSlot' in key:
                data[bert_case]['Slot'] = int(val.split('_')[1])
            elif 'CtlBertPort' in key:
                data[bert_case]['Port'] = int(val.split('_')[1])
            elif 'CtlBertStatus' in key:
                data[bert_case]['Status'] = val
            elif 'CtlBertElapsedTime' in key:
                data[bert_case]['Elapsed Time'] = int(val)
            elif 'CtlBertBitError' in key:
                data[bert_case]['Bit Error'] = int(val)
            elif 'CtlBertErrorSecond' in key:
                data[bert_case]['Error Second'] = int(val)
            elif 'CtlBertUnsyncSecond' in key:
                data[bert_case]['Unsync Second'] = int(val)
        data[bert_case]['Result'] = 'pass'
        if data[bert_case]['Bit Error'] != 0:
            data[bert_case]['Result'] = 'fail'
            fail_item[bert_case].add('Bit Error')      
        if data[bert_case]['Error Second'] != 0:
            data[bert_case]['Result'] = 'fail'
            fail_item[bert_case].add('Error Seconds')   
        if data[bert_case]['Status'] != 'Sync':
            data[bert_case]['Result'] = 'fail'
            fail_item[bert_case].add('Status')   
        if data[bert_case]['Result'] != 'pass':
            fail_item[bert_case].add('Result')
            test_summary[bert_case]["fail"] += 1
        else:
            test_summary[bert_case]["pass"] += 1
    return data, fail_item
def SSM_Clock_Source_parse(info,test_case,case_translation_table1):
    data = {}
    result = 'pass'
    fail_item = generate_fail_item_set(test_case,case_translation_table1)
    if test_case == 'CcbClockSourceEntry':
        data['SSM Clock Source'] = {}
        for x,(key,val) in enumerate(info.items()):
            if 'Dut' in key:
                data['SSM Clock Source']['DUT'] = val
            elif 'CcbClockIndex' in key:
                data['SSM Clock Source']['Clock Index'] = int(val)
            elif 'CcbClockSource' in key:
                data['SSM Clock Source']['Clock Source'] = val
            elif 'CcbClockPort' in key:
                data['SSM Clock Source']['Clock Port'] = val
            elif 'CcbClockState' in key:
                pattern = r'RxSSM(\d+)'
                result_val = hex(int(re.search(pattern, val).group(1), 2))[2:].upper()
                data['SSM Clock Source']['Clock State'] = result_val
    elif test_case == 'CcbTimingSource':
        data['SSM Timing Source'] = {}
        for x,(key,val) in enumerate(info.items()):
            if 'Dut' in key:
                data['SSM Timing Source']['DUT'] = int(val)
            elif 'CcbClockMode' in key:
                data['SSM Timing Source']['Clock Mode'] = val
            elif 'CcbCurrentClock' in key:
                data['SSM Timing Source']['Current Clock'] = val
            elif 'Result' in key:
                data['SSM Timing Source']['Result'] = val
        if 'Result' not in data['SSM Timing Source']:
            data['SSM Timing Source']['Result'] = ''
        if data['SSM Timing Source']['Result'] == 'Fail':
            fail_item['SSM Timing Source'].add('Result')
            fail_item['SSM Timing Source'].add('Current Clock')
            fail_item['SSM Timing Source'].add('Clock State')
            test_summary['SSM Clock Source']["fail"] += 1
        elif data['SSM Timing Source']['Result'] == 'Pass':
            test_summary['SSM Clock Source']["pass"] += 1
    return data, fail_item
# === parser function === 
def count_pw_test_summary(test_summary, working_pw_array,standby_pw_array):
    found_fail = False
    for working_status, standby_status in itertools.zip_longest(working_pw_array, standby_pw_array):
        if working_status is not None and working_status['result'] == 'fail':
            if working_status['fail_reason'] == 'L_bit':
                test_summary["PW Status(L-bit error)"]['fail'] += 1
            elif working_status['fail_reason'] == 'jitter_max':
                test_summary["PW Status(Jitter Max)"]['fail'] += 1
            elif working_status['fail_reason'] == 'DP+-1':
                test_summary["PW Status(DP +- 1)"]['fail'] += 1
            elif working_status['fail_reason'] == 'DP+-2':
                test_summary["PW Status(DP +- 2)"]['fail'] += 1
            elif working_status['fail_reason'] == 'link down':
                test_summary["PW Status(Link down)"]['fail'] += 1
            test_summary["PW Status"]['fail'] += 1
            found_fail = True
        if standby_status is not None and standby_status['result'] == 'fail':
            if standby_status['fail_reason'] == 'L_bit':
                test_summary["PW Status(L-bit error)"]['fail'] += 1
            elif standby_status['fail_reason'] == 'jitter_max':
                test_summary["PW Status(Jitter Max)"]['fail'] += 1
            elif standby_status['fail_reason'] == 'DP+-1':
                test_summary["PW Status(DP +- 1)"]['fail'] += 1
            elif standby_status['fail_reason'] == 'DP+-2':
                test_summary["PW Status(DP +- 2)"]['fail'] += 1
            elif standby_status['fail_reason'] == 'link down':
                test_summary["PW Status(Link down)"]['fail'] += 1
            test_summary["PW Status"]['fail'] += 1
            found_fail = True
    if found_fail != True:
        test_summary["PW Status"]['pass'] += 1
def IP6704_aps_data_parse(time_stamp, text_line,f_read):
    data = {}
    data[time_stamp] = text_line

    time_stamp, cur_text_line = read_line_and_split_time(f_read)
    while "IP6704 APS" in cur_text_line:
        if cur_text_line != text_line:
            data[time_stamp] = cur_text_line
            text_line = cur_text_line
        time_stamp, cur_text_line = read_line_and_split_time(f_read)
    return data
def ctrl_bert_fail_data_parse(text_line, f_read):
    bit_error = ""
    unsync_sec = ""
    if "BitError" in text_line:
        bit_error = re.findall(
            r'\bBitError=[^ ]*', text_line)[0].split('=')[1]
    elif "UnsyncSecond" in text_line:
        unsync_sec = re.findall(
            r'\bUnsyncSecond=[^ ]*', text_line)[0].split('=')[1]
    error_id = re.findall(
        r'\berr_id=[^\n]*', text_line)[0].split('=')[1]
    time_stamp, text_line = read_line_and_split_time(f_read)
    if "ElapsedTime" in text_line:
        elapsed_time = re.split(
            r'=', text_line)[-1].strip()
        if bit_error != "":
            data = {"Time": time_stamp,
                    "Bit Error": bit_error,
                    "Error ID": error_id,
                    "Elapsed Time": elapsed_time
                    }
        elif unsync_sec != "":
            data = {
                "Time": time_stamp,
                "UnsyncSecond": unsync_sec,
                "Error ID": error_id,
                "Elapsed Time": elapsed_time
            }
    # print(json.dumps(data,indent=4))
    return data
def IP6704_50ms_fail_data_parse(time_stamp, text_line):
    # DUT 3	50ms pass :Bit error=46784, Curr value=18
    if "Bit error" in text_line:
        bit_error = re.findall(
            r'\bBit error=[^ ]*', text_line)[0].split('=')[1].strip(',')
    if "Curr value" in text_line:
        cur_val = re.findall(
            r'\bCurr value=[^ ]*', text_line)[0].split('=')[1].strip('\n')
    data = {
        "Time": time_stamp,
        "Bit error": bit_error,
        "Cur value": cur_val
    }
    return data
def IP6704_50ms_result_parse(time_stamp, text_line, f_read):
    result_str = time_stamp + " : " + "=======================================================\n"
    result_str += time_stamp + " : " + text_line
    text_line = f_read.readline()
    while 'End' not in text_line and text_line != '':
        result_str += text_line
        if "50ms result" in text_line:
            if 'Pass' in text_line:
                result = 'pass'
            else:
                result = 'fail'
        text_line = f_read.readline()
    result_str += text_line
    text_line = f_read.readline()
    result_str += text_line

    return result_str, result
# === file path
date = datetime.today().strftime('%m%d')
file_path = F"D:/autotest/G7800"
excel_path = Path(F'D:/autotest/G7800/summary_json_parser_{date}.xlsx')
# with open(F"{file_path}/test_result_modified.json",'r') as fr1:
#     test_result = json.load(fr1)
report_log = F"{file_path}/Report_log.txt"
dut_info = get_dut_info(report_log)
# === file path

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"
wb.create_sheet('DUTs')

s1 = wb['Summary']
s4 = wb['DUTs']

# ===== main =====
case_mapping_table = {
    'RS485 Status': {'first_case':'Rs485SlotStatistEntry', 'second_case': 'Rs485StatistTable'},
    'PW Status': {'first_case':'PwBundleStaticsEntry','second_case': 'PwStandbyStaticsEntry'},
    'PW Statist': {'first_case':'PwPriProtectStatistEntry','second_case': 'PwSecProtectStatistEntry'},
    'Card Port Status': {'first_case':'Qe1PortStatusEntry','second_case': 'FomStatusEntry'},
    'DPLL Status': {'first_case':'DpllPriStatusTable','second_case': 'DpllStandbyStatusTable'},
    'HDLC Status': {'first_case':'HdlcPriStatusTable','second_case': 'HdlcSecStatusTable'},
    'PWACR Status': {'first_case':'PwACRStatusEntry','second_case':''},
    'Long Time Bert': {'first_case':'Qe1LineDiagControlEntry','second_case':''},
    'CTRL Bert': {'first_case':'CtlBertTable','second_case':''},
    'TSI in out': {'first_case':'TsiDataCasInOutTable','second_case':''},
    'SSM Clock Source': {'first_case': 'CcbClockSourceEntry','second_case':"CcbTimingSource"}
}
# case_translation_table1 = {
#     'Rs485SlotStatistEntry': 'RS485 info by slot',
#     'Rs485StatistTable': 'RS485 info',
#     'PwBundleStaticsEntry': 'Working PW Status',
#     'PwStandbyStaticsEntry': 'Standby PW Status',
#     'PwPriProtectStatistEntry': 'Working PW Statist',
#     'PwSecProtectStatistEntry': 'Standby PW Statist',
#     'Qe1PortStatusEntry': 'QE1',
#     'FomStatusEntry': 'FOM',
#     'DpllPriStatusTable': 'Working',
#     'DpllStandbyStatusTable': 'Standby',
#     'Qe1LineDiagControlEntry': 'QE1 Bert',
#     'HdlcPriStatusTable': 'Primary',
#     'HdlcSecStatusTable': 'Secondary',
#     'PwACRStatusEntry': 'PWACR Status',
#     'CtlBertTable': 'CTRL Bert',
#     'TsiDataCasInOutTable': 'TSI in out data',
#     'CcbClockSourceEntry': 'SSM Clock Source',
#     'CcbTimingSource': 'SSM Timing Source'
# }
# test_category = {
#     'CTRL Bert': {'Warm Reset':{}, "Cold Reset(T1 Unframe)":{}, 'Cold Reset(others)':{}, 'Ctrl1Eth1DisEn':{}, 'Power Reset':{}, 'System Reset':{}, 'Check setup status':{}, 'Cold Reset':{}}, 
#     'PW' : {'PW Status':{}, 'PW Statist':{}}, 
#     'RS485': {'RS485 Status':{}}, 
#     'DPLL' : {'DPLL Status':{}}, 
#     'Card' : {'Card port status':{}}, 
#     'QE1 Bert' : {'Long Time Bert':{}},
#     'TSI in out' : {'Basic TSI in out(Signalling)':{},'Basic TSI in out(Data)':{},'Cross TSI in out(Signalling)':{},'Cross TSI in out(Data)':{},'TSI in out(AU_Law)':{}, 'TSI in out':{}}, 
#     'Voice' : {'Tone':{},'Ring':{},'Voice Bert':{},'PW status(Voice test)':{},'RS485 status(Voice test)':{}}, 
#     'Bundle Test': {'Bundle test(bert)':{}, 'Bundle test(PW status)':{}, 'Bundle test(RS485 status)':{}},
#     'Long Time': {'Long Time(bert)':{}, 'Long Time(PW status)':{}, 'Long Time(RS485 status)':{}},
#     'MIB Scan' : {'MIB Scan':{}}, 
#     'Firmware' : {'Firmware test':{}},
#     'Others': {'PWACR Status':{}, 'HDLC Status':{}},
#     'Clock': {'SSM Clock Source':{}}
#     }
case_translation_table, test_case_length, test_counter = generate_test_case(case_mapping_table)
# bert_case = ["Warm Reset",
#         "Cold Reset(T1 Unframe)",
#         "Cold Reset(others)",
#         "Cold Reset",
#         "portdisable",
#         "Ctrl1Eth1DisEn",
#         "Power Reset",
#         "System Reset",
#         "ptn eth port disable",
#         "Check setup status",
#         "config_download",
#         'Download cfg']
# test_case = ''
# detail_items = {}
# test_summary = {}
# cur_pass = {}
# cur_fail = {}
# cur_warning = {}
# row_index = {}
# fail_info = {}
# detail_start_col = 9

# for x,action_item in enumerate(test_result):
#     card_info = get_card_config_action(action_item)
#     test_case_count = test_counter.copy()
#     sec_test_case_count = test_counter.copy()
#     first_row_index = test_counter.copy()
#     bert_counter = 0
#     CTRL_bert_counter = 0
#     if action_item not in fail_info:
#         fail_info[action_item] = {}
#     for test_item in test_result[action_item].keys():
#         if "Qe1LineDiagControlEntry" in test_item:
#             bert_counter += 1
#         elif 'CtlBertTable' in test_item:
#             CTRL_bert_counter += 1
#     for x1,timecase_item in enumerate(test_result[action_item]):
#         time_item = timecase_item.split('_')[0]
#         # counttime_item = timecase_item.split('_')[1]
#         ini_test_case = timecase_item.split('_')[2]
#         if ini_test_case in case_translation_table:
#             test_case = case_translation_table[ini_test_case]
#         else:
#             test_case = ini_test_case
#         if test_case not in wb.sheetnames:
#             wb.create_sheet(test_case)
#             sheet = wb[test_case]
#             max_row = sheet.max_row
#             first_row = max_row
#             set_info_title_for_sheet(sheet,first_row)
#             detail_items[test_case] = {}
#             if test_case == 'CTRL Bert':
#                 for i in range(len(bert_case)):
#                     add_item_in_test_summary(test_summary, bert_case[i], test_case)
#             else:
#                 add_item_in_test_summary(test_summary, test_case, test_case)
#         sheet = wb[test_case]
#         index = sheet.max_row + 1
#         if ini_test_case in case_translation_table and (ini_test_case != 'Qe1LineDiagControlEntry' and ini_test_case != 'CtlBertTable'):
#             for trans_name, ini_name in case_mapping_table.items():
#                 if ini_test_case in ini_name.values():
#                     cur_test_case = trans_name
#                     break
#                 else:
#                     pass
#             if cur_test_case in case_mapping_table:
#                 mapping_case = case_mapping_table[cur_test_case]
#             if ini_test_case == mapping_case['second_case']:
#                 sec_test_case_count[cur_test_case] += 1
#             else:
#                 test_case_count[cur_test_case] += 1
#                 row_index[cur_test_case] = sheet.max_row + 1
#             if test_case_count[cur_test_case] == 1:
#                 first_row_index[cur_test_case] = row_index[cur_test_case] - 1
#             function_name = cur_test_case.replace(' ', '_') + "_parse"
#             data, fail_item = globals()[function_name](test_result[action_item][timecase_item],ini_test_case,case_translation_table1)
#             detail_items[test_case] = data
#             for num1,infor in enumerate(detail_items[test_case]):
#                 infor_key = {}
#                 infor_val = {}
#                 fail_col = set()
#                 for num,(key, val) in enumerate(detail_items[test_case][infor].items()):
#                     infor_key[num] = key
#                     infor_val[num] = val
#                     if infor == case_translation_table1.get(mapping_case['first_case']):
#                         sheet.cell(row_index[cur_test_case],detail_start_col).value = time_item
#                         if test_case_length[mapping_case['first_case']] == 0:
#                             set_title_for_sheet(infor,sheet,1,detail_start_col,detail_start_col,"5cb800")
#                             sheet.merge_cells(start_row=1,end_row=1,start_column=detail_start_col,end_column=detail_start_col+len(detail_items[test_case][infor]))
#                         set_title_for_sheet('Time',sheet,2,detail_start_col,detail_start_col,"5cb800")
#                         set_title_for_sheet(key,sheet,2,detail_start_col+num+1,detail_start_col+num+1,"5cb800")
#                         test_case_length[mapping_case['first_case']] = len(detail_items[test_case][infor]) + 1
#                         sheet.cell(row_index[cur_test_case], detail_start_col+num+1).value = val
#                         for fail_key, fail_name in enumerate(fail_item):
#                             if fail_name == infor:
#                                 for xnum,fail_item_name in enumerate(fail_item[fail_name]):
#                                     if key == fail_item_name:
#                                         sheet.cell(row_index[cur_test_case], detail_start_col+num+1).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                         fail_col.add(detail_start_col+num+1)
#                                         if cur_test_case == 'RS485 Status' and fail_item_name != 'FCS error':
#                                             sheet.cell(row_index[cur_test_case], detail_start_col+num+1).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
#                                     else:
#                                         pass
#                         # === create fail sheet
#                         if key == 'Result' and (val == 'fail' or val == 'Fail'):
#                             if ('F-{}'.format(cur_test_case)) not in wb.sheetnames:
#                                 wb.create_sheet('F-{}'.format(cur_test_case))
#                                 fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row
#                                 fail_sheet_first_row = fail_sheet_max_row
#                                 set_info_title_for_fail_sheet(wb['F-{}'.format(cur_test_case)],fail_sheet_first_row)
#                                 fail_first_info_count = False
#                             fail_sheet_detail_start_col = 5
#                             minus_col = detail_start_col - fail_sheet_detail_start_col
#                             set_title_for_sheet(infor,wb['F-{}'.format(cur_test_case)],1,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                             wb['F-{}'.format(cur_test_case)].merge_cells(start_row=1,end_row=1,start_column=fail_sheet_detail_start_col,end_column=fail_sheet_detail_start_col+len(detail_items[test_case][infor]))
#                             set_title_for_sheet('Time',wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                             fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row + 1
#                             for i in range(len(infor_key)):
#                                 set_title_for_sheet(infor_key[i],wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col+i+1,fail_sheet_detail_start_col+i+1,"5cb800")
#                                 wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row,fail_sheet_detail_start_col).value = time_item
#                                 wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, fail_sheet_detail_start_col+i+1).value = infor_val[i]
#                                 for col in fail_col:
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, col-minus_col).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                             for a in range(len(card_info)):
#                                 merge_row_title_for_sheet(card_info[a],wb['F-{}'.format(cur_test_case)],fail_sheet_max_row,1+a,fail_sheet_max_row,'FFFFFF')
#                             fail_first_info_count = True
#                         # === create fail sheet
#                     elif infor == case_translation_table1.get(mapping_case['second_case']):
#                         sec_info_row = first_row_index[cur_test_case] + sec_test_case_count[cur_test_case]
#                         sheet.cell(sec_info_row,detail_start_col+test_case_length[mapping_case['first_case']]).value = time_item
#                         if test_case_length[mapping_case['second_case']] == 0:
#                             set_title_for_sheet(infor,sheet,1,detail_start_col+test_case_length[mapping_case['first_case']],detail_start_col+test_case_length[mapping_case['first_case']],"5cb800")
#                             sheet.merge_cells(start_row=1,end_row=1,start_column=detail_start_col+test_case_length[mapping_case['first_case']],
#                                               end_column=detail_start_col+test_case_length[mapping_case['first_case']]+len(detail_items[test_case][infor]))
#                         set_title_for_sheet('Time',sheet,2,detail_start_col+test_case_length[mapping_case['first_case']],detail_start_col+test_case_length[mapping_case['first_case']],"5cb800")
#                         set_title_for_sheet(key,sheet,2,detail_start_col+test_case_length[mapping_case['first_case']]+num+1,detail_start_col+test_case_length[mapping_case['first_case']]+num+1,"5cb800")
#                         test_case_length[mapping_case['second_case']] = len(detail_items[test_case][infor])
#                         sheet.cell(sec_info_row, detail_start_col+test_case_length[mapping_case['first_case']]+num+1).value = val
#                         for fail_key, fail_name in enumerate(fail_item):
#                             if fail_name == infor:
#                                 for xnum,fail_item_name in enumerate(fail_item[fail_name]):
#                                     if key == fail_item_name:
#                                         sheet.cell(sec_info_row, detail_start_col+test_case_length[mapping_case['first_case']]+num+1).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                         fail_col.add(detail_start_col+test_case_length[mapping_case['first_case']]+num+1)
#                                         if cur_test_case == 'RS485 Status' and fail_item_name != 'FCS error':
#                                             sheet.cell(row_index[cur_test_case], detail_start_col+num+1).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
#                                     else:
#                                         pass
#                         # === create fail sheet
#                         if key == 'Result' and (val == 'fail' or val == 'Fail'):
#                             if ('F-{}'.format(cur_test_case)) not in wb.sheetnames:
#                                 wb.create_sheet('F-{}'.format(cur_test_case))
#                                 fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row
#                                 fail_sheet_first_row = fail_sheet_max_row
#                                 set_info_title_for_fail_sheet(wb['F-{}'.format(cur_test_case)],fail_sheet_first_row)
#                                 fail_first_info_count = False
#                             fail_sheet_detail_start_col = 5
#                             minus_col = detail_start_col - fail_sheet_detail_start_col
#                             if fail_first_info_count == False:
#                                 set_title_for_sheet(infor,wb['F-{}'.format(cur_test_case)],1,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                                 wb['F-{}'.format(cur_test_case)].merge_cells(start_row=1,end_row=1,start_column=fail_sheet_detail_start_col,end_column=fail_sheet_detail_start_col+len(detail_items[test_case][infor]))
#                                 set_title_for_sheet('Time',wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                                 fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row + 1
#                                 for i in range(len(infor_key)):
#                                     set_title_for_sheet(infor_key[i],wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col+i+1,fail_sheet_detail_start_col+i+1,"5cb800")
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row,fail_sheet_detail_start_col).value = time_item
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, fail_sheet_detail_start_col+i+1).value = infor_val[i]
#                                     for col in fail_col:
#                                         wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, fail_sheet_detail_start_col+col+1).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                 for a in range(len(card_info)):
#                                     merge_row_title_for_sheet(card_info[a],wb['F-{}'.format(cur_test_case)],fail_sheet_max_row,1+a,fail_sheet_max_row,'FFFFFF')
#                             elif fail_first_info_count == True:
#                                 set_title_for_sheet(infor,wb['F-{}'.format(cur_test_case)],1,fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']],fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']],"5cb800")
#                                 wb['F-{}'.format(cur_test_case)].merge_cells(start_row=1,end_row=1,start_column=fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']],end_column=fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']]+len(detail_items[test_case][infor]))
#                                 set_title_for_sheet('Time',wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']],fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']],"5cb800")
#                                 fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row + 1
#                                 for i in range(len(infor_key)):
#                                     set_title_for_sheet(infor_key[i],wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']]+i+1,fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']]+i+1,"5cb800")
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row,fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']]).value = time_item
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, fail_sheet_detail_start_col+test_case_length[mapping_case['first_case']]+i+1).value = infor_val[i]
#                                     for col in fail_col:
#                                         wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, col-minus_col).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                 for a in range(len(card_info)):
#                                     merge_row_title_for_sheet(card_info[a],wb['F-{}'.format(cur_test_case)],fail_sheet_max_row,1+a,fail_sheet_max_row,'FFFFFF')
#                         # === create fail sheet
#         elif ini_test_case == 'Qe1LineDiagControlEntry':
#             cur_test_case = 'Long Time Bert'
#             test_case_count[cur_test_case] += 1
#             if test_case_count[cur_test_case] == bert_counter:
#                 row_index[cur_test_case] = sheet.max_row + 1
#                 data, fail_item = Long_Time_Bert_parse(test_result[action_item][timecase_item],ini_test_case,case_translation_table1)
#                 detail_items[test_case] = data
#                 for num1,infor in enumerate(detail_items[test_case]):
#                     for num,(key, val) in enumerate(detail_items[test_case][infor].items()):
#                         if infor == 'QE1 Bert':
#                             sheet.cell(row_index[cur_test_case],detail_start_col).value = time_item
#                             set_title_for_sheet(infor,sheet,1,detail_start_col,detail_start_col,"5cb800")
#                             sheet.merge_cells(start_row=1,end_row=1,start_column=detail_start_col,end_column=detail_start_col+len(detail_items[test_case][infor]))
#                             set_title_for_sheet('Time',sheet,2,detail_start_col,detail_start_col,"5cb800")
#                             set_title_for_sheet(key,sheet,2,detail_start_col+num+1,detail_start_col+num+1,"5cb800")
#                             test_case_length['Qe1LineDiagControlEntry'] = len(detail_items[test_case][infor]) + 1
#                             sheet.cell(row_index[cur_test_case], detail_start_col+num+1).value = val
#                             for fail_key, fail_name in enumerate(fail_item):
#                                 if fail_name == infor:
#                                     for xnum,fail_item_name in enumerate(fail_item[fail_name]):
#                                         if key == fail_item_name:
#                                             sheet.cell(row_index[cur_test_case], detail_start_col+num+1).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                         else:
#                                             pass
#                             if key == 'Result':
#                                 if val == 'pass':
#                                     sheet.cell(row_index[cur_test_case], 5).value = int('1')
#                                     sheet.cell(row_index[cur_test_case], 6).value = int('0')
#                                 else:
#                                     sheet.cell(row_index[cur_test_case], 5).value = int('0')
#                                     sheet.cell(row_index[cur_test_case], 6).value = int('1')
#                                     sheet.cell(row_index[cur_test_case], 6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                             # === create fail sheet
#                             elif key == 'Result' and (val == 'fail' or val == 'Fail'):
#                                 if ('F-{}'.format(cur_test_case)) not in wb.sheetnames:
#                                     wb.create_sheet('F-{}'.format(cur_test_case))
#                                     fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row
#                                     fail_sheet_first_row = fail_sheet_max_row
#                                     set_info_title_for_fail_sheet(wb['F-{}'.format(cur_test_case)],fail_sheet_first_row)
#                                 fail_sheet_detail_start_col = 5
#                                 minus_col = detail_start_col - fail_sheet_detail_start_col
#                                 set_title_for_sheet(infor,wb['F-{}'.format(cur_test_case)],1,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                                 wb['F-{}'.format(cur_test_case)].merge_cells(start_row=1,end_row=1,start_column=fail_sheet_detail_start_col,end_column=fail_sheet_detail_start_col+len(detail_items[test_case][infor]))
#                                 set_title_for_sheet('Time',wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                                 fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row + 1
#                                 for i in range(len(infor_key)):
#                                     set_title_for_sheet(infor_key[i],wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col+i+1,fail_sheet_detail_start_col+i+1,"5cb800")
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row,fail_sheet_detail_start_col).value = time_item
#                                     wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, fail_sheet_detail_start_col+i+1).value = infor_val[i]
#                                     for col in fail_col:
#                                         wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, col-minus_col).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                 for a in range(len(card_info)):
#                                     merge_row_title_for_sheet(card_info[a],wb['F-{}'.format(cur_test_case)],fail_sheet_max_row,1+a,fail_sheet_max_row,'FFFFFF')
#                             # === create fail sheet
#                 for i in range(len(card_info)):
#                     merge_row_title_for_sheet(card_info[i],wb[cur_test_case],row_index[cur_test_case],1+i,row_index[cur_test_case],'FFFFFF')
#         elif ini_test_case == 'CtlBertTable':
#             cur_test_case = 'CTRL Bert'
#             test_case_count[cur_test_case] += 1
#             if test_case_count[cur_test_case] == CTRL_bert_counter or CTRL_bert_counter == 1000:
#                 row_index[cur_test_case] = sheet.max_row + 1
#                 action_type = card_info[2]
#                 if 'Before' in action_type or 'After' in action_type or 'Check setup status' in action_type:
#                     for i in range(len(bert_case)):
#                         if bert_case[i] in action_type:
#                             data, fail_item = CTRL_Bert_parse(test_result[action_item][timecase_item],ini_test_case,bert_case[i],case_translation_table1)
#                             detail_items[test_case] = data
#                             for num1,infor in enumerate(detail_items[test_case]):
#                                 for num,(key, val) in enumerate(detail_items[test_case][infor].items()):
#                                     sheet.cell(row_index[cur_test_case],detail_start_col).value = time_item
#                                     if test_case_length['CtlBertTable'] == 0:
#                                         set_title_for_sheet(infor,sheet,1,detail_start_col,detail_start_col,"5cb800")
#                                         sheet.merge_cells(start_row=1,end_row=1,start_column=detail_start_col,end_column=detail_start_col+len(detail_items[test_case][infor]))
#                                     set_title_for_sheet('Time',sheet,2,detail_start_col,detail_start_col,"5cb800")
#                                     set_title_for_sheet(key,sheet,2,detail_start_col+num+1,detail_start_col+num+1,"5cb800")
#                                     test_case_length['CtlBertTable'] = len(detail_items[test_case][infor])+1
#                                     sheet.cell(row_index[cur_test_case], detail_start_col+num+1).value = val
#                                     for fail_key, fail_name in enumerate(fail_item):
#                                         if fail_name == infor:
#                                             for xnum,fail_item_name in enumerate(fail_item[fail_name]):
#                                                 if key == fail_item_name:
#                                                     sheet.cell(row_index[cur_test_case], detail_start_col+num+1).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                                 else:
#                                                     pass
#                                     if key == 'Result':
#                                         if val == 'pass':
#                                             sheet.cell(row_index[cur_test_case], 5).value = int('1')
#                                             sheet.cell(row_index[cur_test_case], 6).value = int('0')
#                                         else:
#                                             sheet.cell(row_index[cur_test_case], 5).value = int('0')
#                                             sheet.cell(row_index[cur_test_case], 6).value = int('1')
#                                             sheet.cell(row_index[cur_test_case], 6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                     # === create fail sheet
#                                     elif key == 'Result' and (val == 'fail' or val == 'Fail'):
#                                         if ('F-{}'.format(cur_test_case)) not in wb.sheetnames:
#                                             wb.create_sheet('F-{}'.format(cur_test_case))
#                                             fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row
#                                             fail_sheet_first_row = fail_sheet_max_row
#                                             set_info_title_for_fail_sheet(wb['F-{}'.format(cur_test_case)],fail_sheet_first_row)
#                                         fail_sheet_detail_start_col = 5
#                                         minus_col = detail_start_col - fail_sheet_detail_start_col
#                                         set_title_for_sheet(infor,wb['F-{}'.format(cur_test_case)],1,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                                         wb['F-{}'.format(cur_test_case)].merge_cells(start_row=1,end_row=1,start_column=fail_sheet_detail_start_col,end_column=fail_sheet_detail_start_col+len(detail_items[test_case][infor]))
#                                         set_title_for_sheet('Time',wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col,fail_sheet_detail_start_col,"5cb800")
#                                         fail_sheet_max_row = wb['F-{}'.format(cur_test_case)].max_row + 1
#                                         for i in range(len(infor_key)):
#                                             set_title_for_sheet(infor_key[i],wb['F-{}'.format(cur_test_case)],2,fail_sheet_detail_start_col+i+1,fail_sheet_detail_start_col+i+1,"5cb800")
#                                             wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row,fail_sheet_detail_start_col).value = time_item
#                                             wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, fail_sheet_detail_start_col+i+1).value = infor_val[i]
#                                             for col in fail_col:
#                                                 wb['F-{}'.format(cur_test_case)].cell(fail_sheet_max_row, col-minus_col).fill = PatternFill(start_color="e53b31", fill_type = "solid")
#                                         for a in range(len(card_info)):
#                                             merge_row_title_for_sheet(card_info[a],wb['F-{}'.format(cur_test_case)],fail_sheet_max_row,1+a,fail_sheet_max_row,'FFFFFF')
#                                     # === create fail sheet
#                             for i in range(len(card_info)):
#                                 merge_row_title_for_sheet(card_info[i],wb[cur_test_case],row_index[cur_test_case],1+i,row_index[cur_test_case],'FFFFFF')

#     cur_pass[x] = {}
#     cur_fail[x] = {}
#     cur_warning[x] = {}
#     for case in test_case_count:
#         if case != 'Long Time Bert' and case != 'CTRL Bert':
#             pass_start_cloumn = 5
#             if test_case_count[case] == 0:
#                 if x == 0:
#                     cur_pass[x][case] = 0
#                     cur_fail[x][case] = 0
#                     cur_warning[x][case] = 0
#                 else:
#                     cur_pass[x][case] = cur_pass[x-1][case]
#                     cur_fail[x][case] = cur_fail[x-1][case]
#                     cur_warning[x][case] = cur_warning[x-1][case]
#             else:
#                 cur_pass[x][case] = test_summary[case]['pass']
#                 cur_fail[x][case] = test_summary[case]['fail']
#                 cur_warning[x][case] = test_summary[case]['warning']
#                 for i in range(len(card_info)):
#                     merge_row_title_for_sheet(card_info[i],wb[case],row_index[case]-test_case_count[case]+1,1+i,row_index[case],'FFFFFF')
#                 if x == 0:
#                     merge_row_title_for_sheet(cur_pass[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn,row_index[case],'FFFFFF')
#                     if cur_fail[x][case] != 0:
#                         merge_row_title_for_sheet(cur_fail[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+1,row_index[case],'e53b31')
#                         merge_row_title_for_sheet(cur_warning[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+2,row_index[case],'FFFFFF')
#                     elif cur_warning[x][case] != 0:
#                         merge_row_title_for_sheet(cur_warning[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+2,row_index[case],'f5bf1a')
#                         merge_row_title_for_sheet(cur_fail[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+1,row_index[case],'FFFFFF')
#                     else:
#                         merge_row_title_for_sheet(cur_fail[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+1,row_index[case],'FFFFFF')
#                         merge_row_title_for_sheet(cur_warning[x][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+2,row_index[case],'FFFFFF')
#                 else:
#                     merge_row_title_for_sheet(cur_pass[x][case]-cur_pass[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn,row_index[case],'FFFFFF')
#                     if cur_fail[x][case]-cur_fail[x-1][case] != 0:
#                         merge_row_title_for_sheet(cur_fail[x][case]-cur_fail[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+1,row_index[case],'e53b31')
#                         merge_row_title_for_sheet(cur_warning[x][case]-cur_warning[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+2,row_index[case],'FFFFFF')
#                     elif cur_warning[x][case]-cur_warning[x-1][case] != 0:
#                         merge_row_title_for_sheet(cur_warning[x][case]-cur_warning[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+2,row_index[case],'f5bf1a')
#                         merge_row_title_for_sheet(cur_fail[x][case]-cur_fail[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+1,row_index[case],'FFFFFF')
#                     else:
#                         merge_row_title_for_sheet(cur_fail[x][case]-cur_fail[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+1,row_index[case],'FFFFFF')
#                         merge_row_title_for_sheet(cur_warning[x][case]-cur_warning[x-1][case],wb[case],row_index[case]-test_case_count[case]+1,pass_start_cloumn+2,row_index[case],'FFFFFF')

# ===== main =====
max_row = s4.max_row
for x, item in enumerate(list(dut_info.keys()),start=1):
    row = 1
    start_column = x*2-1
    end_column = x*2
    set_title_for_sheet(F"DUT-{item} {dut_info[item]['Chassis']}",s4,row,start_column,end_column,"ffc000")
    s4.cell(row+1,start_column).value = F"Chassis"
    s4.cell(row+1,end_column).value = F"{dut_info[item]['Chassis']}"
    s4.cell(row+2,start_column).value = F"Slots"
    s4.cell(row+2,end_column).value = F"{dut_info[item]['Total slots']}"
    row += 3
    for slot_item in dut_info[item]['Slot Info']:
        for val in list(slot_item.keys()):
            if 'Name' == val:
                set_title_for_sheet(F"{slot_item[val]}",s4,row,start_column,end_column,"b8cce4")
            else:
                s4.cell(row,start_column).value = val
                s4.cell(row,end_column).value = slot_item[val]
            row+=1


summary_title = ['Category','Test Item','Pass','Fail','Warning']
max_row = s1.max_row
for x, item in enumerate(summary_title):
    set_title_for_sheet(item,s1,max_row,x+1,1,"5cb800")
    s1.cell(max_row,x+1).font= Font(name="Calibri")
max_row = s1.max_row + 1

pcnt = 0
fcnt = 0
# ===== every sheet hyperlink =====
sheet_names = wb.sheetnames
for sheet_name in sheet_names:
    current_sheet = wb[sheet_name]
    current_sheet.cell(row=1, column=1).hyperlink = f"{ntpath.basename(excel_path)}#'{wb.sheetnames[0]}'!A1"
# ===== every sheet hyperlink =====
# ===== adjust width and place in center =====
for sheet in wb.sheetnames:
    adjust_column_width(wb[sheet])
    add_border_in_excel(wb[sheet])
    if sheet != 'Summary':
        alignment = Alignment(horizontal='center', vertical='center')
        for row in wb[sheet].iter_rows(min_row=1, max_row=wb[sheet].max_row, min_col=1, max_col=wb[sheet].max_column):
            for cell in row:
                cell.alignment = alignment
        change_font_for_sheet(wb[sheet],font='Calibri')
        if sheet != 'DUTs' and not sheet.startswith("F-"):
            wb[sheet].freeze_panes = 'H3'
        elif sheet.startswith("F-"):
            wb[sheet].freeze_panes = 'E3'
    print('{:-<30}'.format(sheet),end="")
    print('Done!')
# ===== adjust width and place in center =====
# --- add DUT info in Summary sheet --- #
max_column = s1.max_column + 1

Summary_for_DUT_first_row = max_row
Summary_for_DUT_row = max_row + 1

set_title_for_sheet('Category',s1,Summary_for_DUT_row,1,1,"5cb800")
set_title_for_sheet('Item',s1,Summary_for_DUT_row,2,2,"5cb800")
merge_row_title_for_sheet('Info',s1,Summary_for_DUT_row+1,1,Summary_for_DUT_row+1,"FFFFFF")
merge_row_title_for_sheet('DUTs',s1,Summary_for_DUT_row+1,2,Summary_for_DUT_row+1,"FFFFFF")
s1.cell(row=Summary_for_DUT_row+1, column=2).hyperlink = f"{ntpath.basename(excel_path)}#'{wb.sheetnames[1]}'!A1"

thin = Side(border_style="thin", color="000000")
for i in range(Summary_for_DUT_first_row,Summary_for_DUT_row+2):
    for j in range(1,3):
        s1.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        s1.cell(i,j).font= Font(name="Calibri")
s1.cell(Summary_for_DUT_row+1,2).font= Font(name="Calibri",underline='single',color='0563C1')
# --- add DUT info in Summary sheet --- # 
print('{:-<30}'.format("Formatting"),end="")
print('Done!')

wb.save(F'{file_path}/summary_json_parser_{date}.xlsx')