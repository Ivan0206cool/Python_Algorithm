import openpyxl
from pathlib import Path
import ntpath
from openpyxl.styles import PatternFill, Alignment, Font
file_path = F"C:/Users/User/Desktop/VScode_Project/Python_Algorithm"
excel_path = Path(F'C:/Users/User/Desktop/VScode_Project/Python_Algorithm/summary.xlsx')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"
wb.create_sheet('Test1')
wb.create_sheet('Test2')
s1 = wb['Summary']
summary_title = ['Category','Test Item','Pass','Fail','Warning']  #first row


def set_title_for_sheet(name,sheet,row,col,end_col,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_col-col >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=row,start_column=col,end_column=end_col)
def set_hyperlink_for_excel_cell(sheet1,row,col,test_num):
    s1 = wb[sheet1]
    s1.cell(row=row, column=col).hyperlink = f"{ntpath.basename(excel_path)}#'{wb.sheetnames[test_num]}'!A1"
    s1.cell(row=row, column=col).font= Font(name="Calibri",underline='single',color='0563C1') #change font
def create_Summary_excel_sheet():
    max_row = s1.max_row        #first row
    for x, item in enumerate(summary_title):
        set_title_for_sheet(item,s1,max_row,x+1,1,"5cb800")
        s1.cell(max_row,x+1).font= Font(name="Calibri") #change font

    max_row = s1.max_row + 1     #second row


    set_hyperlink_for_excel_cell('Summary',1,1,1)
    set_hyperlink_for_excel_cell('Summary',1,5,2)
    wb.save(F'{file_path}/summary.xlsx')
























create_Summary_excel_sheet()
