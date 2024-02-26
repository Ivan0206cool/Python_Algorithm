import os
import re
import json
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from xlsx2html import xlsx2html
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side



f_read = open('Error_logccpb1125.txt','r',encoding="utf-8")

if not os.path.exists('./Summary'):
    os.makedirs('./Summary')
if not os.path.exists('./Summary/Report_1125.txt'):
    open('./Summary/Report_1125.txt', 'w').close()
f_write = open('./Summary/Report_1125.txt', 'a')

test_summary = {
    "tsi_in_out" : {
        "pass":0,
        "fail":0
    },
    "E1/T1 au_law": {
        "pass":0,
        "fail":0
    },
    "any_ds1" :{
        "pass":0,
        "fail":0
    },
    "warmreset":{
        "pass":0,
        "fail":0
    },
    "coldreset":{
        "pass":0,
        "fail":0
    },
    "Ctrl1Eth1DisEn":{
        "pass":0,
        "fail":0
    },
    "powereset":{
        "pass":0,
        "fail":0
    }
}
test_result = {
    "tsi_in_out" : {
        "Round":{}
    },
    "E1/T1 au_law": {},
    "any_ds1" :{
        "via PW pass": 0,
        "via PW fail":0,
        "via QT1 pass":0,
        "via QT1 fail":0
    }
}

result_50ms_test=[]
fail_cnt = []
result_cnt= {}
def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False

def sort_time_for_50ms_test(array):
    group = {
        "< 10 ms":0,
        "10-50 ms":0,
        "> 50ms" :0
    }
    array.sort()
    for time in array:
        if float(time) < 10 :
            group["< 10 ms"] +=1
        elif float(time) > 10 and float(time) < 50:
            print(time)
            group["10-50 ms"] +=1
        elif float(time) > 50:
            group["> 50ms"] +=1
    f_write.write(F"{json.dumps(group, indent=4)}")
    # print(group)


def set_title_for_sheet(name,sheet,row,col,len,color):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    sheet.merge_cells(start_row=row,end_row=row,start_column=col,end_column=len)

def add_border_in_excel(sheet):
    thin = Side(border_style="thin", color="000000")
    for i in range(2,sheet.max_row+1):
        for j in range(2,sheet.max_column+1):
            for merged_cell in sheet.merged_cells.ranges:
                if sheet.cell(i,j).coordinate in merged_cell:
                    sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if not sheet.cell(i,j).value == None:
                sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)

def append_img_to_html(html_page,img_path):
    html = open(html_page,encoding='utf-8').read()
    html = html.replace('</table>', F'</table>\n<img src="{img_path}" alt="no file">')
    with open(html_page,'w') as file:
        file.write(html)

def add_style_to_html(*html_pages):
    for item in html_pages:
        html = open(item,encoding='utf-8').read()
        html = html.replace('<body>', '<body style="display: inline-grid">')
        with open(item,'w') as file:
            file.write(html)

def export_to_html(workbook,filepath):
    # print(workbook.sheetnames)
    dirname = os.path.dirname(filepath)
    for name in workbook.sheetnames:
        xlsx2html(filepath=filepath, sheet=name, output=F'{dirname}/{name}.html')


def write_data_to_excel():
    tsi_title = ['Round','RX Name','Pass','Fail','Unknown']
    card_title = ['Frame type','Card type','Action', 'Pass', 'Fail']
    summary_title = ['Test Item','Pass','Fail']
    thin = Side(border_style="thin", color="000000")
    wb = openpyxl.Workbook()
    wb.create_sheet('Action',0)
    wb.create_sheet('Case',0)
    wb.create_sheet('Summary',0)


    s1 = wb['Summary']
    set_title_for_sheet("Summary",s1,2,2,len(summary_title)+1,"5cb800")
    max_row = s1.max_row +1

    for x, item in enumerate(summary_title):
        s1.cell(max_row,x+2).value = item
        s1.cell(max_row,x+2).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # s1.column_dimensions[s1.cell(max_row,x+2).column_letter].width = (len(str(s1.cell(max_row,x+2).value)) + 2) * 1.2


    max_row = s1.max_row + 1

    pcnt = 0
    fcnt = 0
    max_len = 0
    for x, item in enumerate(list(test_summary.keys())):
        # print(list(test_summary.keys()))
        s1.cell(max_row+x,2).value = item
        s1.cell(max_row+x,3).number_format = '0'
        s1.cell(max_row+x,3).value = test_summary[item]['pass']
        s1.cell(max_row+x,4).number_format = '0'
        s1.cell(max_row+x,4).value = test_summary[item]['fail']
        if len(str(s1.cell(max_row+x,2).value)) > max_len:
            max_len = len(str(s1.cell(max_row+x,2).value))

        pcnt += test_summary[item]['pass']
        fcnt += test_summary[item]['fail']
    s1.column_dimensions['B'].width = (max_len + 2) * 1.2
    data  = [pcnt,fcnt]
    label = 'Pass', 'Fail'
    plt.pie(data,labels=label, autopct='%1.1f%%')
    plt.title('Rate')
    plt.axis('equal')
    plt.savefig('./Summary/summary_pie.png',dpi=70)

    add_border_in_excel(s1)
    s2 = wb['Case']
    # max_row = s1.max_row + 2
    max_row = s2.max_row + 1
    for test_case in list(test_result.keys()):

        if test_case == 'tsi_in_out':
            set_title_for_sheet(test_case,s2,max_row,2,len(tsi_title)+1,"5cb800")
            s2.merge_cells(start_row=s2.max_row,end_row=s2.max_row,start_column=2,end_column=len(tsi_title)+1)
            for x, item in enumerate(tsi_title):
                s2.cell(max_row+1,x+2).value = item
            max_row = s2.max_row + 1
            for x, item in enumerate(list(test_result[test_case]['Round'].keys())):
                s2.cell(max_row+x,4).number_format = '0'
                s2.cell(max_row+x,5).number_format = '0'
                s2.cell(max_row+x,6).number_format = '0'
                s2.cell(max_row+x,2).value = int(item)
                s2.cell(max_row+x,3).value = test_result[test_case]['Round'][item]['RX Name']
                s2.cell(max_row+x,4).value = test_result[test_case]['Round'][item]['Pass']
                s2.cell(max_row+x,5).value = test_result[test_case]['Round'][item]['Fail']
                s2.cell(max_row+x,6).value = test_result[test_case]['Round'][item]['Unknown']
            max_row = s2.max_row + 2
        elif test_case == 'E1/T1 au_law' or test_case == 'any_ds1':
            set_title_for_sheet(test_case,s2,max_row,2,len(list(test_result[test_case].keys()))+1,"5cb800")
            for x, item in enumerate(list(test_result[test_case].keys())):
                s2.cell(max_row+1,x+2).value = item
                s2.cell(max_row+2,x+2).number_format = '0'
                s2.cell(max_row+2,x+2).value = test_result[test_case][item]

            max_row = s2.max_row + 2
    add_border_in_excel(s2)
    s3 = wb['Action']
    max_row = s3.max_row + 1
    set_title_for_sheet("Full",s3,max_row,2,len(card_title)+1,"5cb800")

    max_row = s3.max_row + 1

    for x, item in enumerate(card_title):
        s3.cell(max_row,x+2).value = item
    max_row = s3.max_row + 1

    max_len = 0
    for x, item in enumerate(list(result_cnt.keys())):
        val_array = item.split('-')
        s3.cell(max_row,2).value = val_array[0]
        s3.cell(max_row,3).value = val_array[1]
        s3.cell(max_row,4).value = val_array[2]
        s3.cell(max_row,5).number_format = '0'
        s3.cell(max_row,6).number_format = '0'
        s3.cell(max_row,5).value = result_cnt[item]['pass']
        s3.cell(max_row,6).value = result_cnt[item]['fail']
        if len(str(s3.cell(max_row,2).value)) > max_len:
            max_len = len(str(s3.cell(max_row,2).value))
        max_row = s3.max_row + 1
    s3.column_dimensions['B'].width = (max_len + 2) * 1.2
    add_border_in_excel(s3)
    wb.save('./Summary/summary.xlsx')
    export_to_html(wb,'./Summary/summary.xlsx')


    df = pd.read_excel("./Summary/summary.xlsx", sheet_name='Action')
    df = df.drop(axis =1,columns='Unnamed: 0')
    df = df.drop(index = df.index[[0,1]])
    df.columns = card_title

    plot = df.groupby(['Card type']).sum().plot(kind='pie',y='Fail', title="Card type Fail Rate", autopct="%1.1f%%")
    fig = plot.get_figure()
    fig.savefig('./Summary/action_pie.png')


    add_style_to_html('./Summary/Summary.html','./Summary/Action.html')
    append_img_to_html('./Summary/Summary.html','./summary_pie.png')

    img = openpyxl.drawing.image.Image('./Summary/summary_pie.png')
    img.anchor = 'F2'
    s1.add_image(img)

    img = openpyxl.drawing.image.Image('./Summary/action_pie.png')
    img.anchor = 'H2'
    s3.add_image(img)
    wb.save('./Summary/summary.xlsx')
def get_cfg_str_and_code(cfg_type):
    if cfg_type == 0:
        str_cfg_type = 'Unframe'
        str_cfg_type_code = '00'
    elif cfg_type == 1:
        str_cfg_type = 'Frame CRC CAS'
        str_cfg_type_code = '01'
    elif cfg_type == 2:
        str_cfg_type = 'Frame CRC'
        str_cfg_type_code = '02'
    elif cfg_type == 3:
        str_cfg_type = 'Frame'
        str_cfg_type_code = '03'
    elif cfg_type == 4:
        str_cfg_type = 'Frame CAS'
        str_cfg_type_code = '04'
    elif cfg_type == 5:
        str_cfg_type = 'Unframe192'
        str_cfg_type_code = '05'
    elif cfg_type == 6:
        str_cfg_type = 'Unframe193'
        str_cfg_type_code = '06'
    elif cfg_type == 7:
        str_cfg_type = 'D4 CAS'
        str_cfg_type_code = '07'
    elif cfg_type == 8:
        str_cfg_type = 'ESF CAS'
        str_cfg_type_code = '08'
    elif cfg_type == 9:
        str_cfg_type = 'D4'
        str_cfg_type_code = '00'
    elif cfg_type == 10:
        str_cfg_type = 'ESF'
        str_cfg_type_code = '00'
    else:
        str_cfg_type = 'no define'
        str_cfg_type_code = '99'
    return str_cfg_type, str_cfg_type_code

def get_card_type_str_and_code(card_type):
    if card_type == 'QT1':
        str_card_type_code = "00"
        str_card_type = "QT1"
    elif card_type == 'FT1':
        str_card_type_code = "01"
        str_card_type = "FT1"
    elif card_type == 'MQT1':
        str_card_type_code = "02"
        str_card_type = "MQT1"
    elif card_type == 'QE1':
        str_card_type_code = "10"
        str_card_type = "QE1"
    elif card_type == 'FE1':
        str_card_type_code = "11"
        str_card_type = "FE1"
    elif card_type == 'MQE1':
        str_card_type_code = "12"
        str_card_type = "MQE1"
    elif card_type == '4C37':
        str_card_type_code = "13"
        str_card_type = "4C37"
    elif card_type == '8CD':
        str_card_type_code = "14"
        str_card_type = "8CD"
    elif card_type == 'FOM':
        str_card_type_code = "15"
        str_card_type = "FOM"
    elif card_type == 'Voice':
        str_card_type_code = "20"
        str_card_type = "Voice"
    else:
        str_card_type_code = "99"
        str_card_type = "no define"
    return str_card_type,str_card_type_code

# f_write.write("Begin\n")
while True:
    round_num = 0
    text_line = f_read.readline()
    if text_line == '':
        break

    if "TEST Begin" in text_line:

        test_case = text_line.rsplit(' ',2)[0]
        f_write.write(F'{test_case}\n')
        if test_case == "tsi_in_out":
            pass_count = 0
            fail_count = 0
            unknown_count = 0
            rx_name = ''
            text_line = f_read.readline()
            while "TEST End" not in text_line:
                # print(text_line)
                if "Round" in text_line:
                    if round_num != 0 :
                        test_result["tsi_in_out"]["Round"][round_num] = {}
                        test_result["tsi_in_out"]["Round"][round_num]["RX Name"] = rx_name
                        test_result["tsi_in_out"]["Round"][round_num]["Pass"] =pass_count
                        test_result["tsi_in_out"]["Round"][round_num]["Fail"] =fail_count
                        test_result["tsi_in_out"]["Round"][round_num]["Unknown"] =unknown_count
                        test_summary[test_case]["pass"] +=pass_count
                        test_summary[test_case]["fail"] +=fail_count
                        f_write.write(F"-----Round {round_num}-------------------------------------------------\n")
                        f_write.write(F"{rx_name} Pass:{pass_count} Fail:{fail_count}\n")
                    pass_count = 0
                    fail_count = 0
                    unknown_count = 0
                    round_num = re.split(r'\-|:',text_line)[3]
                elif "TS" in text_line:
                    rx_name = re.split(r' |\.',text_line)[7]
                    if "pass." in text_line:
                        pass_count+=1
                    elif "fail" in text_line:
                        fail_count+=1
                    else:
                        unknown_count+=1
                    result = re.split(r' |\.',text_line)[9]

                text_line = f_read.readline()
            if round_num != 0 :
                test_result["tsi_in_out"]["Round"][round_num] = {}
                test_result["tsi_in_out"]["Round"][round_num]["RX Name"] = rx_name
                test_result["tsi_in_out"]["Round"][round_num]["Pass"] =pass_count
                test_result["tsi_in_out"]["Round"][round_num]["Fail"] =fail_count
                test_result["tsi_in_out"]["Round"][round_num]["Unknown"] =unknown_count
                test_summary[test_case]["pass"] +=pass_count
                test_summary[test_case]["fail"] +=fail_count
                f_write.write(F"-----Round {round_num}-------------------------------------------------\n")
                f_write.write(F"{rx_name} Pass:{pass_count} Fail:{fail_count}\n")
                f_write.write("-------------------------------------------------------------\n")
        elif test_case == "E1/T1 au_law":
            pass_count = 0
            fail_count = 0
            unknown_count = 0
            text_line = f_read.readline()
            while "TEST End" not in text_line:
                if "pass." in text_line:
                    pass_count+=1
                elif "fail" in text_line:
                    fail_count+=1
                else:
                    unknown_count+=1
                text_line = f_read.readline()
            test_result["E1/T1 au_law"]["Pass"] =pass_count
            test_result["E1/T1 au_law"]["Fail"] =fail_count
            test_result["E1/T1 au_law"]["Unknown"] =unknown_count
            f_write.write(F"E1 to T1 Pass: {pass_count} Fail: {fail_count}\n")
            f_write.write("-------------------------------------------------------------\n")
            test_summary["tsi_in_out"]["pass"] +=pass_count
            test_summary["tsi_in_out"]["fail"] +=fail_count
            test_summary[test_case]["pass"] +=pass_count
            test_summary[test_case]["fail"] +=fail_count
        elif test_case == "any_ds1":
            text_line = f_read.readline()
            while "TEST End" not in text_line:
                if "Begin" in text_line:
                    test_item = re.split(r'\/|\ ', text_line)[2]+re.split(r'\/|\ ', text_line)[3]

                    text_line = f_read.readline()
                    if "pass" not in text_line:
                        print("init check fail")
                elif "action" in text_line:
                    if "Round" in text_line:
                        round_num = re.split(':|\n', text_line)[1]
                        # print(round_num)
                    setup_pos = re.split('set | port', text_line)[1]
                    # print(setup_pos)
                    text_line = f_read.readline()
                    if "QT1" in setup_pos:
                        if "pass" in text_line:
                            test_result["any_ds1"]["via PW pass"]+=1
                        else:
                            test_result["any_ds1"]["via PW fail"]+=1
                    elif "eth" in setup_pos:
                        if "pass" in text_line:
                            test_result["any_ds1"]["via QT1 pass"]+=1
                        else:
                            test_result["any_ds1"]["via QT1 fail"]+=1
                text_line = f_read.readline()
                test_summary[test_case]["pass"] =test_result["any_ds1"]["via PW pass"]+test_result["any_ds1"]["via QT1 pass"]
                test_summary[test_case]["fail"] +=test_result["any_ds1"]["via PW fail"]+test_result["any_ds1"]["via QT1 fail"]
        # print(json.dumps(test_summary, indent=4))

    elif "-----Begin----" in text_line:
        if round_num != re.split(r"\ |-----",text_line)[6]:
            round_num = re.split(r"\ |-----",text_line)[6]
            # print(F'-----Round {round_num}-------------------------------------------------')
        # print(text_line)
        if '$cfg_type' in text_line:
            cfg_type = re.split(r'\$cfg_type\ ',text_line)[1].split(' ')[0]
            # print(type(cfg_type), cfg_type)
        str_cfg_type, str_cfg_type_code = get_cfg_str_and_code(int(cfg_type))
        # print(str_cfg_type,str_cfg_type_code)

        card_type = re.split(r'dut[\d+]_|_slot',text_line)[1]
        # print(text_line,card_type)
        str_card_type,str_card_type_code = get_card_type_str_and_code(card_type)
        # print(str_card_type,str_card_type_code)
    action_type = re.split(r'\ ',text_line)[-1].split('_cnt=')
	# print(len(action_type), action_type)
    if len(action_type) >= 2 :
        action_type = action_type[-2]
		# print(action_type)
        if action_type in test_summary:
            if str_cfg_type+'-'+str_card_type+'-'+action_type not in result_cnt:
                result_cnt[str_cfg_type+'-'+str_card_type+'-'+action_type] = {"pass":0,"fail":0}
				# print(json.dumps(result_cnt, indent=4))
                text_line = f_read.readline()
                if "Error" in text_line or "fail" in text_line:
                    result_cnt[str_cfg_type+'-'+str_card_type+'-'+action_type]["fail"]+=1
                    # print(text_line)
                    test_summary[action_type]["fail"]+=1
                else:
                    result_cnt[str_cfg_type+'-'+str_card_type+'-'+action_type]["pass"]+=1
                    test_summary[action_type]["pass"]+=1

# print(json.dumps(result_cnt, indent=4))

f_read.close
# print(json.dumps(test_result, indent=4))
write_data_to_excel()

