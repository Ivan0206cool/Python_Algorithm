import json
import itertools
import os
from openpyxl import load_workbook
from xlsx2html import xlsx2html
import re

def create_main_page(export_html_path="./Summary/Main page/Main.html"):
    export_dir = os.path.dirname(export_html_path)
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    if not os.path.exists(export_html_path):
        open(export_html_path, 'w').close()


    html = '\
        <!DOCTYPE html> \n\
            <html lang="en"> \n\
            <head>\n\
                <script src="./js/jquery-3.6.3.min.js"></script>\n\
                <script src="./js/action.js"></script>\n\
                <link rel="stylesheet" href="./css/main.css">\n\
                <meta charset="utf-8">\n\
                <meta name="viewport" content="width=device-width, initial-scale=1.0">\n\
                <title>Document</title>\n\
            </head>\n\
            <body>\n\
            </body>\n\
        </html>\n'
    ori_path = os.getcwd()
    data = get_json_data()
    os.chdir(export_dir)
    # print(os.getcwd())
    table = create_main_table(data)
    table += create_sub_table(data)
    html = html.replace('<body>', '<body>'+table)
    os.chdir(ori_path)
    with open(export_html_path,'w') as file:
        file.write(html)

def create_main_table(data):

    top_button= '\
        <div>\n\
            <button id="home" class="top_button" onClick="home_click()">Home</button>\n\
        </div>\n'

    all_table = '<table class = "center" id ="main">'
    all_table += "  <tr>\n"
    all_table += "    <td colspan='2' class='main_title device_color'>Device</td>\n"
    all_table += "  </tr>\n"
    for item in list(data['Device']):
        all_table += "  <tr>\n"
        all_table += F"    <td>{item}</td>\n"
        all_table += F"    <td>{data['Device'][item]}</td>\n"
        all_table += "  </tr>\n"
    all_table += "  <tr>\n"
    all_table += "    <td class='main_title display_color'>Display</td>\n"
    all_table += "    <td class='main_title setup_color'>Ctrl Regression</td>\n"
    all_table += "  </tr>\n"

    for x,y in list(itertools.zip_longest(list(data['Display']), list(data['Ctrl Regression']),fillvalue='')):
        all_table += "  <tr>\n"
        all_table += F"    <td><button class ='sub_button' id ='{'button_'+x.replace(' ','_')}' onClick='reply_click(this.id)'>{x}</button></td>\n"
        all_table += F"    <td><button class ='sub_button' id ='{'button_'+y.replace(' ','_')}' onClick='reply_click(this.id)'>{y}</button></td>\n"
        all_table += "  </tr>\n"

    all_table += "  <tr>\n"
    all_table += "    <td class='main_title card_regress_color'>Card Regression</td>\n"
    all_table += "  </tr>\n"

    for x in list(data['Card Regression']):
        all_table += "  <tr>\n"
        all_table += F"    <td><button class ='sub_button' id ='{'button_'+x.replace(' ','_')}' onClick='reply_click(this.id)'>{x}</button></td>\n"
        all_table += "  </tr>\n"
    all_table = top_button + all_table
    # print(all_table)
    return all_table

def get_json_data(json_path='./config/Data.json'):
    json_file = open(json_path)
    data = json.load(json_file)
    return data

def create_sub_table(data):
    sub_table = ''
    for category in data:
        for subitem in data[category]:
            if isinstance(data[category][subitem],list):
                # dict = {"category":category,"subitem":subitem}
                sub_table += add_table(subitem, data[category][subitem])
    return sub_table

def add_table(name,data):
    table = F"<table style='display:none' class = 'center' id ='{name.replace(' ','_')}'>\n"
    table += "  <tr>\n"
    table += F"    <td class='sub_title'>{name}</td>\n"
    table += "  </tr>\n"

    for item in data:
        # TODO: get correct html path
        page_src = F"../Test cases/{item}.html"
        table += "<tr>\n"
        # print(page_src)

        if os.path.exists(page_src):
            table += F"<td><button class ='item_button_clickable' id ='{'button_'+item.replace(' ','_')}'>{item}</button>\
                <iframe style='display:none' src='{page_src}' width='100%' height='700px'></iframe></td>\n"
        else:
            table += F"<td><button class ='item_button_unclick' id ='{'button_'+item.replace(' ','_')}'>{item} - No test Report</button>\
                </td>\n"
        table += "</tr>\n"
    table += '</table>\n'
    return table




def add_hyperlink_for_summary_page(html_page,des_name,des_html):
    html = open(html_page,encoding='utf-8').read()
    html = re.sub(rf'\<a href=".+{des_name}.+">',F'<a href="{des_html}">',html)
    # html = html.replace(des_name,F'<a href="{des_html}">{des_name}</a>')
    with open(html_page,'w') as file:
        file.write(html)

def make_sheet_to_html(excel_path,export_folder):
    # print(workbook.sheetnames)
    workbook = load_workbook(excel_path)
    if not os.path.exists(export_folder):
        os.makedirs(export_folder)
    for name in workbook.sheetnames:
        xlsx2html(filepath=excel_path, sheet=name, output=F'{export_folder}/{name}.html',locale='en')
        # Read the HTML file
        with open(F'{export_folder}/{name}.html', 'r', encoding='utf-8') as file:
            html_content = file.read()

        # Replace "\n" with "<br>" at the end of lines (excluding ">\n")
        modified_html_content = re.sub(r'(?<!>)\n', '<br>', html_content)

        # Save the modified content back to the original file
        with open(F'{export_folder}/{name}.html', 'w', encoding='utf-8') as file:
            file.write(modified_html_content)
    for name in workbook.sheetnames:
        if name != 'Summary':
            add_hyperlink_for_summary_page(F'{export_folder}/Summary.html',name,F'./{name}.html')


def append_img_to_html(html_page,img_path):
    html = open(html_page,encoding='utf-8').read()
    html = html.replace('</table>', F'</table>\n<img src="{img_path}" alt="no file">')
    with open(html_page,'w') as file:
        file.write(html)

def append_table_to_html(html_page, table_element):
    html = open(html_page,encoding='utf-8').read()
    html = html.replace('</table>', table_element)
    with open(html_page,'w') as file:
        file.write(html)

def add_style_to_html(html_dir):
    html_pages = os.listdir(html_dir)
    # print(html_pages)
    for item in html_pages:
        item = html_dir+'/'+item
        if os.path.exists(item):
            # print(item)
            title_name = os.path.splitext(item)[0].split('/')[-1]
            # print(title_name)
            html = open(item,encoding='utf-8').read()
            html = html.replace('Title', title_name)
            html = html.replace('</title>','</title> \
            <style> \
            table { \
            font-family: arial, sans-serif; \
            border-collapse: collapse; \
            width: 100%; \
            } \
            td, th { \
            border: 1px solid #dddddd; \
            text-align: left; \
            padding: 8px; \
            } \
            tr:nth-child(even) { \
            background-color: #dddddd; \
            }</style>')
            with open(item,'w') as file:
                file.write(html)

# def create_folder(data,export_dir_path):
#     # based on json data to create related folder
#     for category in data:
#         if category != 'Device':
#             for item in data[category]:
#                 os.makedirs(F"{export_dir_path}/{category}/{item}",exist_ok=True)
#                 for subitem in data[category][item]:
#                     fp = open(F"{export_dir_path}/{category}/{item}/{subitem}.html","w")
#                     fp.close
#                     # print(F"Create {export_dir_path}/{category}/{item}/{subitem}...", end="")
#                     # os.makedirs(F"{export_dir_path}/{category}/{item}/{subitem}",exist_ok=True)
#                     # print("Done")



def add_test_description_to_html(html_dir,img_dir):
    html_pages = os.listdir(html_dir)
    for page in html_pages:
        title_name = page.split('.')[0]
        img_path = img_dir+'/'+title_name+'.png'
        page = html_dir+'/'+ page
        if os.path.exists(img_path):
            img_path = img_path.replace('./Summary','..')
            # print(page,img_path,end='')
            append_img_to_html(page,img_path)

def export_to_html(excel_filepath,export_folder):
    make_sheet_to_html(excel_filepath,export_folder)
    # add_style_to_html(export_folder)