import pandas as pd
#from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles import DEFAULT_FONT
from ChartGen import *
import re
import ntpath
import json, itertools
import os
from pathlib import Path

def get_card_config_action(info):
    if "init" in info:
        card = ""
        config_type = ""
        card_action = info
        card_round = ""
    elif "after action" in info or "before action" in info or "Reget action" in info :
        if "after action" in info:
            action = "After"
        elif "before action" in info:
            action = "Before"
        elif "Reget action" in info:
            action = "Reget after"
        card_info = re.sub(r"(-record after action|-record before action|-Reget action)","", info)
        card = card_info.split('-')[1]
        config_type = card_info.split('-')[2]
        card_action = action + " " +card_info.split('-')[0]
        caround = card_info.split('-')[3]
        card_round = float(re.match(r"([a-zA-Z]+)(\d+)", caround).group(2))
    else:
        card = info.split('-')[1]
        config_type = info.split('-')[2]
        card_action = info.split('-')[0]
        caround = info.split('-')[3]
        card_round = float(re.match(r"([a-zA-Z]+)(\d+)", caround).group(2))
    return card,config_type,card_action,card_round

def merge_row_title_for_sheet(name,sheet,row,col,end_row,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center',vertical='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_row-row >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=end_row,start_column=col,end_column=col)

def read_line_and_split_time(f_read):
    full_line = f_read.readline()
    # print(full_line)
    if not full_line:
        return full_line, ''
    else:
        text_line = re.split(
            r'\d{4}\-\d{2}-\d{2} \d{2}:\d{2}:\d{2} : ', full_line)[1]
        time_stamp = re.findall(
            r'\d{4}\-\d{2}-\d{2} \d{2}:\d{2}:\d{2} : ', full_line)[0].strip(" : ")
        return time_stamp, text_line
    
def get_dut_info(input_report_path):
    f_read = open(input_report_path, 'r', encoding="utf-8")

    time_stamp, text_line = read_line_and_split_time(
        f_read)  # first line is 'begin'
    if 'begin' in text_line:
        time_stamp, text_line = read_line_and_split_time(f_read)
    dut_info = {}
    while "dut" in text_line.lower():
        if re.search(r'DUT[\d]{1} is', text_line):
            dut_num = text_line.split('is')[0].split('DUT')[1].strip()
            matches = re.findall(r'\b(LOOP|GE)([^\,]+)', text_line)
            for match in matches:
                dut_chassis = match[0] + ' ' + match[1].strip()
            print(dut_chassis)
            dut_slot_cnt = text_line.split(',slot#')[1].split(',')[
                0].split('\n')[0]
            dut_info[dut_num] = {
                "Chassis": dut_chassis,
                "Total slots": dut_slot_cnt,
                "Slot Info": []
            }
            # print(json.dumps(dut_info, indent=4))
        elif "dut:" in text_line.lower():
            dut_num = text_line.split(':')[1].split(' ')[0]
            dut_slot = text_line.split(':')[1].split(' ')[1]
            if "ctrl" in dut_slot:
                ctrl_hw_ver = text_line.split('"')[1].split(",")[0]
                ctrl_sw_ver = text_line.split('sw:"')[1].split('"')[0]
                ctrl_serial_ver = text_line.split('serial:')[1].split('\n')[0]
                if "ctrl" == dut_slot:  # ctrl1
                    ctrl_fpga1_ver = text_line.split('FPGA1:')[1].split(",")[0]
                    ctrl_fpga2_ver = text_line.split('FPGA2:')[1].split(",")[0]
                    ctrl_cpld_ver = text_line.split('CPLD:')[1].split(" ")[0]
                    dut_info[dut_num]["Slot Info"].append(
                        {
                            "Name": "Ctrl 1",
                            "HW Version": ctrl_hw_ver,
                            "SW Version": ctrl_sw_ver,
                            "Serial Version": ctrl_serial_ver,
                            "FPGA1 Version": ctrl_fpga1_ver,
                            "FPGA2 Version": ctrl_fpga2_ver,
                            "CPLD Version": ctrl_cpld_ver
                        })
                    # print(dut_cpld_ver,dut_slot)
                else:
                    dut_info[dut_num]["Slot Info"].append(
                        {
                            "Name": "Ctrl 2",
                            "HW Version": ctrl_hw_ver,
                            "SW Version": ctrl_sw_ver,
                            "Serial Version": ctrl_serial_ver
                        })
            elif "slot" in dut_slot:
                slot_info = re.findall(r'slot:(.*)\s+ver', text_line)
                # print(text_line,slot_info)
                slot_num = str(slot_info).split()[1].strip()
                slot_card_type = str(slot_info).split()[2].strip()
                # slot_card_type = re.split(r'\ ',text_line.split('slot: ')[1])[2]
                # print(slot_num, slot_card_type)
                slot_ver = text_line.split('ver:')[1].split('"')[1]
                # print(slot_ver)
                dut_info[dut_num]["Slot Info"].append(
                    {
                        "Name": F"Slot {slot_num}",
                        "Card Type": slot_card_type,
                        "Version": slot_ver
                    })
            #print(dut_slot)

        time_stamp, text_line = read_line_and_split_time(f_read)
    # print(list(dut_info.keys()))
    # for x , item in enumerate(list(dut_info.keys())):
    #     print(x,item,dut_info[item])
    f_read.close
    with open(F'{os.path.dirname(input_report_path)}/dut_info.json', 'w') as fw:
        fw.write(json.dumps(dut_info, indent=4))
    return dut_info

def change_font_for_sheet(sheet, font):
    # print('{:-<30}'.format("Change font"),end="")
    # DEFAULT_FONT.name = "Calibri"
    max_row = sheet.max_row
    max_column  = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            # for merged_cell in sheet.merged_cells.ranges:
            #     if sheet.cell(i,j).coordinate in merged_cell:
            #         sheet.cell(i,j).font = Font(name=font)
            sheet.cell(i,j).font = Font(name=font)
    # print('Done!')

def adjust_column_width(sheet):
    new_column_length = 0
    # print('{:-<30}'.format("Adjust width"),end="")
    for column_cells in sheet.columns:
        # for cell in column_cells:
        #     if str(cell.value).find('\n') != -1:
        #         new_column_length = str(cell.value).find('\n')
        # if new_column_length == 0 :
        break_line_count = max(str(cell.value).count('\n') for cell in column_cells)
        # print(break_line_count)

        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        if break_line_count > 0 :
            new_column_length /= break_line_count
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            sheet.column_dimensions[new_column_letter].width = new_column_length*1.23
    # print('Done!')

def add_border_in_excel(sheet):
    thin = Side(border_style="thin", color="000000")
    max_row = sheet.max_row
    max_column  = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            # for merged_cell in sheet.merged_cells.ranges:
            #     if sheet.cell(i,j).coordinate in merged_cell:
            #         sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_title_for_sheet(name,sheet,row,col,end_col,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_col-col >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=row,start_column=col,end_column=end_col)

def add_item_in_test_summary(test_summary, item, sheet_link_name):
    test_summary[item] = {
        "pass": 0,
        "fail": 0,
        "sheet link": sheet_link_name
    }

# === file path
#date = datetime.today().strftime('%m%d')
#log_path = Path('C:/Users/jeff-hsu/Desktop/Report_python_Jeff/error_logs')
#report_log_cisco = F"{log_path}/cisco/cisco{date}/Report_log.txt"
with open("C:/Users/jeff-hsu/Desktop/Newest JSON/sw-autoit-2020/Test_Case/case_344CCPA/5_GE_ccpb_2ge/test_results/0119/test_result_modified.json",'r') as fr1:
    test_result = json.load(fr1)
report_log = F"C:/Users/jeff-hsu/Desktop/json autoit/sw-autoit-2020/Test_Case/case_344CCPA/5_GE_ccpb_2ge/Report_log.txt"
dut_info = get_dut_info(report_log)
    #json_result2 = json.dumps(test_result, indent=2)
# === file path

#print(test_result)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"
wb.create_sheet('DUTs')

s1 = wb['Summary']
s4 = wb['DUTs']

# ===== main =====
test_case = ''
excel_path = Path("C:/Users/jeff-hsu/Desktop/test_result_ini/20240103/test_1.xlsx")
pw_detail_items = {}
test_summary = {}
for x,action_item in enumerate(test_result):
    card_info = get_card_config_action(action_item)
    for x1,timecase_item in enumerate(test_result[action_item]):
        time_item = timecase_item.split('_')[0]
        test_case = timecase_item.split('_')[1]
        add_item_in_test_summary(test_summary, test_case, test_case)
        if test_case not in wb.sheetnames:
            wb.create_sheet(test_case)
            sheet = wb[test_case]
            max_row = sheet.max_row
            first_row = max_row
            set_title_for_sheet('Card',sheet,first_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,first_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,first_row,3,3,"5cb800")
            set_title_for_sheet('Round',sheet,first_row,4,4,"5cb800")
            set_title_for_sheet('Time',sheet,first_row,5,5,"5cb800")
            pw_detail_items[test_case] = []
        sheet = wb[test_case]
        index = sheet.max_row + 1
        for i in range(len(card_info)):
            sheet.cell(index,1+i).value = card_info[i]

        max_row = sheet.max_row
        sheet.cell(index,5).value = time_item
        pwstatus_cnt = 0
        pwsecstatus_cnt = 0
        for x2,(key,val) in enumerate(test_result[action_item][timecase_item].items()):
            if '_' in key:
                pw_detail_items[test_case] += [key.split('_')[1]]
            else:
                pw_detail_items[test_case] += [key]
            if test_case == 'PwBundleStatusEntry':
                if 'PwStatus' in key and '_' in key:
                    pwstatus_title = (key.split('_')[1]).replace('PwStatus', "")
                    set_title_for_sheet(pwstatus_title,sheet,first_row,7+pwstatus_cnt,7+pwstatus_cnt,"5cb800")
                    if any(char.isalpha() for char in str(val)) == 1 or str(val).count('.') >= 2 or\
                        str(val) == '':
                        sheet.cell(index,7+pwstatus_cnt).value = val
                    else:
                        sheet.cell(index,7+pwstatus_cnt).value = float(val)
                    pwstatus_cnt += 1
                elif 'PwSec' in key:
                    if any(char.isalpha() for char in str(val)) == 1 or str(val).count('.') >= 2 or\
                        str(val) == '':
                        sheet.cell(index+1,7+pwsecstatus_cnt).value = val
                    else:
                        sheet.cell(index+1,7+pwsecstatus_cnt).value = float(val)
                    pwsecstatus_cnt += 1
                elif 'Dut' in key:
                    set_title_for_sheet('DUT',sheet,first_row,6,6,"5cb800")
                    merge_row_title_for_sheet(float(val),sheet,index,6,index+1,"FFFFFF")
                elif 'Pwstatus' in key:
                    set_title_for_sheet('ProtectionType',sheet,first_row,27,27,"5cb800")
                    sheet.cell(index,27).value = val
                for i in range(len(card_info)):
                    merge_row_title_for_sheet(card_info[i],sheet,index,i+1,index+1,"FFFFFF")
                merge_row_title_for_sheet(time_item,sheet,index,5,index+1,"FFFFFF")
            else:
                set_title_for_sheet(pw_detail_items[test_case][x2],sheet,first_row,6+x2,6+x2,"5cb800")            
                if any(char.isalpha() for char in str(val)) == 1 or str(val).count('.') >= 2 or\
                    str(val) == '':
                    sheet.cell(index,6+x2).value = val
                else:
                    sheet.cell(index,6+x2).value = float(val)
# ===== main =====
                
max_row = s4.max_row
for x, item in enumerate(list(dut_info.keys()),start=1):
    row = 1
    start_column = x*2-1
    end_column = x*2
    set_title_for_sheet(F"DUT-{item} {dut_info[item]['Chassis']}",s4,row,start_column,end_column,"ffc000")
    # s4.cell(row,s4.max_column+1).value = ""
    s4.cell(row+1,start_column).value = F"Chassis"
    s4.cell(row+1,end_column).value = F"{dut_info[item]['Chassis']}"
    s4.cell(row+2,start_column).value = F"Slots"
    s4.cell(row+2,end_column).value = F"{dut_info[item]['Total slots']}"
    row += 3
    # print(dut_info[item]['Slot Info'])
    for slot_item in dut_info[item]['Slot Info']:
        for val in list(slot_item.keys()):
            if 'Name' == val:
                set_title_for_sheet(F"{slot_item[val]}",s4,row,start_column,end_column,"b8cce4")
            else:
                s4.cell(row,start_column).value = val
                s4.cell(row,end_column).value = slot_item[val]
            row+=1
#copy_mr_sheet(wb,"D:\Jason-Huang\RPM20230511.xlsx")
            
summary_title = ['Test Item','Pass','Fail']
max_row = s1.max_row
for x, item in enumerate(summary_title):
    # s1.cell(max_row,x+1).value = item
    set_title_for_sheet(item,s1,max_row,x+1,1,"5cb800")
    s1.cell(max_row,x+1).font= Font(name="Calibri")
    # s1.cell(max_row,x+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
max_row = s1.max_row + 1

pcnt = 0
fcnt = 0
for x, item in enumerate(list(test_summary.keys())):
    # print(list(test_summary.keys()))
    if test_summary[item]['pass'] == 0 or test_summary[item]['fail'] ==0 or \
        item == 'MIB Scan' or item == 'Firmware test' or item == 'Alarm Trap check' or item =='MIB Diff':
        s1.cell(max_row,1).value = item
        # print(item)
        link = F"{ntpath.basename(excel_path)}#'{test_summary[item]['sheet link']}'!A1"
        # print(link)
        s1.cell(max_row,1).hyperlink = link
        s1.cell(max_row,1).font= Font(name="Calibri",underline='single',color='0563C1')
        # s1.cell(max_row,1).style= "Hyperlink"
        s1.cell(max_row,2).number_format = '0'
        s1.cell(max_row,2).value = test_summary[item]['pass']
        s1.cell(max_row,2).font= Font(name="Calibri")
        s1.cell(max_row,3).number_format = '0'
        s1.cell(max_row,3).value = test_summary[item]['fail']
        s1.cell(max_row,3).font= Font(name="Calibri")
# ===== hyperlink ===== #
        if s1.cell(max_row,3).value != 0:
            link_fail = F"{ntpath.basename(excel_path)}#'{'F-{}'.format(test_summary[item]['sheet link'])}'!A1"
            s1.cell(max_row,3).hyperlink = link_fail
            s1.cell(max_row,3).font= Font(name="Calibri",underline='single',color='0563C1')
# ===== hyperlink ===== #
        pcnt += test_summary[item]['pass']
        fcnt += test_summary[item]['fail']
        max_row += 1

# ===== every sheet hyperlink =====
sheet_names = wb.sheetnames
for sheet_name in sheet_names:
    current_sheet = wb[sheet_name]
    current_sheet.cell(row=1, column=1).hyperlink = f"{ntpath.basename(excel_path)}#'{wb.sheetnames[0]}'!A1"
# ===== every sheet hyperlink =====

# ===== adjust width and place in center =====
for sheet in wb.sheetnames:
    adjust_column_width(wb[sheet])
    add_border_in_excel(wb[sheet])
    #print(sheet)
    if sheet != 'Summary':
        alignment = Alignment(horizontal='center', vertical='center')
        for row in wb[sheet].iter_rows(min_row=1, max_row=wb[sheet].max_row, min_col=1, max_col=wb[sheet].max_column):
            for cell in row:
                cell.alignment = alignment
        change_font_for_sheet(wb[sheet],font='Calibri')
        if sheet != 'DUTs':
            wb[sheet].freeze_panes = 'F2'
    print('{:-<30}'.format(sheet),end="")
    print('Done!')
# ===== adjust width and place in center =====
print('{:-<30}'.format("Formatting"),end="")
print('Done!')

wb.save('C:/Users/jeff-hsu/Desktop/test_result_ini/20240103/test_1.xlsx')
