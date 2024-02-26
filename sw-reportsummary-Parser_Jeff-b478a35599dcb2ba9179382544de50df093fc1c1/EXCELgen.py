import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.styles import DEFAULT_FONT
from ChartGen import *
import re
import ntpath
import json, itertools
import os
from pathlib import Path

def collect_rs485_data(result,rs485_array):
    # print(result)
    if "detail" in result:
        for item in result["detail"]:
            for x, rs485_data in enumerate(rs485_array):
                if type(item['seq']) == list:
                    if rs485_data['seq'] in item['seq']:
                        rs485_data['date'] = item['date']
                        if 'oldVer' in item:
                            rs485_data['oldVer'] = item['oldVer']
                        if 'newVer' in item:
                            rs485_data['newVer'] = item['newVer']
                else:
                    if rs485_data['seq'] == item['seq']:
                        rs485_data['date'] = item['date']
                        if 'oldVer' in item:
                            rs485_data['oldVer'] = item['oldVer']
                        if 'newVer' in item:
                            rs485_data['newVer'] = item['newVer']

def write_data_to_excel(dut_info,test_summary,test_result,excel_path=Path('./summary_autoit.xlsx')):
    def append_fail_item_to_cell(error_item:list):
        append_data = ""
        for count, fail_item in enumerate(error_item,start=1):
            for key,val in fail_item.items():
                append_data += F"{key}: {val} "
            if count != len(error_item):
                append_data += '\n'
        return append_data
    summary_title = ['Category','Test Item','Pass','Fail','Warning']
    tsi_title_signalling = ['Card/PW','Check Type','Round','Set OutCAS','Get InCAS','Pass','Fail','Fail TS']
    tsi_title_data = ['Card/PW','Check Type','Round','Set OutData','Get InData','Pass','Fail','Fail TS']
    card_title = ['Card type','Frame type','Action', 'Pass', 'Fail', 'Pass Rate', 'Fail info']

    # --- test --- #
    time_title = ['Total time']
    voice_title = ['Card type','Frame type','Action','au-law', 'Impedance', 'Pass', 'Fail', 'Pass Rate', 'Fail info']
    cisco_voice_title = ['Round','Action','Stage','au-law', 'Impedance','Pass', 'Fail', 'Pass Rate', 'Fail info']
    # --- test --- #

    # title = re.findall(r'(ccpa|ccpb|g7800|cisco|8ge)',excel_path)[0] if len(re.findall(r'(ccpa|ccpb|g7800|cisco|8ge)',excel_path)) else 'autoit'

    with open(Path(F'{os.path.dirname(excel_path)}/test_result.json'),'w') as fw:
        fw.write(json.dumps(test_result,indent=4))
    with open(Path(F'{os.path.dirname(excel_path)}/test_summary.json'),'w') as fw:
        fw.write(json.dumps(test_summary,indent=4))

    # --- test --- #
    #with open(Path(F'{os.path.dirname(excel_path)}/test_result.json'),'r') as fr1:
    #    test_result = json.load(fr1)
    #with open(Path(F'{os.path.dirname(excel_path)}/test_summary.json'),'r') as fr2:
    #    test_summary = json.load(fr2)

    # --- test --- #

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    wb.create_sheet('DUTs')

    s1 = wb['Summary']
    s4 = wb['DUTs']


    for test_case in list(test_result.keys()):
        print('{:-<30}'.format(test_case),end="")
        if test_case == 'Bert' or test_case == 'Tone' or test_case == 'Ring' or test_case == '50ms' or test_case == 'Voice Bert':
            if test_result[test_case]['Total'] > 0:
                wb.create_sheet(test_case)
            else:
                print("Skip!")
                continue
        else:
            wb.create_sheet(test_case)
        sheet = wb[test_case]
        max_row = sheet.max_row
        if test_case == 'Basic TSI in out(Signalling)' or \
            test_case == 'Basic TSI in out(Data)' or \
            test_case == 'Cross TSI in out(Signalling)' or\
            test_case == 'Cross TSI in out(Data)':
            # set_title_for_sheet(test_case,sheet,max_row,1,len(tsi_title_signalling),"5cb800")
            # s2.merge_cells(start_row=s2.max_row,end_row=s2.max_row,start_column=1,end_column=len(tsi_title))
            if 'Signalling' in test_case:
                set_title = tsi_title_signalling
            else:
                set_title = tsi_title_data
            for x, tsi_title_item in enumerate(set_title,start=1):
                set_title_for_sheet(tsi_title_item,sheet,max_row,x,1,"5cb800")
            max_row = sheet.max_row + 1
            for pw_or_card in test_result[test_case]:
                for cfg_type in test_result[test_case][pw_or_card]:
                    for x, item in enumerate(list(test_result[test_case][pw_or_card][cfg_type]),start=max_row):
                        sheet.cell(x,3).number_format = '0'
                        sheet.cell(x,6).number_format = '0'
                        sheet.cell(x,7).number_format = '0'
                        if pw_or_card == 'PW':
                            sheet.cell(x,1).value = pw_or_card
                            sheet.cell(x,2).value = cfg_type
                        else:
                            words = cfg_type.split()
                            sheet.cell(x,1).value = words[0]
                            sheet.cell(x,2).value = words[1]
                        sheet.cell(x,3).value = int(item)
                        if 'Signalling' in test_case:
                            sheet.cell(x,4).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['OutCAS']
                            sheet.cell(x,5).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['InCAS']
                        elif 'Data' in test_case:
                            sheet.cell(x,4).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['OutData']
                            sheet.cell(x,5).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['InData']
                        sheet.cell(x,6).value = test_result[test_case][pw_or_card][cfg_type][item]['Pass']
                        sheet.cell(x,7).value = test_result[test_case][pw_or_card][cfg_type][item]['Fail']
                        fail_ts = ""
                        for time_slot_info in test_result[test_case][pw_or_card][cfg_type][item]["Item"]:
                            if time_slot_info['Status'] == 'Fail':
                                fail_ts += time_slot_info['TS'] + ','
                                # --- create fail sheet ---
                                if f'F-{test_case}' not in wb.sheetnames and test_case in ['Basic TSI in out(Signalling)', 
                                                                                              'Basic TSI in out(Data)',
                                                                                              'Cross TSI in out(Signalling)',
                                                                                              'Cross TSI in out(Data)']:
                                    wb.create_sheet('F-{}'.format(test_case),index=100)
                                    row_index = wb['F-{}'.format(test_case)].max_row+1
                                    set_title_for_sheet('Test case',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                                    set_title_for_sheet('Card/PW',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                                    set_title_for_sheet('Check Type',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                                    set_title_for_sheet('Round',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                                    set_title_for_sheet('Set OutData',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
                                    set_title_for_sheet('Get InData',wb['F-{}'.format(test_case)],1,6,6,"5cb800")
                                    set_title_for_sheet('Fail TS',wb['F-{}'.format(test_case)],1,7,7,"5cb800")
                                wb['F-{}'.format(test_case)].cell(row_index,1).value = test_case
                                if pw_or_card == 'PW':
                                    wb['F-{}'.format(test_case)].cell(row_index,2).value = pw_or_card
                                    wb['F-{}'.format(test_case)].cell(row_index,3).value = cfg_type
                                else:
                                    words = cfg_type.split()
                                    wb['F-{}'.format(test_case)].cell(row_index,2).value = words[0]
                                    wb['F-{}'.format(test_case)].cell(row_index,3).value = words[1]
                                wb['F-{}'.format(test_case)].cell(row_index,4).value = int(item)
                                if 'Signalling' in test_case:
                                    wb['F-{}'.format(test_case)].cell(row_index,5).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['OutCAS']
                                    wb['F-{}'.format(test_case)].cell(row_index,6).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['InCAS']
                                elif 'Data' in test_case:
                                    wb['F-{}'.format(test_case)].cell(row_index,5).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['OutData']
                                    wb['F-{}'.format(test_case)].cell(row_index,6).value = test_result[test_case][pw_or_card][cfg_type][item]["Item"][0]['InData']
                                wb['F-{}'.format(test_case)].cell(row_index,7).value = fail_ts[:-1]
                        if fail_ts != '':
                            fail_ts = fail_ts[:-1]
                            sheet.cell(x,8).value = fail_ts
                    if f'F-{test_case}' in wb.sheetnames:
                        row_index = wb['F-{}'.format(test_case)].max_row+1
                        # --- placed in center ---
                        alignment = Alignment(horizontal='center', vertical='center')
                        for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                            for cell in row:
                                cell.alignment = alignment
                        wb['F-{}'.format(test_case)].freeze_panes = 'A2'
                    max_row = sheet.max_row+1
            max_row = sheet.max_row + 2
            # --- placed in center ---
            alignment = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = alignment
            sheet.freeze_panes = 'A2'

        elif test_case == 'TSI in out(AU_Law)' :
            for x, item in enumerate(['Card','Covert','Set Data','Get Data','Result'],start=1):
                set_title_for_sheet(item,sheet,max_row,x,1,"5cb800")
            x = max_row + 1
            for card_key,card_dict in test_result[test_case].items():
                # print(card_key,card_dict)
                for covert_key,covert_dict in card_dict.items():
                    for item in covert_dict:
                        sheet.cell(x,1).value = card_key
                        sheet.cell(x,2).value = covert_key
                        sheet.cell(x,3).value = item['setData']
                        sheet.cell(x,4).value = item['getData']
                        sheet.cell(x,5).value = item['status']
                        x+=1
            max_row = sheet.max_row + 2
            sheet.freeze_panes = 'A2'
        elif test_case == "Any DS1" and len(test_result[test_case]) > 1:
            for item in test_result[test_case]:
                set_title_for_sheet(F"{item['Card']}-{item['Voice datapath']}",sheet,max_row,1,7,"5cb800")
                sheet.cell(max_row+1,2).value = "Init Pass"
                sheet.cell(max_row+1,3).value = "Init Fail"
                sheet.cell(max_row+1,4).value = "Card Disable Pass"
                sheet.cell(max_row+1,5).value = "Card Disable Fail"
                sheet.cell(max_row+1,6).value = "Eth Disable Pass"
                sheet.cell(max_row+1,7).value = "Eth Disable Fail"
                sheet.cell(max_row+2,1).value = "Tone"
                sheet.cell(max_row+3,1).value = "Bert"
                sheet.cell(max_row+4,1).value = "Ring"
                sheet.cell(max_row+2,2).value = item["Init Pass"]["Tone"]
                sheet.cell(max_row+3,2).value = item["Init Pass"]["Bert"]
                sheet.cell(max_row+4,2).value = item["Init Pass"]["Ring"]
                sheet.cell(max_row+2,3).value = item["Init Fail"]["Tone"]
                sheet.cell(max_row+3,3).value = item["Init Fail"]["Bert"]
                sheet.cell(max_row+4,3).value = item["Init Fail"]["Ring"]
                sheet.cell(max_row+2,4).value = item["Card Disable Pass"]["Tone"]
                sheet.cell(max_row+3,4).value = item["Card Disable Pass"]["Bert"]
                sheet.cell(max_row+4,4).value = item["Card Disable Pass"]["Ring"]
                sheet.cell(max_row+2,5).value = item["Card Disable Fail"]["Tone"]
                sheet.cell(max_row+3,5).value = item["Card Disable Fail"]["Bert"]
                sheet.cell(max_row+4,5).value = item["Card Disable Fail"]["Ring"]
                sheet.cell(max_row+2,6).value = item["Eth Disable Pass"]["Tone"]
                sheet.cell(max_row+3,6).value = item["Eth Disable Pass"]["Bert"]
                sheet.cell(max_row+4,6).value = item["Eth Disable Pass"]["Ring"]
                sheet.cell(max_row+2,7).value = item["Eth Disable Fail"]["Tone"]
                sheet.cell(max_row+3,7).value = item["Eth Disable Fail"]["Bert"]
                sheet.cell(max_row+4,7).value = item["Eth Disable Fail"]["Ring"]
                max_row = sheet.max_row+1
        elif test_case == "Tone Test":
            set_title_for_sheet(test_case,sheet,max_row,1,3,"5cb800")
            sheet.cell(max_row+1,1).value = 'Round'
            sheet.cell(max_row+1,2).value = 'Pass'
            sheet.cell(max_row+1,3).value = 'Fail'
            max_row = sheet.max_row + 1
            for x, item in enumerate(list(test_result[test_case]['Round'].keys())):
                sheet.cell(max_row+x,2).number_format = '0'
                sheet.cell(max_row+x,3).number_format = '0'
                sheet.cell(max_row+x,1).value = int(item)
                sheet.cell(max_row+x,2).value = test_result[test_case]['Round'][item]['pass']
                sheet.cell(max_row+x,3).value = test_result[test_case]['Round'][item]['fail']
            max_row = sheet.max_row + 2
        elif test_case == "MIB Scan":
            sheet.cell(max_row,1).value = 'MIB Version'
            sheet.cell(max_row,2).value = test_result[test_case]["MIB version"]
            sheet.cell(max_row+1,1).value = 'Registered cards'
            sheet.cell(max_row+1,2).value = test_result[test_case]["Registered cards"]
            set_title_for_sheet("In MIB only",sheet,max_row+2,1,1,"5cb800")
            set_title_for_sheet("In DUT only",sheet,max_row+2,2,1,"f79646")
            row_index = 4
            for x,y in itertools.zip_longest(
                test_result[test_case]["In MIB only result"],
                test_result[test_case]["In DUT only result"],
                fillvalue=''):
                sheet.cell(row_index,1).value = x
                sheet.cell(row_index,2).value = y
                row_index += 1
        elif test_case == "Firmware test":
            color_array = ["ffffff","d9d9d9"]
            slot_dict = test_result[test_case]['Slot']
            slot_dict["32"] = "CTRL"
            del test_result[test_case]['Slot']
            rs485_array = test_result[test_case]['Detail485']
            del test_result[test_case]['Detail485']
            color = ["5cb800","87B8EA"]
            highlight_color = "ffff00"
            G7800_cards = ['FXO', 'FXS', 'QE1', 'RS232X50', '4C37','EMA', 'RTB', 'MQE1','SQEMA']
            row_index = 1
            col_index = 1
            color_index = 0
            title_row_index = 0
            for ftp_mode, test_item in test_result[test_case]['IPV4'].items():
                if ftp_mode == 'TFTP' or ftp_mode == 'SFTP':
                    # print(ftp_mode)
                    # print(row_index,col_index)
                    set_title_for_sheet(F"IPV4-{ftp_mode}",sheet,row_index,col_index,col_index+1,color[color_index])
                    col_index += 2
                    title_row_index = row_index
                    row_index += 1
                    for connection, test_item in test_item.items():
                        if connection != 'VT100(COM)' and connection != 'WEB':
                            set_title_for_sheet(connection,sheet,title_row_index,col_index,col_index,color[color_index])
                            row_index = title_row_index + 1
                            for test_name,test_item in test_item.items():
                                # print(test_item)
                                if 'Pass' in test_item:
                                    if test_item['Pass'] > 0 or test_item['Fail']:
                                        sheet.cell(row_index,col_index).value = F"Pass:{test_item['Pass']}, Fail:{test_item['Fail']}, Total:{test_item['Pass']+test_item['Fail']}"
                                        if test_item['Fail'] > 0:
                                            sheet.cell(row_index,col_index).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        test_summary[test_case]['pass'] += test_item['Pass']
                                        test_summary[test_case]['fail'] += test_item['Fail']
                                    collect_rs485_data(test_item,rs485_array)
                                    if col_index == 3:
                                        set_title_for_sheet(test_name, sheet,row_index,col_index-2,col_index-1,color[color_index])
                                    row_index += 1
                                else:
                                    if col_index == 3:
                                        merge_row_title_for_sheet(test_name,sheet,row_index,col_index-2,row_index+len(list(test_item.keys()))-1,color[color_index])
                                    for x, card in enumerate(test_item.keys()):
                                        if col_index == 3:
                                            if card in G7800_cards:
                                                set_title_for_sheet(card,sheet,row_index+x,col_index-1,col_index-1,highlight_color)
                                            else:
                                                set_title_for_sheet(card,sheet,row_index+x,col_index-1,col_index-1,color[color_index])
                                        if test_item[card]['Pass'] > 0 or test_item[card]['Fail']:
                                            sheet.cell(row_index+x,col_index).value = F"Pass:{test_item[card]['Pass']}, Fail:{test_item[card]['Fail']}, Total:{test_item[card]['Pass']+test_item[card]['Fail']}"
                                            if test_item[card]['Fail'] > 0:
                                                sheet.cell(row_index+x,col_index).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            test_summary[test_case]['pass'] += test_item[card]['Pass']
                                            test_summary[test_case]['fail'] += test_item[card]['Fail']
                                        collect_rs485_data(test_item[card],rs485_array)
                                    row_index += len(list(test_item.keys()))
                                    # print(row_index)
                            col_index += 1
                    col_index = 1
            # print(sheet.max_column,sheet.max_row)
            color_index = 1
            row_index = 1
            col_index = sheet.max_column + 1
            for ftp_mode, test_item in test_result[test_case]['IPV6'].items():
                if ftp_mode == 'TFTP' or ftp_mode == 'SFTP':
                    # print(ftp_mode)
                    # print(row_index,col_index)
                    set_title_for_sheet(F"IPV6-{ftp_mode}",sheet,row_index,col_index,col_index+1,color[color_index])
                    col_index += 2
                    title_row_index = row_index
                    row_index += 1
                    for connection, test_item in test_item.items():
                        if connection != 'VT100(COM)' and connection != 'WEB':
                            set_title_for_sheet(connection,sheet,title_row_index,col_index,col_index,color[color_index])
                            row_index = title_row_index + 1
                            for test_name,test_item in test_item.items():
                                # print(test_item)
                                if 'Pass' in test_item:
                                    if test_item['Pass'] > 0 or test_item['Fail']:
                                        sheet.cell(row_index,col_index).value = F"Pass:{test_item['Pass']}, Fail:{test_item['Fail']}, Total:{test_item['Pass']+test_item['Fail']}"
                                        if test_item['Fail'] > 0:
                                            sheet.cell(row_index,col_index).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        test_summary[test_case]['pass'] += test_item['Pass']
                                        test_summary[test_case]['fail'] += test_item['Fail']
                                    collect_rs485_data(test_item,rs485_array)
                                    if col_index == 9:
                                        set_title_for_sheet(test_name, sheet,row_index,col_index-2,col_index-1,color[color_index])
                                    row_index += 1
                                else:
                                    if col_index == 9:
                                        merge_row_title_for_sheet(test_name,sheet,row_index,col_index-2,row_index+len(list(test_item.keys()))-1,color[color_index])
                                    for x, card in enumerate(test_item.keys()):
                                        if col_index == 9:
                                            if card in G7800_cards:
                                                set_title_for_sheet(card,sheet,row_index+x,col_index-1,col_index-1,highlight_color)
                                            else:
                                                set_title_for_sheet(card,sheet,row_index+x,col_index-1,col_index-1,color[color_index])
                                        if test_item[card]['Pass'] > 0 or test_item[card]['Fail']:
                                            sheet.cell(row_index+x,col_index).value = F"Pass:{test_item[card]['Pass']}, Fail:{test_item[card]['Fail']}, Total:{test_item[card]['Pass']+test_item[card]['Fail']}"
                                            if test_item[card]['Fail'] > 0:
                                                sheet.cell(row_index+x,col_index).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            test_summary[test_case]['pass'] += test_item[card]['Pass']
                                            test_summary[test_case]['fail'] += test_item[card]['Fail']
                                        collect_rs485_data(test_item[card],rs485_array)
                                    row_index += len(list(test_item.keys()))
                                    # print(row_index)
                            col_index += 1
                    col_index = 7
            # print(rs485_array)
            with open('./rs485.json','w') as f:
                json.dump(rs485_array, f)

            row_index = sheet.max_row + 2
            merge_row_title_for_sheet('Seq. ID',sheet,row_index,1,row_index+1,color[color_index])
            # merge_row_title_for_sheet('Slot',sheet,row_index,3,row_index+1,color[color_index])
            merge_row_title_for_sheet('Action',sheet,row_index,2,row_index+1,color[color_index])
            set_title_for_sheet('RS485',sheet,row_index,3,6,color[color_index])
            set_title_for_sheet('Slot-Card',sheet,row_index+1,3,1,color[color_index])
            set_title_for_sheet('No RX count',sheet,row_index+1,4,1,color[color_index])
            set_title_for_sheet('Timeout count',sheet,row_index+1,5,1,color[color_index])
            set_title_for_sheet('Non-alive count',sheet,row_index+1,6,1,color[color_index])
            merge_row_title_for_sheet('Old Ver.',sheet,row_index,7,row_index+1,color[color_index])
            merge_row_title_for_sheet('New Ver.',sheet,row_index,8,row_index+1,color[color_index])
            merge_row_title_for_sheet('Datetime',sheet,row_index,9,row_index+1,color[color_index])
            cur_row = sheet.max_row + 1
            for x,data in enumerate(rs485_array):
                merge_len = 0
                for slot_num,value in data['RS485'].items():
                    if F"{slot_num}" in slot_dict and slot_num != str(data['slot']) and value != -1:
                        merge_len+=1
                merge_len-=1
                merge_row_title_for_sheet(data['seq'],sheet,cur_row,1,cur_row+merge_len,color_array[x%2])


                # merge_row_title_for_sheet(data['slot'],sheet,cur_row,3,cur_row+merge_len,color_array[x%2])
                merge_row_title_for_sheet(data['action'],sheet,cur_row,2,cur_row+merge_len,color_array[x%2])

                index = 0
                for slot_num,value in data['RS485'].items():
                    if F"{slot_num}" in slot_dict and slot_num != str(data['slot']) and value != -1:
                        sheet.cell(cur_row+index,3).value = F"{slot_num}-{slot_dict[F'{slot_num}']}"
                        sheet.cell(cur_row+index,3).fill = PatternFill("solid", start_color=color_array[x%2])
                        sheet.cell(cur_row+index,4).value = value['norxcount']
                        if value['norxcount'] > 0 :
                            sheet.cell(cur_row+index,4).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        else:
                            sheet.cell(cur_row+index,4).fill = PatternFill("solid", start_color=color_array[x%2])
                        sheet.cell(cur_row+index,5).value = value['timeoutcount']
                        if value['timeoutcount'] > 0 :
                            sheet.cell(cur_row+index,5).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        else:
                            sheet.cell(cur_row+index,5).fill = PatternFill("solid", start_color=color_array[x%2])

                        sheet.cell(cur_row+index,6).value = value['nonalivecount']
                        if value['nonalivecount'] > 0 :
                            sheet.cell(cur_row+index,6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        else:
                            sheet.cell(cur_row+index,6).fill = PatternFill("solid", start_color=color_array[x%2])
                        index +=1
                if 'oldVer' in data:
                    merge_row_title_for_sheet(data['oldVer'],sheet,cur_row,7,cur_row+merge_len,color_array[x%2])
                    merge_row_title_for_sheet(data['newVer'],sheet,cur_row,8,cur_row+merge_len,color_array[x%2])
                    if data['oldVer'] == data['newVer']:
                        sheet.cell(cur_row,7).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        sheet.cell(cur_row,8).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                else:
                    merge_row_title_for_sheet('',sheet,cur_row,7,cur_row+merge_len,color_array[x%2])
                    merge_row_title_for_sheet('',sheet,cur_row,8,cur_row+merge_len,color_array[x%2])
                    sheet.cell(cur_row,7).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                    sheet.cell(cur_row,8).fill = PatternFill(start_color="e53b31", fill_type = "solid")

                if 'date' in data:
                    merge_row_title_for_sheet(data['date'],sheet,cur_row,9,cur_row+merge_len,color_array[x%2])
                else:
                    merge_row_title_for_sheet('',sheet,cur_row,9,cur_row+merge_len,color_array[x%2])
                cur_row += merge_len + 1
        elif test_case == "Alarm Trap check":
            set_title_for_sheet('Alarm Type',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Alarm',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Value',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Alarm Message Check',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Alarm Message',sheet,max_row,6,6,"5cb800")
            for x, item in enumerate(test_result[test_case].items(),start=2):
                # print(item)
                sheet.cell(x,1).value = item[1]['ccAlarmType']
                sheet.cell(x,2).value = item[0]
                sheet.cell(x,3).value = item[1]['value']
                sheet.cell(x,4).value = item[1]['action']
                if int(item[1]['value']) > 0:
                    sheet.cell(x,5).value = item[1]['alarmMessageCheck']
                    if 'error' in item[1]['alarmMessageCheck']:
                        sheet.cell(x,5).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        test_summary['Alarm Trap check']['fail'] += 1
                    else:
                        test_summary['Alarm Trap check']['pass'] += 1
                if 'alarmMessage' in item[1]:
                    sheet.cell(x,6).value = item[1]['alarmMessage']

            # print(json.dumps(flatten_data,indent=4))
        elif test_case == "MIB Diff":
            for x, item in enumerate(test_result[test_case],start=1):
                # print(item)
                set_title_for_sheet(F"Only in {item['file']}",sheet,1,x,x,"5cb800")
                for row ,oid_item in enumerate(item['diff_oid'],start=2):
                    sheet.cell(row,x).value = oid_item
            # set_title_for_sheet('Value',sheet,max_row,2,2,"5cb800")
        elif test_case == "DPLL Status":
            set_title_for_sheet('Card',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('DPLL Status',sheet,max_row,4,4,"5cb800")

            set_title_for_sheet('Time',sheet,max_row+1,4,4,"5cb800")
            set_title_for_sheet('DUT',sheet,max_row+1,5,5,"5cb800")
            set_title_for_sheet('Working/Standby',sheet,max_row+1,6,6,"5cb800")
            set_title_for_sheet('Input Status',sheet,max_row+1,7,7,"5cb800")
            set_title_for_sheet('Lock Status',sheet,max_row+1,8,8,"5cb800")
            set_title_for_sheet('Input Sticky',sheet,max_row+1,9,9,"5cb800")
            set_title_for_sheet('Lock Sticky',sheet,max_row+1,10,10,"5cb800")
            set_title_for_sheet('Result',sheet,max_row+1,11,11,"5cb800")
            # --- merge ---
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=4,
                              end_column=11)
            for i in range(1,4):
                sheet.merge_cells(start_row=1,end_row=2,start_column=i,
                              end_column=i)
            row_index = 3
            color_array = ["ffffff","d9d9d9"]
            for x, dpll_item in enumerate(test_result[test_case]):
                info = list(dpll_item.keys())[0]
                card, config_type,card_action = get_card_config_action(info)
                merge_row_title_for_sheet(card,sheet,row_index,1,row_index+len(dpll_item[info])-1,color_array[x%2])
                merge_row_title_for_sheet(config_type,sheet,row_index,2,row_index+len(dpll_item[info])-1,color_array[x%2])
                merge_row_title_for_sheet(card_action,sheet,row_index,3,row_index+len(dpll_item[info])-1,color_array[x%2])
                for index in range(row_index, row_index+len(dpll_item[info])):
                    sheet.cell(index,4).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                append_data = ""
                for item in dpll_item[info]:
                    #print(dpll_item[info])
                    for ind,(key,val) in enumerate(item.items()):
                        if ind == 0:
                            sheet.cell(row_index,4).value = key
                            dut_num = re.search(r'DUT (\d+)', F"{val}")
                            if dut_num:
                                sheet.cell(row_index,5).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                            sheet.cell(row_index,6).value = val.split()[-1]
                        else:
                            sheet.cell(row_index,6+ind).value = val
                        append_data += F"{key}: {val} "
                    if 'result: fail' in append_data:
                        #create_fail_info_sheet(wb,test_case,card,config_type,card_action,append_data)
                        # ----- create fail sheet ----- #
                        if f'F-{test_case}' not in wb.sheetnames and test_case in ['DPLL Status']:
                            wb.create_sheet('F-{}'.format(test_case),index=100)
                            max_row = wb['F-{}'.format(test_case)].max_row
                            set_title_for_sheet('Card',wb['F-{}'.format(test_case)],max_row,1,1,"5cb800")
                            set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],max_row,2,2,"5cb800")
                            set_title_for_sheet('Action',wb['F-{}'.format(test_case)],max_row,3,3,"5cb800")
                            set_title_for_sheet('DPLL Status',wb['F-{}'.format(test_case)],max_row,4,4,"5cb800")
                            set_title_for_sheet('Time',wb['F-{}'.format(test_case)],max_row+1,4,4,"5cb800")
                            set_title_for_sheet('DUT',wb['F-{}'.format(test_case)],max_row+1,5,5,"5cb800")
                            set_title_for_sheet('Working/Standby',wb['F-{}'.format(test_case)],max_row+1,6,6,"5cb800")
                            set_title_for_sheet('Input Status',wb['F-{}'.format(test_case)],max_row+1,7,7,"5cb800")
                            set_title_for_sheet('Lock Status',wb['F-{}'.format(test_case)],max_row+1,8,8,"5cb800")
                            set_title_for_sheet('Input Sticky',wb['F-{}'.format(test_case)],max_row+1,9,9,"5cb800")
                            set_title_for_sheet('Lock Sticky',wb['F-{}'.format(test_case)],max_row+1,10,10,"5cb800")
                            set_title_for_sheet('Result',wb['F-{}'.format(test_case)],max_row+1,11,11,"5cb800")
                            # --- merge ---
                            wb['F-{}'.format(test_case)].merge_cells(start_row=max_row,end_row=max_row,start_column=4,
                                            end_column=11)
                            for i in range(1,4):
                                wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,
                                            end_column=i)
                        max_row = wb['F-{}'.format(test_case)].max_row+1
                        merge_row_title_for_sheet(card,wb['F-{}'.format(test_case)],max_row,1,max_row,color_array[x%2])
                        merge_row_title_for_sheet(config_type,wb['F-{}'.format(test_case)],max_row,2,max_row,color_array[x%2])
                        merge_row_title_for_sheet(card_action,wb['F-{}'.format(test_case)],max_row,3,max_row,color_array[x%2])
                        for ind,(key,val) in enumerate(item.items()):
                            if ind == 0:
                                wb['F-{}'.format(test_case)].cell(max_row,4).value = key
                                dut_num = re.search(r'DUT (\d+)', F"{val}")
                                if dut_num:
                                    wb['F-{}'.format(test_case)].cell(max_row,5).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                wb['F-{}'.format(test_case)].cell(max_row,6).value = val.split()[-1]
                            else:
                                wb['F-{}'.format(test_case)].cell(max_row,6+ind).value = val
                        # --- placed in center ---
                        alignment = Alignment(horizontal='center', vertical='center')
                        for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                            for cell in row:
                                cell.alignment = alignment
                        wb['F-{}'.format(test_case)].freeze_panes = 'D3'
                    append_data = "" 
                    # --- color for gray ---
                    for index in range(row_index, sheet.max_row+1):
                        for index2 in range(4,sheet.max_column+1):
                            if "fail" in str(sheet.cell(index,index2).value):
                                sheet.cell(index,index2).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            if sheet.cell(index,index2).fill != PatternFill(start_color="e53b31", fill_type = "solid"):
                                sheet.cell(index,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                    row_index += 1
                # --- placed in center ---
                alignment = Alignment(horizontal='center', vertical='center')
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = alignment
                sheet.freeze_panes = 'D3'
        elif test_case == "RS485 Status":
            slot_second_name = ['DUT','Slot','TX count', "RX count", "No RX count", "Timeout count", 
                           "Alive count","Non-alive count","Oversize count", "Init count", "Bad CMD count",
                            "Bad CMD num","Result"]
            sta_second_name = ['DUT', 'FCS error', 'Slip error', 'Buffer error', 'Result']
            set_title_for_sheet('Card',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Pass',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Fail',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Warning',sheet,max_row,6,6,"5cb800")
            set_title_for_sheet('RS485 info by slot',sheet,max_row,7,7,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=7,end_column=7+len(slot_second_name)-1)
            set_title_for_sheet('RS485 info',sheet,max_row,20,20,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=20,end_column=20+len(sta_second_name)-1)
            for i in range(0,len(slot_second_name)):
                set_title_for_sheet(slot_second_name[i],sheet,max_row+1,7+i,7+i,"5cb800")
            for i in range(0,len(sta_second_name)):
                set_title_for_sheet(sta_second_name[i],sheet,max_row+1,20+i,20+i,"5cb800")
            # --- merge ---
            for i in range(1,7):
                sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
            color_array = ["ffffff","d9d9d9"]
            for x, RS485_item in enumerate(test_result[test_case]):
                index1 = sheet.max_row + 1
                info = list(RS485_item.keys())[0]
                card, config_type,card_action = get_card_config_action(info)
                long_row = len(RS485_item[info]) -1
                warning_cnt = 0
                fail_cnt = 0
                pass_cnt = 0
                minus_count = 0
                append_data = ""
                for count,item in enumerate(RS485_item[info], start = 1):
                    for ind,(key,val) in enumerate(item.items()):
                        if len(item) > 6:
                            if ind == 0:
                                info_data = F"{val}"
                                dut_num = re.search(r'DUT (\d+)', info_data)
                                slot_num = re.search(r'slot (\d+)', info_data)
                                slot_abc = re.search(r'slot ([a-zA-Z])', info_data)
                                if dut_num:
                                    sheet.cell(index1+count-minus_count-1,7).value = int(dut_num.group(1))
                                if slot_num:
                                    sheet.cell(index1+count-minus_count-1,8).value = int(slot_num.group(1))
                                elif slot_abc:
                                    sheet.cell(index1+count-minus_count-1,8).value = slot_abc.group(1)
                            elif ind == 1:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,9).value = float(info_data)
                            elif ind == 2:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,10).value = float(info_data) 
                            elif ind == 3:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,11).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+count-minus_count-1,11).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                            elif ind == 4:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,12).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+count-minus_count-1,12).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                            elif ind == 5:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,13).value = float(info_data)
                            elif ind == 6:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,14).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+count-minus_count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            elif ind == 7:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,15).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+count-minus_count-1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            elif ind == 8:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,16).value = float(info_data)
                            elif ind == 9:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,17).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+count-minus_count-1,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            elif ind == 10:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,18).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+count-minus_count-1,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            elif ind == 11:
                                info_data = F"{val}"
                                sheet.cell(index1+count-minus_count-1,19).value = info_data
                                if info_data == 'fail':
                                    sheet.cell(index1+count-minus_count-1,19).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    fail_cnt += 1
                                elif info_data == 'warning':
                                    sheet.cell(index1+count-minus_count-1,19).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    warning_cnt += 1
                                elif info_data == 'ok':
                                    pass_cnt += 1
                        else:
                            if ind == 0:
                                info_data = F"{val}"
                                dut_num = re.search(r'DUT (\d+)', info_data)
                                if dut_num:
                                    sheet.cell(index1+minus_count,20).value = int(dut_num.group(1))
                            elif ind == 1:
                                info_data = F"{val}"
                                sheet.cell(index1+minus_count,21).value = float(info_data)
                            elif ind == 2:
                                info_data = F"{val}"
                                sheet.cell(index1+minus_count,22).value = float(info_data) 
                            elif ind == 3:
                                info_data = F"{val}"
                                sheet.cell(index1+minus_count,23).value = float(info_data)
                                if float(info_data) != 0.0:
                                    sheet.cell(index1+minus_count,23).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            elif ind == 4:
                                info_data = F"{val}"
                                sheet.cell(index1+minus_count,24).value = info_data
                                if info_data == 'fail':
                                    sheet.cell(index1+minus_count,24).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    fail_cnt += 1
                                elif info_data == 'warning':
                                    sheet.cell(index1+count-minus_count-1,19).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    warning_cnt += 1
                                elif info_data == 'ok':
                                    pass_cnt += 1
                    if len(item) <= 6:
                        minus_count += 1    
                    append_data = ""
                merge_row_title_for_sheet(card,sheet,index1,1,index1+count-minus_count-1,color_array[x%2])
                merge_row_title_for_sheet(config_type,sheet,index1,2,index1+count-minus_count-1,color_array[x%2])
                merge_row_title_for_sheet(card_action,sheet,index1,3,index1+count-minus_count-1,color_array[x%2])
                merge_row_title_for_sheet(pass_cnt,sheet,index1,4,index1+count-minus_count-1,color_array[x%2])
                merge_row_title_for_sheet(fail_cnt,sheet,index1,5,index1+count-minus_count-1,color_array[x%2])
                merge_row_title_for_sheet(warning_cnt,sheet,index1,6,index1+count-minus_count-1,color_array[x%2])
                for index3 in range(index1, index1+count-minus_count):
                    for index4 in range(6,sheet.max_column+1):
                        if sheet.cell(index3,index4).fill != PatternFill(start_color="f5bf1a", fill_type = "solid"):
                            sheet.cell(index3,index4).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                if fail_cnt != 0:
                    sheet.cell(index1,5).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                if warning_cnt != 0:
                    sheet.cell(index1,6).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")                # --- create fail sheet ---
                    if f'F-{test_case}' not in wb.sheetnames and test_case in ['RS485 Status']:
                        wb.create_sheet('F-{}'.format(test_case),index=100)
                        set_title_for_sheet('Test case',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                        set_title_for_sheet('Card',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                        set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                        set_title_for_sheet('Card action',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                        set_title_for_sheet('RS485 info by slot',wb['F-{}'.format(test_case)],max_row,5,5,"5cb800")
                        wb['F-{}'.format(test_case)].merge_cells(start_row=max_row,end_row=max_row,start_column=5,end_column=5+len(slot_second_name)-1)
                        set_title_for_sheet('RS485 info',wb['F-{}'.format(test_case)],max_row,18,18,"5cb800")
                        wb['F-{}'.format(test_case)].merge_cells(start_row=max_row,end_row=max_row,start_column=18,end_column=18+len(sta_second_name)-1)
                        # --- second row ---
                        for i in range(0,len(slot_second_name)):
                            set_title_for_sheet(slot_second_name[i],wb['F-{}'.format(test_case)],max_row+1,5+i,5+i,"5cb800")
                        for i in range(0,len(sta_second_name)):
                            set_title_for_sheet(sta_second_name[i],wb['F-{}'.format(test_case)],max_row+1,18+i,18+i,"5cb800")    
                        # --- merge ---
                        for i in range(1,5):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                    fail_count = 0
                    minus_count = 0   
                    for count, fail_item in enumerate(RS485_item[info],start=1):
                        if RS485_item[info][count-1]['result'] == 'fail' or RS485_item[info][count-1]['result'] == 'warning':
                            fail_count += 1
                            index2 = wb['F-{}'.format(test_case)].max_row + 1
                            for ind, (key,val) in enumerate(fail_item.items()):
                                if len(fail_item) > 6:
                                    if ind == 0:
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        slot_num = re.search(r'slot (\d+)', info_data)
                                        slot_abc = re.search(r'slot ([a-zA-Z])', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,5).value = int(dut_num.group(1))
                                        if slot_num:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,6).value = int(slot_num.group(1))
                                        if slot_abc:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,6).value = slot_abc.group(1)
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,7).value = float(info_data)
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,8).value = float(info_data) 
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,9).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,9).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,10).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,10).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,11).value = float(info_data)
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,12).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,12).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 7:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,13).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,13).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 8:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,14).value = float(info_data)
                                    elif ind == 9:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,15).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,15).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 10:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,16).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,16).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 11:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2-minus_count,17).value = info_data
                                        if info_data == 'warning':
                                            wb['F-{}'.format(test_case)].cell(index2+minus_count,17).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")                                        
                                        if info_data == 'fail':
                                            wb['F-{}'.format(test_case)].cell(index2-minus_count,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                else:
                                    if ind == 0:
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index2+minus_count,18).value = int(dut_num.group(1))
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+minus_count,19).value = float(info_data)
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+minus_count,20).value = float(info_data) 
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+minus_count,21).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+minus_count,21).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+minus_count,22).value = info_data
                                        if info_data == 'warning':
                                            wb['F-{}'.format(test_case)].cell(index2+minus_count,22).fill = PatternFill(start_color="f5bf1a", fill_type = "solid")
                                        if info_data == 'fail':
                                            wb['F-{}'.format(test_case)].cell(index2+minus_count,22).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                wb['F-{}'.format(test_case)].cell(index2,1).value = test_case
                                wb['F-{}'.format(test_case)].cell(index2,2).value = card
                                wb['F-{}'.format(test_case)].cell(index2,3).value = config_type
                                wb['F-{}'.format(test_case)].cell(index2,4).value = card_action
                    # --- placed in center ---
                    alignment = Alignment(horizontal='center', vertical='center')
                    for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                        for cell in row:
                            cell.alignment = alignment
                    wb['F-{}'.format(test_case)].freeze_panes = 'E3'
            # --- placed in center ---
            alignment = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = alignment
            sheet.freeze_panes = 'G3'
        elif test_case == "PW Status":
            second_name = ['DUT','CTRL','Bundle ID','Primary/Secondary', "Jitter buffer", "L bit count", "R bit count", 
                           "TX Good count","RX Good count","RX lost count", "Link", "Fail reason","Result"]
            jitter_name = ['Under run','Over run','Depth','MIN','MAX']
            set_title_for_sheet('Card',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Pass',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Fail',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Working PW Status',sheet,max_row,6,6,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=6,end_column=6+len(second_name)+len(jitter_name)-2)
            set_title_for_sheet('Standby PW Status',sheet,max_row,23,23,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=23,end_column=23+len(second_name)+len(jitter_name)-2)
            # --- second row ---
            for i in range(0,5):
                set_title_for_sheet(second_name[i],sheet,max_row+1,6+i,6+i,"5cb800")
            for i in range(5,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,6+len(jitter_name)+i-1,6+len(jitter_name)+i-1,"5cb800")
            for i in range(0,5):
                set_title_for_sheet(second_name[i],sheet,max_row+1,23+i,23+i,"5cb800")
            for i in range(5,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,23+len(jitter_name)+i-1,23+len(jitter_name)+i-1,"5cb800")
            # --- third row ---
            for i in range(len(jitter_name)):
                set_title_for_sheet(jitter_name[i],sheet,max_row+2,10+i,10+i,"5cb800")
            for i in range(len(jitter_name)):
                set_title_for_sheet(jitter_name[i],sheet,max_row+2,27+i,27+i,"5cb800")
            # --- merge ---
            for i in range(1,6):
                sheet.merge_cells(start_row=1,end_row=3,start_column=i,end_column=i)
                #sheet.merge_cells(start_row=index,end_row=index+count-1,start_column=i,end_column=i)
            for i in range(6,10):
                sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
            for i in range(15,23):
                sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
            for i in range(23,27):
                sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
            for i in range(32,40):
                sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
            sheet.merge_cells(start_row=2,end_row=2,start_column=10,end_column=14)
            sheet.merge_cells(start_row=2,end_row=2,start_column=27,end_column=31)
            # --- placed in center ---
            alignment = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = alignment
            row_index = 4
            color_array = ["ffffff","d9d9d9"]
            for x,pw_item in enumerate(test_result[test_case]):
                index = sheet.max_row + 1
                info = list(pw_item.keys())[0]
                card, config_type,card_action = get_card_config_action(info)
                long_row = max(len(list(pw_item[info][status])) for status in pw_item[info].keys()) -1
                merge_row_title_for_sheet(card,sheet,row_index,1,row_index+long_row,color_array[x%2])
                merge_row_title_for_sheet(config_type,sheet,row_index,2,row_index+long_row,color_array[x%2])
                merge_row_title_for_sheet(card_action,sheet,row_index,3,row_index+long_row,color_array[x%2])
                for index1 in range(row_index, row_index+long_row+1):
                    for index2 in range(4,sheet.max_column+1):
                        sheet.cell(index1,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                        sheet.cell(index1,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                pass_cnt = 0
                fail_cnt = 0
                for status in pw_item[info].keys():
                    append_data = ""
                    cur_row_index = row_index
                    if status == 'Working PW status':
                        info_data = ''
                        for count,item in enumerate(pw_item[info][status],start=1):
                            for ind,(key,val) in enumerate(item.items()):
                                if ind == 0:
                                    info_data = F"{val}"
                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                    ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                    bundle_num = re.search(r'Bundle (\d+)', info_data)
                                    if dut_num:
                                        sheet.cell(index+count-1,6).value = float(dut_num.group(1))
                                    if ctrl_num:
                                        sheet.cell(index+count-1,7).value = float(ctrl_num.group(1))
                                    if bundle_num:
                                        sheet.cell(index+count-1,8).value = float(bundle_num.group(1))
                                    sheet.cell(index+count-1,9).value = (info_data.split())[-1]
                                elif ind == 1:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,10).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 2:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,11).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 3:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,12).value = info_data
                                    info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                    if info_data1:
                                        result = float(info_data1.group(1))
                                        if result > 12.0 or result < 4.0:
                                            sheet.cell(index+count-1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 4:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,13).value = float(info_data)
                                    if float(info_data) < 4.0:
                                        sheet.cell(index+count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 5:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,14).value = float(info_data)
                                    if float(info_data) > 20.0:
                                        sheet.cell(index+count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 6:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,15).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 7:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,16).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 8:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,17).value = float(info_data)
                                elif ind == 9:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,18).value = float(info_data)
                                elif ind == 10:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,19).value = float(info_data)
                                    if float(info_data) > 0.0:
                                        sheet.cell(index+count-1,19).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 11:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,20).value = info_data
                                    if info_data != "up":
                                        sheet.cell(index+count-1,20).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                if key == "fail_reason":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,21).value = info_data
                                    if info_data != '':
                                        sheet.cell(index+count-1,21).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif key == "result":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,22).value = info_data
                                    if info_data == 'fail':
                                        sheet.cell(index+count-1,22).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        fail_cnt += 1
                                    elif info_data == 'ok':
                                        pass_cnt += 1                  
                    elif status == 'Standby PW status':
                        #sheet.cell(cur_row_index,7).value = append_data
                        info_data = ''
                        for count,item in enumerate(pw_item[info][status],start=1):
                            for ind,(key,val) in enumerate(item.items()):
                                if ind == 0:
                                    info_data = F"{val}"
                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                    ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                    bundle_num = re.search(r'Bundle (\d+)', info_data)
                                    if dut_num:
                                        sheet.cell(index+count-1,23).value = float(dut_num.group(1))
                                    if ctrl_num:
                                        sheet.cell(index+count-1,24).value = float(ctrl_num.group(1))
                                    if bundle_num:
                                        sheet.cell(index+count-1,25).value = float(bundle_num.group(1))
                                    sheet.cell(index+count-1,26).value = (info_data.split())[-1]
                                elif ind == 1:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,27).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,27).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 2:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,28).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,28).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 3:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,29).value = info_data
                                    info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                    if info_data1:
                                        result = float(info_data1.group(1))
                                        if result > 12.0 or result < 4.0:
                                            sheet.cell(index+count-1,29).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 4:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,30).value = float(info_data)
                                    if float(info_data) < 4.0:
                                        sheet.cell(index+count-1,30).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 5:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,31).value = float(info_data)
                                    if float(info_data) > 20.0:
                                        sheet.cell(index+count-1,31).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 6:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,32).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,32).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 7:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,33).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,33).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 8:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,34).value = float(info_data)
                                elif ind == 9:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,35).value = float(info_data)
                                elif ind == 10:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,36).value = float(info_data)
                                    if float(info_data) > 0.0:
                                        sheet.cell(index+count-1,36).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 11:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,37).value = info_data
                                    if info_data != "up":
                                        sheet.cell(index+count-1,37).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                if key == "fail_reason":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,38).value = info_data
                                    if info_data != '':
                                        sheet.cell(index+count-1,38).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif key == "result":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,39).value = info_data
                                    if info_data == 'fail':
                                        sheet.cell(index+count-1,39).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        fail_cnt += 1
                                    elif info_data == 'ok':
                                        pass_cnt += 1
                    sheet.freeze_panes = 'F4'
                    append_data =""
                    cur_row_index+=1
                    merge_row_title_for_sheet(pass_cnt,sheet,row_index,4,row_index+long_row,color_array[x%2])
                    merge_row_title_for_sheet(fail_cnt,sheet,row_index,5,row_index+long_row,color_array[x%2])
                if fail_cnt != 0:
                    sheet.cell(row_index,5).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                # --- create fail sheet ---
                    if f'F-{test_case}' not in wb.sheetnames and test_case in ['PW Status']:
                        wb.create_sheet('F-{}'.format(test_case),index=100)
                        max_row = wb['F-{}'.format(test_case)].max_row
                        set_title_for_sheet('Card',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                        set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                        set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                        set_title_for_sheet('Working PW error info',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=4,end_column=4+len(second_name)+len(jitter_name)-2)
                        set_title_for_sheet('Standby PW error info',wb['F-{}'.format(test_case)],1,21,21,"5cb800")
                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=21,end_column=21+len(second_name)+len(jitter_name)-2)
                        # --- second row ---
                        for i in range(0,5):
                            set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,4+i,4+i,"5cb800")
                        for i in range(5,len(second_name)):
                            set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,4+len(jitter_name)+i-1,4+len(jitter_name)+i-1,"5cb800")
                        for i in range(0,5):
                            set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,21+i,21+i,"5cb800")
                        for i in range(5,len(second_name)):
                            set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,21+len(jitter_name)+i-1,21+len(jitter_name)+i-1,"5cb800")
                        # --- third row ---
                        for i in range(len(jitter_name)):
                            set_title_for_sheet(jitter_name[i],wb['F-{}'.format(test_case)],max_row+2,8+i,8+i,"5cb800")
                        for i in range(len(jitter_name)):
                            set_title_for_sheet(jitter_name[i],wb['F-{}'.format(test_case)],max_row+2,25+i,25+i,"5cb800")
                        # --- merge ---
                        for i in range(1,4):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=3,start_column=i,end_column=i)
                        for i in range(4,8):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        for i in range(13,21):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        for i in range(21,25):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        for i in range(30,38):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=2,start_column=8,end_column=12)
                        wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=2,start_column=25,end_column=29)
                    fail_count = 0
                    index2 = wb['F-{}'.format(test_case)].max_row + 1
                    if len(pw_item[info]['Working PW status']) > 0:
                        for count, fail_item in enumerate(pw_item[info]['Working PW status'],start=1):
                            if pw_item[info]['Working PW status'][count-1]['result'] == 'fail':
                                fail_count += 1
                                for ind, (key,val) in enumerate(fail_item.items()):
                                    if ind == 0:
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                        bundle_num = re.search(r'Bundle (\d+)', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,4).value = float(dut_num.group(1))
                                        if ctrl_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,5).value = float(ctrl_num.group(1))
                                        if bundle_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,6).value = float(bundle_num.group(1))
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,7).value = (info_data.split())[-1]
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,8).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,8).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,9).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,10).value = info_data
                                        info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                        if info_data1:
                                            result = float(info_data1.group(1))
                                            if result > 12.0 or result < 4.0:
                                                wb['F-{}'.format(test_case)].cell(index2+fail_count-1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,11).value = float(info_data)
                                        if float(info_data) < 4.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,12).value = float(info_data)
                                        if float(info_data) > 20.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,13).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 7:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,14).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 8:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,15).value = float(info_data)
                                    elif ind == 9:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,16).value = float(info_data)
                                    elif ind == 10:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,17).value = float(info_data)
                                        if float(info_data) > 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 11:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,18).value = info_data
                                        if info_data != "up":
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    if key == "fail_reason":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,19).value = info_data
                                    elif key == "result":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,20).value = info_data
                    if len(pw_item[info]['Standby PW status']) > 0:
                        for count, fail_item in enumerate(pw_item[info]['Standby PW status'],start=1):
                            if pw_item[info]['Standby PW status'][count-1]['result'] == 'fail':
                                fail_count += 1
                                fail_item = pw_item[info]['Standby PW status'][count-1]
                                for ind, (key,val) in enumerate(fail_item.items()):
                                    if ind == 0:
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                        bundle_num = re.search(r'Bundle (\d+)', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,21).value = float(dut_num.group(1))
                                        if ctrl_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,22).value = float(ctrl_num.group(1))
                                        if bundle_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,23).value = float(bundle_num.group(1)) 
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,24).value = (info_data.split())[-1]
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,25).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,25).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,26).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,26).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,27).value = info_data
                                        info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                        if info_data1:
                                            result = float(info_data1.group(1))
                                            if result > 12.0 or result < 4.0:
                                                wb['F-{}'.format(test_case)].cell(index2+fail_count-1,27).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,28).value = float(info_data)
                                        if float(info_data) < 4.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,28).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,29).value = float(info_data)
                                        if float(info_data) > 20.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,29).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,30).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,30).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 7:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,31).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,31).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 8:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,32).value = float(info_data)
                                    elif ind == 9:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,33).value = float(info_data)
                                    elif ind == 10:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,34).value = float(info_data)
                                        if float(info_data) > 0.0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,34).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 11:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,35).value = info_data
                                        if info_data != "up":
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,35).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    if key == "fail_reason":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,36).value = info_data
                                    elif key == "result":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,37).value = info_data
                    for i in range(index2, index2+fail_count):
                        wb['F-{}'.format(test_case)].cell(i,1).value = card
                        wb['F-{}'.format(test_case)].cell(i,2).value = config_type
                        wb['F-{}'.format(test_case)].cell(i,3).value = card_action
                    # --- placed in center ---
                    index1 = wb['F-{}'.format(test_case)].max_row
                    for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    wb['F-{}'.format(test_case)].freeze_panes = 'D4'   
                row_index += long_row+1
            # --- placed in center ---
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        elif test_case == "PW Statist":
            second_name = ['DUT','Bundle ID','Primary/Secondary', "bid_index","Eth1LinkStat", "Eth1RxSeqNum", "Eth1RxSeqLos", 
                           "Eth1RxJBLos","DiffDelay","Result"]
            set_title_for_sheet('Card',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Pass',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Fail',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Working PW Statist',sheet,max_row,6,6,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=6,end_column=6+len(second_name)-1)
            set_title_for_sheet('Standby PW Statist',sheet,max_row,6+len(second_name),6+len(second_name),"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=6+len(second_name),end_column=6+2*len(second_name)-1)
            # --- second row ---
            for i in range(0,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,6+i,6+i,"5cb800")
            for i in range(0,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,6+len(second_name)+i,6+len(second_name)+i,"5cb800")
            # --- merge ---
            for i in range(1,6):
                sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
            row_index = 3
            color_array = ["ffffff","d9d9d9"]
            for x,pw_item in enumerate(test_result[test_case]):
                index = sheet.max_row + 1
                info = list(pw_item.keys())[0]
                card, config_type,card_action = get_card_config_action(info)
                long_row = max(len(list(pw_item[info][status])) for status in pw_item[info].keys()) -1
                merge_row_title_for_sheet(card,sheet,row_index,1,row_index+long_row,color_array[x%2])
                merge_row_title_for_sheet(config_type,sheet,row_index,2,row_index+long_row,color_array[x%2])
                merge_row_title_for_sheet(card_action,sheet,row_index,3,row_index+long_row,color_array[x%2])
                for index1 in range(row_index, row_index+long_row+1):
                    for index2 in range(4,sheet.max_column+1):
                        sheet.cell(index1,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                        sheet.cell(index1,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                pass_cnt = 0
                fail_cnt = 0
                for status in pw_item[info].keys():
                    append_data = ""
                    cur_row_index = row_index
                    # print(pw_item[status])
                    if status == 'Working PW Static':
                        info_data = ''
                        for count,item in enumerate(pw_item[info][status],start=1):
                            for ind,(key,val) in enumerate(item.items()):
                                if ind == 0:
                                    info_data = F"{val}"
                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                    bundle_num = re.search(r'Bundle (\d+)', info_data)
                                    if dut_num:
                                        sheet.cell(index+count-1,6).value = float(dut_num.group(1))
                                    if bundle_num:
                                        sheet.cell(index+count-1,7).value = float(bundle_num.group(1))
                                    sheet.cell(index+count-1,8).value = (info_data.split())[-1]
                                elif ind == 1:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,9).value = float(info_data)
                                    #if float(info_data) != 0.0:
                                        #sheet.cell(index+count-1,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 2:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,10).value = info_data
                                    #if info_data != "up":
                                        #sheet.cell(index+count-1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 3:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,11).value = int(info_data)
                                elif ind == 4:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,12).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 5:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,13).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 6:
                                    info_data = F"{val}"
                                    if (info_data.replace('"', '').replace('.', '')).isdigit():
                                        sheet.cell(index+count-1,14).value = float(info_data.replace('"', '').replace('.', ''))
                                    else:
                                        sheet.cell(index+count-1,14).value = info_data.replace('"', '').replace('.', '')
                                    #if float(info_data) != 0.0:
                                        #sheet.cell(index+count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif key == "result":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,15).value = info_data
                                    if info_data == 'fail':
                                        sheet.cell(index+count-1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        fail_cnt += 1
                                    elif info_data == 'ok':
                                        pass_cnt += 1
                    elif status == 'Standby PW Static':
                        info_data = ''
                        for count,item in enumerate(pw_item[info][status],start=1):
                            for ind,(key,val) in enumerate(item.items()):
                                if ind == 0:
                                    info_data = F"{val}"
                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                    bundle_num = re.search(r'Bundle (\d+)', info_data)
                                    if dut_num:
                                        sheet.cell(index+count-1,16).value = float(dut_num.group(1))
                                    if bundle_num:
                                        sheet.cell(index+count-1,17).value = float(bundle_num.group(1))
                                    sheet.cell(index+count-1,18).value = (info_data.split())[-1]
                                elif ind == 1:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,19).value = float(info_data)
                                    #if float(info_data) != 0.0:
                                        #sheet.cell(index+count-1,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 2:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,20).value = info_data
                                    #if info_data != "up":
                                        #sheet.cell(index+count-1,20).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 3:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,21).value = int(info_data)
                                elif ind == 4:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,22).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,22).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 5:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,23).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,23).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 6:
                                    info_data = F"{val}"
                                    if (info_data.replace('"', '').replace('.', '')).isdigit():
                                        sheet.cell(index+count-1,24).value = float(info_data.replace('"', '').replace('.', ''))
                                    else:
                                        sheet.cell(index+count-1,24).value = info_data.replace('"', '').replace('.', '')
                                    #if float(info_data) != 0.0:
                                        #sheet.cell(index+count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif key == "result":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,25).value = info_data
                                    if info_data == 'fail':
                                        sheet.cell(index+count-1,25).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        fail_cnt += 1
                                    elif info_data == 'ok':
                                        pass_cnt += 1
                    
                    sheet.freeze_panes = 'F3'
                    append_data =""
                    cur_row_index+=1
                    merge_row_title_for_sheet(pass_cnt,sheet,row_index,4,row_index+long_row,color_array[x%2])
                    merge_row_title_for_sheet(fail_cnt,sheet,row_index,5,row_index+long_row,color_array[x%2])
                if fail_cnt != 0:
                    sheet.cell(row_index,5).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                # --- create fail sheet ---
                    if f'F-{test_case}' not in wb.sheetnames and test_case in ['PW Statist']:
                        wb.create_sheet('F-{}'.format(test_case),index=100)
                        max_row = wb['F-{}'.format(test_case)].max_row
                        set_title_for_sheet('Card',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                        set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                        set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                        set_title_for_sheet('Working PW error info',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=4,end_column=4+len(second_name)-1)
                        set_title_for_sheet('Standby PW error info',wb['F-{}'.format(test_case)],1,14,14,"5cb800")
                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=14,end_column=14+len(second_name)-1)
                        # --- second row ---
                        for i in range(0,len(second_name)):
                            set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,4+i,4+i,"5cb800")
                        for i in range(0,len(second_name)):
                            set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,4+len(second_name)+i,4+len(second_name)+i,"5cb800")
                        # --- merge ---
                        for i in range(1,4):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                    fail_count = 0
                    index2 = wb['F-{}'.format(test_case)].max_row + 1
                    if len(pw_item[info]['Working PW Static']) > 0:
                        for count, fail_item in enumerate(pw_item[info]['Working PW Static'],start=1):
                            if pw_item[info]['Working PW Static'][count-1]['result'] == 'fail':
                                fail_count += 1
                                for ind, (key,val) in enumerate(fail_item.items()):
                                    if ind == 0:
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        bundle_num = re.search(r'Bundle (\d+)', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,4).value = float(dut_num.group(1))
                                        if bundle_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,5).value = float(bundle_num.group(1))
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,6).value = (info_data.split())[-1]
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,7).value = float(info_data)
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,8).value = info_data
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,9).value = float(info_data)
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,10).value = float(info_data)
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,11).value = float(info_data)
                                        if info_data != 0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        if (info_data.replace('"', '').replace('.', '')).isdigit():
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,12).value = float(info_data.replace('"', '').replace('.', ''))
                                        else:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,12).value = info_data.replace('"', '').replace('.', '')
                                    elif key == "result":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,13).value = info_data
                    if len(pw_item[info]['Standby PW Static']) > 0:
                        for count, fail_item in enumerate(pw_item[info]['Standby PW Static'],start=1):
                            if pw_item[info]['Standby PW Static'][count-1]['result'] == 'fail':
                                fail_count += 1
                                fail_item = pw_item[info]['Standby PW Static'][count-1]
                                for ind, (key,val) in enumerate(fail_item.items()):
                                    if ind == 0:
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        bundle_num = re.search(r'Bundle (\d+)', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,14).value = float(dut_num.group(1))
                                        if bundle_num:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,15).value = float(bundle_num.group(1)) 
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,16).value = (info_data.split())[-1]
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,17).value = float(info_data)
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,18).value = info_data
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,19).value = float(info_data)
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,20).value = float(info_data)
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,21).value = float(info_data)
                                        if info_data != 0:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,21).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        if (info_data.replace('"', '').replace('.', '')).isdigit():
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,22).value = float(info_data.replace('"', '').replace('.', ''))
                                        else:
                                            wb['F-{}'.format(test_case)].cell(index2+fail_count-1,22).value = info_data.replace('"', '').replace('.', '')
                                    elif key == "result":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index2+fail_count-1,23).value = info_data
                    for i in range(index2, index2+fail_count):
                        wb['F-{}'.format(test_case)].cell(i,1).value = card
                        wb['F-{}'.format(test_case)].cell(i,2).value = config_type
                        wb['F-{}'.format(test_case)].cell(i,3).value = card_action
                    # --- placed in center ---
                    index1 = wb['F-{}'.format(test_case)].max_row
                    for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    wb['F-{}'.format(test_case)].freeze_panes = 'D3'   
                row_index += long_row+1
            # --- placed in center ---
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        elif test_case == 'Card port status':
            max_row = sheet.max_row
            set_title_for_sheet('Card',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Port status',sheet,max_row,4,4,"5cb800")
            card_row = ['QE1','FE1/MQE1','FOM']
            qe1_row = ['DUT','Slot','Port','Qe1PortLOS','Qe1PortLOF','Qe1PortRcvAIS','Qe1PortRcvRAI'
                          'Qe1PortXmtAIS','Qe1PortXmtRAI','Qe1PortXmtAIS','Qe1PortBPVcount','Qe1PortEScount','Result']
            fe1_row = ['DUT','Slot','Port','Result']
            fom_row = ['DUT','Slot','Port','FomType','FomOpticalLOS','FomOpticalLOF','FomOpticalFE',
                              'FomOpticalFCS','FomOpticalEOC','FomE1LOF','FomE1CRC','FomE1RxAIS','FomE1RxRAI',
                              'FomE1TxAIS','FomE1TxRAI','FomE1LOS','FomE1BPVCRC','Result']
            set_title_for_sheet('QE1',sheet,max_row+1,4,4,"5cb800")
            set_title_for_sheet('FE1/MQE1',sheet,max_row+1,4+len(qe1_row),4+len(qe1_row),"5cb800")
            set_title_for_sheet('FOM',sheet,max_row+1,4+len(qe1_row)+len(fe1_row),4+len(qe1_row)+len(fe1_row),"5cb800")
            for i in range(len(qe1_row)):
                set_title_for_sheet(qe1_row[i],sheet,max_row+2,4+i,4+i,"5cb800")
            for i in range(len(fe1_row)):
                set_title_for_sheet(fe1_row[i],sheet,max_row+2,4+len(qe1_row)+i,4+len(qe1_row)+i,"5cb800")
            for i in range(len(fom_row)):
                set_title_for_sheet(fom_row[i],sheet,max_row+2,4+len(qe1_row)+len(fe1_row)+i,4+len(qe1_row)+
                                    len(fe1_row)+i,"5cb800")
            # --- merge ---
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=4,
                              end_column=4+len(qe1_row)+len(fe1_row)+len(fom_row)-1)
            sheet.merge_cells(start_row=max_row+1,end_row=max_row+1,start_column=4,
                              end_column=4+len(qe1_row)-1)
            sheet.merge_cells(start_row=max_row+1,end_row=max_row+1,start_column=4+len(qe1_row),
                              end_column=4+len(qe1_row)+len(fe1_row)-1)
            sheet.merge_cells(start_row=max_row+1,end_row=max_row+1,start_column=4+len(qe1_row)+len(fe1_row),
                              end_column=4+len(qe1_row)+len(fe1_row)+len(fom_row)-1)
            for i in range(1,4):
                sheet.merge_cells(start_row=1,end_row=3,start_column=i,
                              end_column=i)
            row_index = 4
            color_array = ["ffffff","d9d9d9"]
            for x, card_port_item in enumerate(test_result[test_case]):
                info = list(card_port_item.keys())[0]
                card, config_type,card_action = get_card_config_action(info)
                merge_row_title_for_sheet(card,sheet,row_index,1,row_index+len(card_port_item[info])-1,color_array[x%2])
                merge_row_title_for_sheet(config_type,sheet,row_index,2,row_index+len(card_port_item[info])-1,color_array[x%2])
                merge_row_title_for_sheet(card_action,sheet,row_index,3,row_index+len(card_port_item[info])-1,color_array[x%2])
                append_data = ""
                for item in card_port_item[info]:
                    if " QE1 " in list(item.values())[0]:
                        for ind,(key,val) in enumerate(item.items()):
                            append_data += F"{key}: {val} "
                            if ind == 0:
                                dut_num = re.search(r'DUT (\d+)', F"{val}")
                                slot_num = re.search(r'Slot (\d+)', F"{val}")
                                port_num = re.search(r'Port (\d+)', F"{val}")
                                sheet.cell(row_index,4).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                sheet.cell(row_index,5).value = float(re.search(r'Slot (\d+)', F"{val}").group(1))
                                sheet.cell(row_index,6).value = float(re.search(r'Port (\d+)', F"{val}").group(1))
                            elif ind == 7 or ind == 8:
                                sheet.cell(row_index,6+ind).value = val
                            else:
                                sheet.cell(row_index,6+ind).value = val
                    if " FE1 " in list(item.values())[0] or " MQE1 " in list(item.values())[0]:
                        for ind,(key,val) in enumerate(item.items()):
                            append_data += F"{key}: {val} "
                            if ind == 0:
                                dut_num = re.search(r'DUT (\d+)', F"{val}")
                                slot_num = re.search(r'Slot (\d+)', F"{val}")
                                port_num = re.search(r'Port (\d+)', F"{val}")
                                sheet.cell(row_index,4+len(qe1_row)).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                sheet.cell(row_index,5+len(qe1_row)).value = float(re.search(r'Slot (\d+)', F"{val}").group(1))
                                sheet.cell(row_index,6+len(qe1_row)).value = float(re.search(r'Port (\d+)', F"{val}").group(1))
                            else:
                                sheet.cell(row_index,6+len(qe1_row)+ind).value = val
                    if " FOM " in list(item.values())[0]:
                        for ind,(key,val) in enumerate(item.items()):
                            append_data += F"{key}: {val} "
                            if ind == 0:
                                dut_num = re.search(r'DUT (\d+)', F"{val}")
                                slot_num = re.search(r'Slot (\d+)', F"{val}")
                                port_num = re.search(r'Port (\d+)', F"{val}")
                                sheet.cell(row_index,4+len(qe1_row)+len(fe1_row)).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                sheet.cell(row_index,5+len(qe1_row)+len(fe1_row)).value = float(re.search(r'Slot (\d+)', F"{val}").group(1))
                                sheet.cell(row_index,6+len(qe1_row)+len(fe1_row)).value = float(re.search(r'Port (\d+)', F"{val}").group(1))
                            elif ind == 4 or ind == 5 or ind == 6 or ind == 8 or ind == 14:
                                sheet.cell(row_index,6+ind+len(qe1_row)+len(fe1_row)).value = float(val)
                            else:
                                sheet.cell(row_index,6+ind+len(qe1_row)+len(fe1_row)).value = val
                    #sheet.cell(row_index,4).value = append_data
                    if 'result: fail' in append_data:
                        if f'F-{test_case}' not in wb.sheetnames and test_case in ['Card port status']:
                            wb.create_sheet('F-{}'.format(test_case),index=100)
                            max_row = wb['F-{}'.format(test_case)].max_row
                            set_title_for_sheet('Card',wb['F-{}'.format(test_case)],max_row,1,1,"5cb800")
                            set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],max_row,2,2,"5cb800")
                            set_title_for_sheet('Action',wb['F-{}'.format(test_case)],max_row,3,3,"5cb800")
                            set_title_for_sheet('Port status',wb['F-{}'.format(test_case)],max_row,4,4,"5cb800")
                            card_row = ['QE1','FE1/MQE1','FOM']
                            qe1_row = ['DUT','Slot','Port','Qe1PortLOS','Qe1PortLOF','Qe1PortRcvAIS','Qe1PortRcvRAI'
                                        'Qe1PortXmtAIS','Qe1PortXmtRAI','Qe1PortXmtAIS','Qe1PortBPVcount','Qe1PortEScount','Result']
                            fe1_row = ['DUT','Slot','Port','Result']
                            fom_row = ['DUT','Slot','Port','FomType','FomOpticalLOS','FomOpticalLOF','FomOpticalFE',
                                            'FomOpticalFCS','FomOpticalEOC','FomE1LOF','FomE1CRC','FomE1RxAIS','FomE1RxRAI',
                                            'FomE1TxAIS','FomE1TxRAI','FomE1LOS','FomE1BPVCRC','Result']
                            set_title_for_sheet('QE1',wb['F-{}'.format(test_case)],max_row+1,4,4,"5cb800")
                            set_title_for_sheet('FE1/MQE1',wb['F-{}'.format(test_case)],max_row+1,4+len(qe1_row),4+len(qe1_row),"5cb800")
                            set_title_for_sheet('FOM',wb['F-{}'.format(test_case)],max_row+1,4+len(qe1_row)+len(fe1_row),4+len(qe1_row)+len(fe1_row),"5cb800")
                            for i in range(len(qe1_row)):
                                set_title_for_sheet(qe1_row[i],wb['F-{}'.format(test_case)],max_row+2,4+i,4+i,"5cb800")
                            for i in range(len(fe1_row)):
                                set_title_for_sheet(fe1_row[i],wb['F-{}'.format(test_case)],max_row+2,4+len(qe1_row)+i,4+len(qe1_row)+i,"5cb800")
                            for i in range(len(fom_row)):
                                set_title_for_sheet(fom_row[i],wb['F-{}'.format(test_case)],max_row+2,4+len(qe1_row)+len(fe1_row)+i,4+len(qe1_row)+
                                                    len(fe1_row)+i,"5cb800")
                            # --- merge ---
                            wb['F-{}'.format(test_case)].merge_cells(start_row=max_row,end_row=max_row,start_column=4,
                                            end_column=4+len(qe1_row)+len(fe1_row)+len(fom_row)-1)
                            wb['F-{}'.format(test_case)].merge_cells(start_row=max_row+1,end_row=max_row+1,start_column=4,
                                            end_column=4+len(qe1_row)-1)
                            wb['F-{}'.format(test_case)].merge_cells(start_row=max_row+1,end_row=max_row+1,start_column=4+len(qe1_row),
                                            end_column=4+len(qe1_row)+len(fe1_row)-1)
                            wb['F-{}'.format(test_case)].merge_cells(start_row=max_row+1,end_row=max_row+1,start_column=4+len(qe1_row)+len(fe1_row),
                                            end_column=4+len(qe1_row)+len(fe1_row)+len(fom_row)-1)
                            for i in range(1,4):
                                wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=3,start_column=i,
                                            end_column=i)
                        max_row = wb['F-{}'.format(test_case)].max_row+1
                        merge_row_title_for_sheet(card,wb['F-{}'.format(test_case)],max_row,1,max_row,color_array[x%2])
                        merge_row_title_for_sheet(config_type,wb['F-{}'.format(test_case)],max_row,2,max_row,color_array[x%2])
                        merge_row_title_for_sheet(card_action,wb['F-{}'.format(test_case)],max_row,3,max_row,color_array[x%2])
                        if " QE1 " in list(item.values())[0]:
                            for ind,(key,val) in enumerate(item.items()):
                                append_data += F"{key}: {val} "
                                if ind == 0:
                                    dut_num = re.search(r'DUT (\d+)', F"{val}")
                                    slot_num = re.search(r'Slot (\d+)', F"{val}")
                                    port_num = re.search(r'Port (\d+)', F"{val}")
                                    wb['F-{}'.format(test_case)].cell(max_row,4).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                    wb['F-{}'.format(test_case)].cell(max_row,5).value = float(re.search(r'Slot (\d+)', F"{val}").group(1))
                                    wb['F-{}'.format(test_case)].cell(max_row,6).value = float(re.search(r'Port (\d+)', F"{val}").group(1))
                                elif ind == 7 or ind == 8:
                                    wb['F-{}'.format(test_case)].cell(max_row,6+ind).value = float(val)
                                else:
                                    wb['F-{}'.format(test_case)].cell(max_row,6+ind).value = val
                        if " FE1 " in list(item.values())[0] or " MQE1 " in list(item.values())[0]:
                            for ind,(key,val) in enumerate(item.items()):
                                append_data += F"{key}: {val} "
                                if ind == 0:
                                    dut_num = re.search(r'DUT (\d+)', F"{val}")
                                    slot_num = re.search(r'Slot (\d+)', F"{val}")
                                    port_num = re.search(r'Port (\d+)', F"{val}")
                                    wb['F-{}'.format(test_case)].cell(max_row,4+len(qe1_row)).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                    wb['F-{}'.format(test_case)].cell(max_row,5+len(qe1_row)).value = float(re.search(r'Slot (\d+)', F"{val}").group(1))
                                    wb['F-{}'.format(test_case)].cell(max_row,6+len(qe1_row)).value = float(re.search(r'Port (\d+)', F"{val}").group(1))
                                else:
                                    wb['F-{}'.format(test_case)].cell(max_row,6+len(qe1_row)+ind).value = val
                        if " FOM " in list(item.values())[0]:
                            for ind,(key,val) in enumerate(item.items()):
                                append_data += F"{key}: {val} "
                                if ind == 0:
                                    dut_num = re.search(r'DUT (\d+)', F"{val}")
                                    slot_num = re.search(r'Slot (\d+)', F"{val}")
                                    port_num = re.search(r'Port (\d+)', F"{val}")
                                    wb['F-{}'.format(test_case)].cell(max_row,4+len(qe1_row)+len(fe1_row)).value = float(re.search(r'DUT (\d+)', F"{val}").group(1))
                                    wb['F-{}'.format(test_case)].cell(max_row,5+len(qe1_row)+len(fe1_row)).value = float(re.search(r'Slot (\d+)', F"{val}").group(1))
                                    wb['F-{}'.format(test_case)].cell(max_row,6+len(qe1_row)+len(fe1_row)).value = float(re.search(r'Port (\d+)', F"{val}").group(1))
                                elif ind == 4 or ind == 5 or ind == 6 or ind == 8 or ind == 14:
                                    wb['F-{}'.format(test_case)].cell(max_row,6+ind+len(qe1_row)+len(fe1_row)).value = float(val)
                                else:
                                    wb['F-{}'.format(test_case)].cell(max_row,6+ind+len(qe1_row)+len(fe1_row)).value = val
                        #create_fail_info_sheet(wb,test_case,card,config_type,card_action,append_data)
                        # --- placed in center ---
                        alignment = Alignment(horizontal='center', vertical='center')
                        for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                            for cell in row:
                                cell.alignment = alignment
                        wb['F-{}'.format(test_case)].freeze_panes = 'D4'
                    append_data = ""
                    # --- color for gray ---
                    for index in range(row_index, sheet.max_row+1):
                        for index2 in range(4,sheet.max_column+1):
                            if "fail" in str(sheet.cell(index,index2).value):
                                sheet.cell(index,index2).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            if sheet.cell(index,index2).fill != PatternFill(start_color="e53b31", fill_type = "solid"):
                                sheet.cell(index,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                    row_index += 1
                # --- placed in center ---
                alignment = Alignment(horizontal='center', vertical='center')
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = alignment
                sheet.freeze_panes = 'D4'
        elif test_case == 'Long Time Bert':
            set_title_for_sheet('Card',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Config type',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('QE1 Bert Info',sheet,max_row,4,4,"5cb800")

            set_title_for_sheet('Time',sheet,max_row+1,4,4,"5cb800")
            set_title_for_sheet('PRBS Status',sheet,max_row+1,5,5,"5cb800")
            set_title_for_sheet('Elapse Seconds',sheet,max_row+1,6,6,"5cb800")
            set_title_for_sheet('Error Counts',sheet,max_row+1,7,7,"5cb800")
            set_title_for_sheet('Error Seconds',sheet,max_row+1,8,8,"5cb800")
            set_title_for_sheet('Result',sheet,max_row+1,9,9,"5cb800")
            # --- merge ---
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=4,end_column=9)
            for i in range(1,4):
                sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)

            row_index = 3
            color_array = ["ffffff","d9d9d9"]
            for x, qe1_bert_item in enumerate(test_result[test_case]):
                info = list(qe1_bert_item.keys())[0]
                card, config_type,card_action = get_card_config_action(info)
                merge_row_title_for_sheet(card,sheet,row_index,1,row_index+len(qe1_bert_item[info])-1,color_array[x%2])
                merge_row_title_for_sheet(config_type,sheet,row_index,2,row_index+len(qe1_bert_item[info])-1,color_array[x%2])
                merge_row_title_for_sheet(card_action,sheet,row_index,3,row_index+len(qe1_bert_item[info])-1,color_array[x%2])
                #for index in range(row_index, row_index+len(qe1_bert_item[info])):
                    #sheet.cell(index,4).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                append_data = ""
                for item in qe1_bert_item[info]:
                    for ind,(key,val) in enumerate(item.items()):
                        append_data += F"{key}: {val} "
                        if ind == 0 or ind == 1 or ind == 5:
                            sheet.cell(row_index,4+ind).value = val
                        else:
                            sheet.cell(row_index,4+ind).value = float(val)
                    append_data = ""
                    # --- color for gray ---
                    for index in range(row_index, sheet.max_row+1):
                        for index2 in range(4,sheet.max_column+1):
                            if "fail" in str(sheet.cell(index,index2).value):
                                sheet.cell(index,index2).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            if sheet.cell(index,index2).fill != PatternFill(start_color="e53b31", fill_type = "solid"):
                                sheet.cell(index,index2).fill = PatternFill(start_color=color_array[x%2], fill_type = "solid")
                    row_index += 1
                # --- placed in center ---
                alignment = Alignment(horizontal='center', vertical='center')
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = alignment
                sheet.freeze_panes = 'D3'
                if 'result: fail' in append_data:
                    if f'F-{test_case}' not in wb.sheetnames and test_case in ['Long Time Bert']:
                        wb.create_sheet('F-{}'.format(test_case),index=100)
                        max_row = wb['F-{}'.format(test_case)].max_row
                        set_title_for_sheet('Card',wb['F-{}'.format(test_case)],max_row,1,1,"5cb800")
                        set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],max_row,2,2,"5cb800")
                        set_title_for_sheet('Action',wb['F-{}'.format(test_case)],max_row,3,3,"5cb800")
                        set_title_for_sheet('QE1 Bert Info',wb['F-{}'.format(test_case)],max_row,4,4,"5cb800")

                        set_title_for_sheet('Time',wb['F-{}'.format(test_case)],max_row+1,4,4,"5cb800")
                        set_title_for_sheet('PRBS Status',wb['F-{}'.format(test_case)],max_row+1,5,5,"5cb800")
                        set_title_for_sheet('Error Counts',wb['F-{}'.format(test_case)],max_row+1,6,6,"5cb800")
                        set_title_for_sheet('Error Seconds',wb['F-{}'.format(test_case)],max_row+1,7,7,"5cb800")
                        set_title_for_sheet('Result',wb['F-{}'.format(test_case)],max_row+1,8,8,"5cb800")
                        # --- merge ---
                        wb['F-{}'.format(test_case)].merge_cells(start_row=max_row,end_row=max_row,start_column=4,end_column=8)
                        for i in range(1,4):
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                    max_row = wb['F-{}'.format(test_case)].max_row+1
                    merge_row_title_for_sheet(card,wb['F-{}'.format(test_case)],max_row,1,max_row,color_array[x%2])
                    merge_row_title_for_sheet(config_type,wb['F-{}'.format(test_case)],max_row,2,max_row,color_array[x%2])
                    merge_row_title_for_sheet(card_action,wb['F-{}'.format(test_case)],max_row,3,max_row,color_array[x%2])
                    for ind,(key,val) in enumerate(item.items()):
                        append_data += F"{key}: {val} "
                        for ind,(key,val) in enumerate(item.items()):
                            if ind == 0 or ind == 1 or ind == 4:
                                wb['F-{}'.format(test_case)].cell(row_index,4+ind).value = val
                            else:
                                wb['F-{}'.format(test_case)].cell(row_index,4+ind).value = float(val)
                    append_data = ""
                    # --- placed in center ---
                    alignment = Alignment(horizontal='center', vertical='center')
                    for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                        for cell in row:
                            cell.alignment = alignment
                    wb['F-{}'.format(test_case)].freeze_panes = 'D3'
        elif test_case == 'Bundle test(bert)':
            set_title_for_sheet('Bundle mode',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Bundle IDs',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Bert type',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Pass',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Fail',sheet,max_row,6,6,"5cb800")
            for i in range(1,7):
                sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
            set_title_for_sheet('Ctrl Fail item',sheet,1,7,7,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=7,end_column=10)
            set_title_for_sheet('Time',sheet,2,7,7,"5cb800")
            set_title_for_sheet('Bit Error',sheet,2,8,8,"5cb800")
            set_title_for_sheet('Error ID',sheet,2,9,9,"5cb800")
            set_title_for_sheet('Elapsed Time',sheet,2,10,10,"5cb800")

            set_title_for_sheet('Test Card Fail item',sheet,1,11,1,"5cb800")
            set_title_for_sheet('Time',sheet,2,11,11,"5cb800")
            set_title_for_sheet('PRBS Status',sheet,2,12,12,"5cb800")
            set_title_for_sheet('Error Counts',sheet,2,13,13,"5cb800")
            set_title_for_sheet('Error ID',sheet,2,14,14,"5cb800")
            set_title_for_sheet('Error Seconds',sheet,2,15,15,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=11,end_column=15)
            max_row = 1
            for vlan_mode,vlan_mode_item in test_result[test_case].items():
                for bundle_ids, bert_item in vlan_mode_item.items():
                    for bert_type1,bert_result1 in bert_item.items():
                        for bert_type,bert_result in bert_result1.items():
                            index = sheet.max_row + 1
                            merge_row_title_for_sheet(vlan_mode,sheet,index,1,index+bert_result['fail']-1,"FFFFFF")
                            merge_row_title_for_sheet(bundle_ids,sheet,index,2,index+bert_result['fail']-1,"FFFFFF")
                            merge_row_title_for_sheet(bert_type,sheet,index,3,index+bert_result['fail']-1,"FFFFFF")
                            merge_row_title_for_sheet(bert_type1,sheet,index,4,index+bert_result['fail']-1,"FFFFFF")
                            merge_row_title_for_sheet(bert_result['pass'],sheet,index,5,index+bert_result['fail']-1,"FFFFFF")
                            merge_row_title_for_sheet(bert_result['fail'],sheet,index,6,index+bert_result['fail']-1,"FFFFFF")
                            bert_fail_data = ''
                            if bert_result['fail'] > 0:
                                sheet.cell(index,6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                for fail_item in bert_result['fail_item']:
                                    if type(fail_item) == dict:
                                        for key in fail_item.keys():
                                            bert_fail_data += key+":"+fail_item[key]+" "
                                    bert_fail_data += "\n"
                            # ----- create fail sheet ----- #
                                    wb.create_sheet('F-{}'.format(test_case),index=100)
                                    set_title_for_sheet('Test Item',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                                    set_title_for_sheet('Bundle mode',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                                    set_title_for_sheet('Bundle IDs',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                                    set_title_for_sheet('Bert type',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                                    set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
                                    for i in range(1,6):
                                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                                    set_title_for_sheet('Ctrl Fail item',wb['F-{}'.format(test_case)],1,6,6,"5cb800")
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=6,end_column=9)
                                    set_title_for_sheet('Time',wb['F-{}'.format(test_case)],2,6,6,"5cb800")
                                    set_title_for_sheet('Unsync Second',wb['F-{}'.format(test_case)],2,7,7,"5cb800")
                                    set_title_for_sheet('Error ID',wb['F-{}'.format(test_case)],2,8,8,"5cb800")
                                    set_title_for_sheet('Elapsed Time',wb['F-{}'.format(test_case)],2,9,9,"5cb800")
                                    
                                    set_title_for_sheet('Test Card Fail item',wb['F-{}'.format(test_case)],1,10,10,"5cb800")
                                    set_title_for_sheet('Time',wb['F-{}'.format(test_case)],2,10,10,"5cb800")
                                    set_title_for_sheet('PRBS Status',wb['F-{}'.format(test_case)],2,11,11,"5cb800")
                                    set_title_for_sheet('Error Counts',wb['F-{}'.format(test_case)],2,12,12,"5cb800")
                                    set_title_for_sheet('Error ID',wb['F-{}'.format(test_case)],2,13,13,"5cb800")
                                    set_title_for_sheet('Error Seconds',wb['F-{}'.format(test_case)],2,14,14,"5cb800")
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=10,end_column=14)
                                    index2 = wb['F-{}'.format(test_case)].max_row + 1
                                    for row_index2 in range(1,1+len(bert_fail_data.splitlines())):
                                        wb['F-{}'.format(test_case)].cell(index2,1).value = test_case
                                        wb['F-{}'.format(test_case)].cell(index2,2).value = vlan_mode
                                        wb['F-{}'.format(test_case)].cell(index2,3).value = bundle_ids
                                        wb['F-{}'.format(test_case)].cell(index2,4).value = bert_type
                                        wb['F-{}'.format(test_case)].cell(index2,5).value = bert_type1
                                        match1 = re.search(r'Time:(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) UnsyncSecond:(\d+) Error ID:(\d+) Elapsed Time:(\d+)', bert_fail_data.splitlines()[row_index2-1])
                                        if match1:
                                            wb['F-{}'.format(test_case)].cell(index2,6).value = match1.group(1)
                                            wb['F-{}'.format(test_case)].cell(index2,7).value = int(match1.group(2))
                                            wb['F-{}'.format(test_case)].cell(index2,8).value = int(match1.group(3))
                                            wb['F-{}'.format(test_case)].cell(index2,9).value = int(match1.group(4))
                                        match2 = re.search(r'Time:(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) PRBSStatus:(\w+) ErrorCounts:(\d+) Error ID:(\d+) ErrorSeconds:(\d+)', bert_fail_data.splitlines()[row_index2-1])
                                        if match2:
                                            wb['F-{}'.format(test_case)].cell(index2,10).value = match2.group(1)
                                            wb['F-{}'.format(test_case)].cell(index2,11).value = match2.group(2)
                                            wb['F-{}'.format(test_case)].cell(index2,12).value = int(match2.group(3))
                                            wb['F-{}'.format(test_case)].cell(index2,13).value = int(match2.group(4))
                                            wb['F-{}'.format(test_case)].cell(index2,14).value = int(match2.group(5))
                                wb['F-{}'.format(test_case)].freeze_panes = 'F3'
                                # --- placed in center ---
                                for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row,
                                                                                    min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                                    for cell in row:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                            # ----- create fail sheet ----- #
                            if len(bert_result['fail_item']) > 0:
                                if len(fail_item) < 5:
                                    for count, fail_item in enumerate(bert_result['fail_item'],start=1):
                                        for key,val in fail_item.items():
                                            if key == "Time":
                                                sheet.cell(index+count-1,7).value = val
                                            if key == "UnsyncSecond":
                                                sheet.cell(index+count-1,8).value = int(val)
                                            if key == "Error ID":
                                                sheet.cell(index+count-1,9).value = int(val)
                                            if key == "Elapsed Time":
                                                sheet.cell(index+count-1,10).value = int(val)
                                else:
                                    for key,val in fail_item.items():
                                        if key == "Time":
                                            sheet.cell(index+count-1,11).value = val
                                        if key == "PRBSStatus":
                                            sheet.cell(index+count-1,12).value = val
                                        if key == "ErrorCounts":
                                            sheet.cell(index+count-1,13).value = int(val)
                                        if key == "Error ID":
                                            sheet.cell(index+count-1,14).value = int(val)
                                        if key == "ErrorSeconds":
                                            sheet.cell(index+count-1,15).value = int(val)

            sheet.freeze_panes = 'G3'
            # --- placed in center ---
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        elif test_case == 'Bundle test(PW status)':
            second_name = ['Time','DUT','CTRL','Primary/Secondary', "Jitter buffer", "L bit count", "R bit count", 
                           "TX Good count","RX Good count","RX lost count", "Link", "Fail reason","Result"]
            jitter_name = ['Under run','Over run','Depth','MIN','MAX']
            set_title_for_sheet('Bundle mode',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Bundle IDs',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Check time',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Pass',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Fail',sheet,max_row,6,6,"5cb800")
            set_title_for_sheet('Working PW',sheet,max_row,7,7,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=7,end_column=7+len(second_name)+len(jitter_name)-2)
            set_title_for_sheet('Standby PW',sheet,max_row,24,24,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=24,end_column=24+len(second_name)+len(jitter_name)-2)
            # --- second row ---
            for i in range(0,5):
                set_title_for_sheet(second_name[i],sheet,max_row+1,7+i,7+i,"5cb800")
            for i in range(5,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,7+len(jitter_name)+i-1,7+len(jitter_name)+i-1,"5cb800")
            for i in range(0,5):
                set_title_for_sheet(second_name[i],sheet,max_row+1,24+i,24+i,"5cb800")
            for i in range(5,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,24+len(jitter_name)+i-1,24+len(jitter_name)+i-1,"5cb800")
            # --- third row ---
            for i in range(len(jitter_name)):
                set_title_for_sheet(jitter_name[i],sheet,max_row+2,11+i,11+i,"5cb800")
            for i in range(len(jitter_name)):
                set_title_for_sheet(jitter_name[i],sheet,max_row+2,28+i,28+i,"5cb800")
            for vlan_mode,vlan_mode_item in test_result[test_case].items():
                for bundle_ids, pw_item in vlan_mode_item.items():
                    for action1,pw_result1 in pw_item.items():
                        for action,pw_result in pw_result1.items():
                            if len(pw_result['Working PW error status']) > 0:
                                index = sheet.max_row + 1
                                sheet.cell(index,1).value = vlan_mode
                                sheet.cell(index,2).value = bundle_ids
                                sheet.cell(index,3).value = action1
                                sheet.cell(index,4).value = action
                                sheet.cell(index,5).value = pw_result['pass']
                                sheet.cell(index,6).value = pw_result['fail']
                                info_data = ''
                                for count, fail_item in enumerate(pw_result['Working PW error status'],start=1):
                                    for ind, (key,val) in enumerate(fail_item.items()):
                                        if ind == 0:
                                            sheet.cell(index+count-1,7).value = key
                                            info_data = F"{val}"
                                            dut_num = re.search(r'DUT (\d+)', info_data)
                                            ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                            if dut_num:
                                                sheet.cell(index+count-1,8).value = int(dut_num.group(1))
                                            if ctrl_num:
                                                sheet.cell(index+count-1,9).value = int(ctrl_num.group(1))
                                            sheet.cell(index+count-1,10).value = (info_data.split())[-1]
                                        elif ind == 1:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,11).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 2:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,12).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 3:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,13).value = info_data
                                            info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                            if info_data1:
                                                result = float(info_data1.group(1))
                                                if result > 12.0 or result < 4.0:
                                                    sheet.cell(index+count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 4:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,14).value = float(info_data)
                                            if float(info_data) < 4.0:
                                                sheet.cell(index+count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 5:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,15).value = float(info_data)
                                            if float(info_data) > 20.0:
                                                sheet.cell(index+count-1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 6:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,16).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 7:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,17).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 8:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,18).value = float(info_data)
                                        elif ind == 9:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,19).value = float(info_data)
                                        elif ind == 10:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,20).value = float(info_data)
                                            if float(info_data) > 0.0:
                                                sheet.cell(index+count-1,20).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 11:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,21).value = info_data
                                            if info_data != "up":
                                                sheet.cell(index+count-1,21).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        if key == "fail_reason":
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,22).value = info_data
                                            if info_data != '':
                                                sheet.cell(index+count-1,22).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif key == "result":
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,23).value = info_data
                                            if info_data == 'fail':
                                                sheet.cell(index+count-1,23).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                # --- merge ---
                                for i in range(1,7):
                                    sheet.merge_cells(start_row=1,end_row=3,start_column=i,end_column=i)
                                    sheet.merge_cells(start_row=index,end_row=index+count-1,start_column=i,end_column=i)
                                for i in range(7,11):
                                    sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                                for i in range(16,28):
                                    sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                                for i in range(33,41):
                                    sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                                sheet.merge_cells(start_row=2,end_row=2,start_column=11,end_column=15)
                                sheet.merge_cells(start_row=2,end_row=2,start_column=28,end_column=32)
                            
                            if len(pw_result['Standby PW error status']) > 0:
                                info_data = ''
                                for count, fail_item in enumerate(pw_result['Standby PW error status'],start=1):
                                    for ind, (key,val) in enumerate(fail_item.items()):
                                        if ind == 0:
                                            sheet.cell(index+count-1,24).value = key
                                            info_data = F"{val}"
                                            dut_num = re.search(r'DUT (\d+)', info_data)
                                            ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                            if dut_num:
                                                sheet.cell(index+count-1,25).value = float(dut_num.group(1))
                                            if ctrl_num:
                                                sheet.cell(index+count-1,26).value = float(ctrl_num.group(1))
                                            sheet.cell(index+count-1,27).value = (info_data.split())[-1]
                                        elif ind == 1:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,28).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,28).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 2:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,29).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,29).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 3:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,30).value = info_data
                                            info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                            if info_data1:
                                                result = float(info_data1.group(1))
                                                if result > 12.0 or result < 4.0:
                                                    sheet.cell(index+count-1,30).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 4:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,31).value = float(info_data)
                                            if float(info_data) < 4.0:
                                                sheet.cell(index+count-1,31).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 5:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,32).value = float(info_data)
                                            if float(info_data) > 20.0:
                                                sheet.cell(index+count-1,32).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 6:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,33).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,33).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 7:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,34).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                sheet.cell(index+count-1,34).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 8:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,35).value = float(info_data)
                                        elif ind == 9:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,36).value = float(info_data)
                                        elif ind == 10:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,37).value = float(info_data)
                                            if float(info_data) > 0.0:
                                                sheet.cell(index+count-1,37).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 11:
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,38).value = info_data
                                            if info_data != "up":
                                                sheet.cell(index+count-1,38).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        if key == "fail_reason":
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,39).value = info_data
                                            if info_data != '':
                                                sheet.cell(index+count-1,39).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif key == "result":
                                            info_data = F"{val}"
                                            sheet.cell(index+count-1,40).value = info_data
                                            if info_data == 'fail':
                                                sheet.cell(index+count-1,40).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            if pw_result['fail'] > 0:
                                sheet.cell(index,6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            # --- create fail sheet ---
                                if f'F-{test_case}' not in wb.sheetnames and test_case in ['Bundle test(PW status)']:
                                    wb.create_sheet('F-{}'.format(test_case),index=100)
                                    set_title_for_sheet('Bundle mode',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                                    set_title_for_sheet('Bundle IDs',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                                    set_title_for_sheet('Check time',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                                    set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                                    set_title_for_sheet('Working PW error info',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=5,end_column=5+len(second_name)+len(jitter_name)-2)
                                    set_title_for_sheet('Standby PW error info',wb['F-{}'.format(test_case)],1,22,22,"5cb800")
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=22,end_column=22+len(second_name)+len(jitter_name)-2)
                                    # --- second row ---
                                    for i in range(0,5):
                                        set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,5+i,5+i,"5cb800")
                                    for i in range(5,len(second_name)):
                                        set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,5+len(jitter_name)+i-1,5+len(jitter_name)+i-1,"5cb800")
                                    for i in range(0,5):
                                        set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,22+i,22+i,"5cb800")
                                    for i in range(5,len(second_name)):
                                        set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,22+len(jitter_name)+i-1,22+len(jitter_name)+i-1,"5cb800")
                                    # --- third row ---
                                    for i in range(len(jitter_name)):
                                        set_title_for_sheet(jitter_name[i],wb['F-{}'.format(test_case)],max_row+2,9+i,9+i,"5cb800")
                                    for i in range(len(jitter_name)):
                                        set_title_for_sheet(jitter_name[i],wb['F-{}'.format(test_case)],max_row+2,26+i,26+i,"5cb800")
                                    # --- merge ---
                                    for i in range(1,5):
                                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=3,start_column=i,end_column=i)
                                    for i in range(5,9):
                                        wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                                    for i in range(14,26):
                                        wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                                    for i in range(31,39):
                                        wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=2,start_column=9,end_column=13)
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=2,start_column=26,end_column=30)
                                fail_count = 0
                                for count, fail_item in enumerate(pw_result['Working PW error status'],start=1):
                                    if pw_result['Working PW error status'][count-1]['result'] == 'fail':
                                        fail_count += 1
                                        index1 = wb['F-{}'.format(test_case)].max_row + 1
                                        for ind, (key,val) in enumerate(fail_item.items()):
                                            if ind == 0:
                                                wb['F-{}'.format(test_case)].cell(index1,5).value = key
                                                info_data = F"{val}"
                                                dut_num = re.search(r'DUT (\d+)', info_data)
                                                ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                                if dut_num:
                                                    wb['F-{}'.format(test_case)].cell(index1,6).value = float(dut_num.group(1))
                                                if ctrl_num:
                                                    wb['F-{}'.format(test_case)].cell(index1,7).value = float(ctrl_num.group(1))
                                                wb['F-{}'.format(test_case)].cell(index1,8).value = (info_data.split())[-1]
                                            elif ind == 1:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,9).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 2:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,10).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 3:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,11).value = info_data
                                                info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                                if info_data1:
                                                    result = float(info_data1.group(1))
                                                    if result > 12.0 or result < 4.0:
                                                        wb['F-{}'.format(test_case)].cell(index1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 4:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,12).value = float(info_data)
                                                if float(info_data) < 4.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 5:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,13).value = float(info_data)
                                                if float(info_data) > 20.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 6:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,14).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 7:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,15).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 8:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,16).value = float(info_data)
                                            elif ind == 9:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,17).value = float(info_data)
                                            elif ind == 10:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,18).value = float(info_data)
                                                if float(info_data) > 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 11:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,19).value = info_data
                                                if info_data != "up":
                                                    wb['F-{}'.format(test_case)].cell(index1,19).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            if key == "fail_reason":
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,20).value = info_data
                                            elif key == "result":
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,21).value = info_data
                                        wb['F-{}'.format(test_case)].cell(index1,1).value = vlan_mode
                                        wb['F-{}'.format(test_case)].cell(index1,2).value = bundle_ids
                                        wb['F-{}'.format(test_case)].cell(index1,3).value = action1
                                        wb['F-{}'.format(test_case)].cell(index1,4).value = action
                                for count, fail_item in enumerate(pw_result['Standby PW error status'],start=1):
                                    if pw_result['Standby PW error status'][count-1]['result'] == 'fail':
                                        fail_count += 1
                                        index1 = wb['F-{}'.format(test_case)].max_row + 1
                                        fail_item = pw_result['Standby PW error status'][count-1]
                                        for ind, (key,val) in enumerate(fail_item.items()):
                                            if ind == 0:
                                                wb['F-{}'.format(test_case)].cell(index1,22).value = key
                                                info_data = F"{val}"
                                                dut_num = re.search(r'DUT (\d+)', info_data)
                                                ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                                if dut_num:
                                                    wb['F-{}'.format(test_case)].cell(index1,23).value = float(dut_num.group(1))
                                                if ctrl_num:
                                                    wb['F-{}'.format(test_case)].cell(index1,24).value = float(ctrl_num.group(1))
                                                wb['F-{}'.format(test_case)].cell(index1,25).value = (info_data.split())[-1]
                                            elif ind == 1:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,26).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,26).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 2:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,27).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,27).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 3:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,28).value = info_data
                                                info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                                if info_data1:
                                                    result = float(info_data1.group(1))
                                                    if result > 12.0 or result < 4.0:
                                                        wb['F-{}'.format(test_case)].cell(index1,28).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 4:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,29).value = float(info_data)
                                                if float(info_data) < 4.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,29).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 5:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,30).value = float(info_data)
                                                if float(info_data) > 20.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,30).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 6:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,31).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,31).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 7:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,32).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,32).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 8:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,33).value = float(info_data)
                                            elif ind == 9:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,34).value = float(info_data)
                                            elif ind == 10:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,35).value = float(info_data)
                                                if float(info_data) > 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,35).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 11:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,36).value = info_data
                                                if info_data != "up":
                                                    wb['F-{}'.format(test_case)].cell(index1,36).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            if key == "fail_reason":
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,37).value = info_data
                                            elif key == "result":
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1,38).value = info_data
                                # --- placed in center ---
                                alignment = Alignment(horizontal='center', vertical='center')
                                for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                                    for cell in row:
                                        cell.alignment = alignment
                                wb['F-{}'.format(test_case)].freeze_panes = 'E4'   
                                # --- merge ---
                                for i in range(1,5):
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=index1-fail_count+1,end_row=index1,start_column=i,end_column=i)
            # --- placed in center ---
            alignment = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = alignment
            sheet.freeze_panes = 'G4'
        elif test_case == 'Bundle test(RS485 status)':
            slot_second_name = ['Time','DUT','Slot','TX count', "RX count", "No RX count", "Timeout count", 
                           "Alive count","Non-alive count","Oversize count", "Init count", "Bad CMD count",
                            "Bad CMD num","Result"]
            sta_second_name = ['Time','DUT', 'FCS error', 'Slip error', 'Buffer error', 'Result']
            set_title_for_sheet('Bundle mode',sheet,max_row,1,1,"5cb800")
            set_title_for_sheet('Bundle IDs',sheet,max_row,2,2,"5cb800")
            set_title_for_sheet('Check time',sheet,max_row,3,3,"5cb800")
            set_title_for_sheet('Action',sheet,max_row,4,4,"5cb800")
            set_title_for_sheet('Pass',sheet,max_row,5,5,"5cb800")
            set_title_for_sheet('Fail',sheet,max_row,6,6,"5cb800")
            set_title_for_sheet('RS485 info by slot',sheet,max_row,7,7,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=7,end_column=7+len(slot_second_name)-1)
            set_title_for_sheet('RS485 info',sheet,max_row,21,21,"5cb800")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=21,end_column=21+len(sta_second_name)-1)
            for i in range(0,len(slot_second_name)):
                set_title_for_sheet(slot_second_name[i],sheet,max_row+1,7+i,7+i,"5cb800")
            for i in range(0,len(sta_second_name)):
                set_title_for_sheet(sta_second_name[i],sheet,max_row+1,21+i,21+i,"5cb800")
            for vlan_mode,vlan_mode_item in test_result[test_case].items():
                for bundle_ids, rs485_item in vlan_mode_item.items():
                    for action1,rs485_result1 in rs485_item.items():
                        for action,rs485_result in rs485_result1.items():
                            if len(rs485_result['RS485 error status']) > 0:
                                index = sheet.max_row + 1
                                sheet.cell(index,1).value = vlan_mode
                                sheet.cell(index,2).value = bundle_ids
                                sheet.cell(index,3).value = action1
                                sheet.cell(index,4).value = action
                                sheet.cell(index,5).value = rs485_result['pass']
                                sheet.cell(index,6).value = rs485_result['fail']
                                info_data = ''
                                minus_count = 0
                                for count, fail_item in enumerate(rs485_result['RS485 error status'],start=1):
                                    for ind, (key,val) in enumerate(fail_item.items()):
                                        if len(fail_item) > 6:
                                            if ind == 0:
                                                sheet.cell(index+count-minus_count-1,7).value = key
                                                info_data = F"{val}"
                                                dut_num = re.search(r'DUT (\d+)', info_data)
                                                slot_num = re.search(r'slot (\d+)', info_data)
                                                slot_abc = re.search(r'slot ([a-zA-Z])', info_data)
                                                if dut_num:
                                                    sheet.cell(index+count-minus_count-1,8).value = int(dut_num.group(1))
                                                if slot_num:
                                                    sheet.cell(index+count-minus_count-1,9).value = int(slot_num.group(1))
                                                if slot_abc:
                                                    sheet.cell(index+count-minus_count-1,9).value = slot_abc.group(1)
                                            elif ind == 1:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,10).value = float(info_data)
                                            elif ind == 2:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,11).value = float(info_data) 
                                            elif ind == 3:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,12).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    sheet.cell(index+count-minus_count-1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 4:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,13).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    sheet.cell(index+count-minus_count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 5:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,14).value = float(info_data)
                                            elif ind == 6:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,15).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    sheet.cell(index+count-minus_count-1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 7:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,16).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    sheet.cell(index+count-minus_count-1,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 8:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,17).value = float(info_data)
                                            elif ind == 9:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,18).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    sheet.cell(index+count-minus_count-1,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 10:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,19).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    sheet.cell(index+count-minus_count-1,19).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 11:
                                                info_data = F"{val}"
                                                sheet.cell(index+count-minus_count-1,20).value = info_data
                                                if info_data == 'fail':
                                                    sheet.cell(index+count-minus_count-1,20).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        else:
                                            if ind == 0:
                                                info_data = F"{val}"
                                                dut_num = re.search(r'DUT (\d+)', info_data)
                                                if dut_num:
                                                    sheet.cell(index+minus_count,21).value = int(dut_num.group(1))
                                            elif ind == 1:
                                                info_data = F"{val}"
                                                sheet.cell(index+minus_count,22).value = float(info_data)
                                            elif ind == 2:
                                                info_data = F"{val}"
                                                sheet.cell(index+minus_count,23).value = float(info_data) 
                                            elif ind == 3:
                                                info_data = F"{val}"
                                                sheet.cell(index+minus_count,24).value = float(info_data)
                                                if float(info_data) != 0:
                                                    sheet.cell(index+minus_count,24).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 4:
                                                info_data = F"{val}"
                                                sheet.cell(index+minus_count,25).value = info_data
                                                if info_data == 'fail':
                                                    sheet.cell(index+minus_count,25).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    if len(fail_item) <= 6:
                                        minus_count += 1
                                # --- merge ---
                                for i in range(1,7):
                                    sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                                    sheet.merge_cells(start_row=index,end_row=index+count-minus_count-1,start_column=i,end_column=i)
                            if rs485_result['fail'] > 0:
                                sheet.cell(index,6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                # --- create fail sheet ---
                                if f'F-{test_case}' not in wb.sheetnames and test_case in ['Bundle test(RS485 status)']:
                                    wb.create_sheet('F-{}'.format(test_case),index=100)
                                    set_title_for_sheet('Bundle mode',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                                    set_title_for_sheet('Bundle IDs',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                                    set_title_for_sheet('Check time',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                                    set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                                    set_title_for_sheet('RS485 error info by slot',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=5,end_column=5+len(slot_second_name)-1)
                                    set_title_for_sheet('RS485 error info',wb['F-{}'.format(test_case)],1,19,19,"5cb800")
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=19,end_column=19+len(sta_second_name)-1)
                                    # --- second row ---
                                    for i in range(0,len(slot_second_name)):
                                        set_title_for_sheet(slot_second_name[i],wb['F-{}'.format(test_case)],max_row+1,5+i,5+i,"5cb800")
                                    for i in range(0,len(sta_second_name)):
                                        set_title_for_sheet(sta_second_name[i],wb['F-{}'.format(test_case)],max_row+1,19+i,19+i,"5cb800")
                                    # --- merge ---
                                    for i in range(1,5):
                                        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                                fail_count = 0
                                for count, fail_item in enumerate(rs485_result['RS485 error status'],start=1):
                                    minus_count = 0
                                    if rs485_result['RS485 error status'][count-1]['result'] == 'fail':
                                        fail_count += 1
                                        index1 = wb['F-{}'.format(test_case)].max_row + 1
                                        for ind, (key,val) in enumerate(fail_item.items()):
                                            if len(fail_item) > 6:
                                                if ind == 0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,5).value = key
                                                    info_data = F"{val}"
                                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                                    slot_num = re.search(r'slot (\d+)', info_data)
                                                    slot_abc = re.search(r'slot ([a-zA-Z])', info_data)
                                                    if dut_num:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,6).value = int(dut_num.group(1))
                                                    if slot_num:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,7).value = int(slot_num.group(1))
                                                    if slot_abc:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,7).value = slot_abc.group(1)
                                                elif ind == 1:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,8).value = float(info_data)
                                                elif ind == 2:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,9).value = float(info_data) 
                                                elif ind == 3:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,10).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 4:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,11).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 5:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,12).value = float(info_data)
                                                elif ind == 6:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,13).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 7:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,14).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 8:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,15).value = float(info_data)
                                                elif ind == 9:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,16).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 10:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,17).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 11:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,18).value = info_data
                                                    if info_data == 'fail':
                                                        wb['F-{}'.format(test_case)].cell(index1-minus_count,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            else:
                                                if ind == 0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,19).value = key
                                                    info_data = F"{val}"
                                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                                    if dut_num:
                                                        wb['F-{}'.format(test_case)].cell(index1+minus_count,20).value = int(dut_num.group(1))
                                                elif ind == 1:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,21).value = float(info_data)
                                                elif ind == 2:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,22).value = float(info_data) 
                                                elif ind == 3:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,23).value = float(info_data)
                                                    if float(info_data) != 0.0:
                                                        wb['F-{}'.format(test_case)].cell(index1+minus_count,23).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                                elif ind == 4:
                                                    info_data = F"{val}"
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,24).value = info_data
                                                    if info_data == 'fail':
                                                        wb['F-{}'.format(test_case)].cell(index1+minus_count,24).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        if len(fail_item) <= 6:
                                            minus_count += 1
                                        wb['F-{}'.format(test_case)].cell(index1,1).value = vlan_mode
                                        wb['F-{}'.format(test_case)].cell(index1,2).value = bundle_ids
                                        wb['F-{}'.format(test_case)].cell(index1,3).value = action1
                                        wb['F-{}'.format(test_case)].cell(index1,4).value = action
                                # --- merge ---
                                for i in range(1,5):
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=index1-fail_count+1,end_row=index1,start_column=i,end_column=i)
                                # --- placed in center ---
                                alignment = Alignment(horizontal='center', vertical='center')
                                for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                                    for cell in row:
                                        cell.alignment = alignment
                                wb['F-{}'.format(test_case)].freeze_panes = 'E3'
                            # --- placed in center ---
                            alignment = Alignment(horizontal='center', vertical='center')
                            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                                for cell in row:
                                    cell.alignment = alignment
                            sheet.freeze_panes = 'G3'
        elif test_case == 'PDH ring':
            set_title_for_sheet('Action',sheet,max_row,1,1,"daeef3")
            set_title_for_sheet('Pass',sheet,max_row,2,1,"daeef3")
            set_title_for_sheet('Fail',sheet,max_row,3,1,"daeef3")
            set_title_for_sheet('Fail item',sheet,max_row,4,1,"daeef3")
            max_row = 2
            for result in test_result[test_case]:
                sheet.cell(max_row,1).value = result
                sheet.cell(max_row,2).value = test_result[test_case][result]['pass']
                sheet.cell(max_row,3).value = test_result[test_case][result]['fail']
                if len(test_result[test_case][result]['fail_item']) > 0:
                    append_data =""
                    for fail_item in test_result[test_case][result]['fail_item']:
                        for key in fail_item.keys():
                            append_data += key+":"+fail_item[key]+" "
                        append_data += '\n'
                    sheet.cell(max_row,4).alignment = Alignment(wrap_text=True)
                    sheet.cell(max_row,4).value = append_data
                max_row += 1
        elif test_case == 'Long Time(bert)':
            set_title_for_sheet('Round',sheet,1,1,1,"daeef3")
            set_title_for_sheet('Action',sheet,1,2,1,"daeef3")
            set_title_for_sheet('Bert type',sheet,1,3,1,"daeef3")
            set_title_for_sheet('Pass',sheet,1,4,1,"daeef3")
            set_title_for_sheet('Fail',sheet,1,5,1,"daeef3")
            for i in range(1,6):
                sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
            set_title_for_sheet('Ctrl Fail item',sheet,1,6,1,"daeef3")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=6,end_column=9)
            set_title_for_sheet('Time',sheet,2,6,1,"daeef3")
            set_title_for_sheet('Bit Error',sheet,2,7,1,"daeef3")
            set_title_for_sheet('Error ID',sheet,2,8,1,"daeef3")
            set_title_for_sheet('Elapsed Time',sheet,2,9,1,"daeef3")

            set_title_for_sheet('Test Card Fail item',sheet,1,10,1,"daeef3")
            set_title_for_sheet('Time',sheet,2,10,1,"daeef3")
            set_title_for_sheet('PRBS Status',sheet,2,11,1,"daeef3")
            set_title_for_sheet('Error Counts',sheet,2,12,1,"daeef3")
            set_title_for_sheet('Error ID',sheet,2,13,1,"daeef3")
            set_title_for_sheet('Error Seconds',sheet,2,14,1,"daeef3")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=10,end_column=14)
            
            max_row = 1
            for round, round_result in test_result[test_case].items():
                for bert_type1, bert_item1 in round_result.items():
                    for bert_type, bert_item in bert_item1.items():
                        index = sheet.max_row + 1
                        sheet.cell(index, 1).number_format = '0'
                        merge_row_title_for_sheet(int(round),sheet,index,1,index+bert_item['fail']-1,"FFFFFF")
                        merge_row_title_for_sheet(bert_type1,sheet,index,2,index+bert_item['fail']-1,"FFFFFF")
                        merge_row_title_for_sheet(bert_type,sheet,index,3,index+bert_item['fail']-1,"FFFFFF")
                        merge_row_title_for_sheet(bert_item['pass'],sheet,index,4,index+bert_item['fail']-1,"FFFFFF")
                        merge_row_title_for_sheet(bert_item['fail'],sheet,index,5,index+bert_item['fail']-1,"FFFFFF")
                        append_data = ''
                        if bert_item['fail'] > 0:
                            sheet.cell(index, 5).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            for count, fail_item in enumerate(bert_item['fail_item'],start=1):
                                for key,val in fail_item.items():
                                    append_data += F"{key}: {val} "
                                append_data += "\n"
                                # ----- create fail sheet ----- #
                                wb.create_sheet('F-{}'.format(test_case),index=100)
                                set_title_for_sheet('Test Item',wb['F-{}'.format(test_case)],1,1,1,"daeef3")
                                set_title_for_sheet('Round',wb['F-{}'.format(test_case)],1,2,2,"daeef3")
                                set_title_for_sheet('Bert type',wb['F-{}'.format(test_case)],1,3,3,"daeef3")
                                set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,4,4,"daeef3")
                                for i in range(1,5):
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                                set_title_for_sheet('Ctrl Fail item',wb['F-{}'.format(test_case)],1,5,1,"daeef3")
                                wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=5,end_column=8)
                                set_title_for_sheet('Time',wb['F-{}'.format(test_case)],2,5,1,"daeef3")
                                set_title_for_sheet('Bit Error',wb['F-{}'.format(test_case)],2,6,1,"daeef3")
                                set_title_for_sheet('Error ID',wb['F-{}'.format(test_case)],2,7,1,"daeef3")
                                set_title_for_sheet('Elapsed Time',wb['F-{}'.format(test_case)],2,8,1,"daeef3")
                                
                                set_title_for_sheet('Test Card Fail item',wb['F-{}'.format(test_case)],1,9,1,"daeef3")
                                set_title_for_sheet('Time',wb['F-{}'.format(test_case)],2,9,1,"daeef3")
                                set_title_for_sheet('PRBS Status',wb['F-{}'.format(test_case)],2,10,1,"daeef3")
                                set_title_for_sheet('Error Counts',wb['F-{}'.format(test_case)],2,11,1,"daeef3")
                                set_title_for_sheet('Error ID',wb['F-{}'.format(test_case)],2,12,1,"daeef3")
                                set_title_for_sheet('Error Seconds',wb['F-{}'.format(test_case)],2,13,1,"daeef3")
                                wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=9,end_column=13)

                                index2 = wb['F-{}'.format(test_case)].max_row + 1
                                for row_index2 in range(1,1+len(append_data.splitlines())):
                                    wb['F-{}'.format(test_case)].cell(index2,1).value = test_case
                                    wb['F-{}'.format(test_case)].cell(index2,2).value = int(round)
                                    wb['F-{}'.format(test_case)].cell(index2,3).value = bert_type
                                    wb['F-{}'.format(test_case)].cell(index2,4).value = bert_type1
                                    match1 = re.search(r'Time: (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) Bit Error: (\d+) Error ID: (\d+) Elapsed Time: (\d+)', append_data.splitlines()[row_index2-1])
                                    if match1:
                                        wb['F-{}'.format(test_case)].cell(index2,5).value = match1.group(1)
                                        wb['F-{}'.format(test_case)].cell(index2,6).value = int(match1.group(2))
                                        wb['F-{}'.format(test_case)].cell(index2,7).value = int(match1.group(3))
                                        wb['F-{}'.format(test_case)].cell(index2,8).value = int(match1.group(4))
                                    match2 = re.search(r'Time: (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) PRBSStatus: (\w+) ErrorCounts: (\d+) Error ID: (\d+) ErrorSeconds: (\d+)', append_data.splitlines()[row_index2-1])
                                    if match2:
                                        wb['F-{}'.format(test_case)].cell(index2,9).value = match2.group(1)
                                        wb['F-{}'.format(test_case)].cell(index2,10).value = match2.group(2)
                                        wb['F-{}'.format(test_case)].cell(index2,11).value = int(match2.group(3))
                                        wb['F-{}'.format(test_case)].cell(index2,12).value = int(match2.group(4))
                                        wb['F-{}'.format(test_case)].cell(index2,13).value = int(match2.group(5))
                            # --- placed in center ---
                            alignment = Alignment(horizontal='center', vertical='center')
                            for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                                for cell in row:
                                    cell.alignment = alignment
                            wb['F-{}'.format(test_case)].freeze_panes = 'E3'

                        if len(bert_item['fail_item']) > 0:
                            sheet.cell(index,6).alignment = Alignment(wrap_text=True)
                            for count, fail_item in enumerate(bert_item['fail_item'],start=1):
                                if len(fail_item) < 5:
                                    for key,val in fail_item.items():
                                        if key == "Time":
                                            sheet.cell(index+count-1,6).value = val
                                        if key == "Bit Error":
                                            sheet.cell(index+count-1,7).value = int(val)
                                        if key == "Error ID":
                                            sheet.cell(index+count-1,8).value = int(val)
                                        if key == "Elapsed Time":
                                            sheet.cell(index+count-1,9).value = int(val)
                                else:
                                    for key,val in fail_item.items():
                                        if key == "Time":
                                            sheet.cell(index+count-1,10).value = val
                                        if key == "PRBSStatus":
                                            sheet.cell(index+count-1,11).value = val
                                        if key == "ErrorCounts":
                                            sheet.cell(index+count-1,12).value = int(val)
                                        if key == "Error ID":
                                            sheet.cell(index+count-1,13).value = int(val)
                                        if key == "ErrorSeconds":
                                            sheet.cell(index+count-1,14).value = int(val)
            sheet.freeze_panes = 'F3'
            # --- placed in center ---
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        elif test_case == 'Long Time(PW status)':
            second_name = ['Time','DUT','CTRL', 'Bundle','Primary/Secondary', "Jitter buffer", "L bit count", "R bit count", 
                           "TX Good count","RX Good count","RX lost count", "Link", "Fail reason","Result"]
            jitter_name = ['Under run','Over run','Depth','MIN','MAX']
            set_title_for_sheet('Round',sheet,max_row,1,1,"daeef3")
            set_title_for_sheet('Action',sheet,max_row,2,2,"daeef3")
            set_title_for_sheet('Pass',sheet,max_row,3,3,"daeef3")
            set_title_for_sheet('Fail',sheet,max_row,4,4,"daeef3")
            set_title_for_sheet('Working PW',sheet,max_row,5,5,"daeef3")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=5,end_column=5+len(second_name)+len(jitter_name)-2)
            set_title_for_sheet('Standby PW',sheet,max_row,23,23,"daeef3")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=23,end_column=23+len(second_name)+len(jitter_name)-2)
            
            color_array = ["ffffff","d9d9d9"]
            # --- second row ---
            for i in range(0,6):
                set_title_for_sheet(second_name[i],sheet,max_row+1,5+i,5+i,"daeef3")
            for i in range(6,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,5+len(jitter_name)+i-1,5+len(jitter_name)+i-1,"daeef3")
            for i in range(0,6):
                set_title_for_sheet(second_name[i],sheet,max_row+1,23+i,23+i,"daeef3")
            for i in range(6,len(second_name)):
                set_title_for_sheet(second_name[i],sheet,max_row+1,23+len(jitter_name)+i-1,23+len(jitter_name)+i-1,"daeef3")
            # --- third row ---
            for i in range(len(jitter_name)):
                set_title_for_sheet(jitter_name[i],sheet,max_row+2,10+i,10+i,"daeef3")
            for i in range(len(jitter_name)):
                set_title_for_sheet(jitter_name[i],sheet,max_row+2,28+i,28+i,"daeef3")
      
            for round, round_result in test_result[test_case].items():
                for act, action_result in round_result.items():
                    if len(action_result['Working PW error status']) > 0:
                        index = sheet.max_row + 1
                        sheet.cell(index, 1).number_format = '0'
                        sheet.cell(index, 1).value = int(round)
                        sheet.cell(index, 2).value = act
                        sheet.cell(index, 3).value = action_result['pass']
                        sheet.cell(index, 4).value = action_result['fail']
                    
                        for count, fail_item in enumerate(action_result['Working PW error status'],start=1):
                            for ind, (key,val) in enumerate(fail_item.items()):
                                if ind == 0:
                                    sheet.cell(index+count-1,5).value = key
                                    info_data = F"{val}"
                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                    ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                    bundle_num = re.search(r'Bundle (\d+)', info_data)
                                    if dut_num:
                                        sheet.cell(index+count-1,6).value = float(dut_num.group(1))
                                    if ctrl_num:
                                        sheet.cell(index+count-1,7).value = float(ctrl_num.group(1))
                                    if bundle_num:
                                        sheet.cell(index+count-1,8).value = float(bundle_num.group(1))
                                    sheet.cell(index+count-1,9).value = (info_data.split())[-1]
                                elif ind == 1:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,10).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 2:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,11).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 3:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,12).value = info_data
                                    info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                    if info_data1:
                                        result = float(info_data1.group(1))
                                        if result > 12.0 or result < 4.0:
                                            sheet.cell(index+count-1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 4:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,13).value = float(info_data)
                                    if float(info_data) < 4.0:
                                        sheet.cell(index+count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 5:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,14).value = float(info_data)
                                    if float(info_data) > 20.0:
                                        sheet.cell(index+count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 6:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,15).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 7:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,16).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 8:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,17).value = float(info_data)
                                elif ind == 9:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,18).value = float(info_data)
                                elif ind == 10:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,19).value = float(info_data)
                                    if float(info_data) > 0.0:
                                        sheet.cell(index+count-1,19).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 11:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,20).value = info_data
                                    if info_data != "up":
                                        sheet.cell(index+count-1,20).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                if key == "fail_reason":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,21).value = info_data
                                    if info_data != '':
                                        sheet.cell(index+count-1,21).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif key == "result":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,22).value = info_data
                                    if info_data == 'fail':
                                        sheet.cell(index+count-1,22).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        # --- merge ---
                        for i in range(1,5):
                            sheet.merge_cells(start_row=1,end_row=3,start_column=i,end_column=i)
                            sheet.merge_cells(start_row=index,end_row=index+count-1,start_column=i,end_column=i)
                        for i in range(5,10):
                            sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        for i in range(15,28):
                            sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        for i in range(33,41):
                            sheet.merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                        sheet.merge_cells(start_row=2,end_row=2,start_column=10,end_column=14)
                        sheet.merge_cells(start_row=2,end_row=2,start_column=28,end_column=32)
                    if len(action_result['Standby PW error status']) > 0:
                        for count, fail_item in enumerate(action_result['Standby PW error status'],start=1):
                            for ind, (key,val) in enumerate(fail_item.items()):
                                if ind == 0:
                                    sheet.cell(index+count-1,23).value = key
                                    info_data = F"{val}"
                                    dut_num = re.search(r'DUT (\d+)', info_data)
                                    ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                    bundle_num = re.search(r'Bundle (\d+)', info_data)
                                    if dut_num:
                                        sheet.cell(index+count-1,24).value = float(dut_num.group(1))
                                    if ctrl_num:
                                        sheet.cell(index+count-1,25).value = float(ctrl_num.group(1))
                                    if bundle_num:
                                        sheet.cell(index+count-1,26).value = float(bundle_num.group(1))
                                    sheet.cell(index+count-1,27).value = (info_data.split())[-1]
                                elif ind == 1:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,28).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,28).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 2:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,29).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,29).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 3:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,30).value = info_data
                                    info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                    if info_data1:
                                        result = float(info_data1.group(1))
                                        if result > 12.0 or result < 4.0:
                                            sheet.cell(index+count-1,30).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 4:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,31).value = float(info_data)
                                    if float(info_data) < 4.0:
                                        sheet.cell(index+count-1,31).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 5:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,32).value = float(info_data)
                                    if float(info_data) > 20.0:
                                        sheet.cell(index+count-1,32).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 6:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,33).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,33).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 7:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,34).value = float(info_data)
                                    if float(info_data) != 0.0:
                                        sheet.cell(index+count-1,34).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 8:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,35).value = float(info_data)
                                elif ind == 9:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,36).value = float(info_data)
                                elif ind == 10:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,37).value = float(info_data)
                                    if float(info_data) > 0.0:
                                        sheet.cell(index+count-1,37).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif ind == 11:
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,38).value = info_data
                                    if info_data != "up":
                                        sheet.cell(index+count-1,38).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                if key == "fail_reason":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,39).value = info_data
                                    if info_data != '':
                                        sheet.cell(index+count-1,39).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                elif key == "result":
                                    info_data = F"{val}"
                                    sheet.cell(index+count-1,40).value = info_data
                                    if info_data == 'fail':
                                        sheet.cell(index+count-1,40).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                    if action_result['fail'] > 0:
                        sheet.cell(index, 4).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                        # --- create fail sheet ---
                        if f'F-{test_case}' not in wb.sheetnames and test_case in ['Long Time(PW status)']:
                            wb.create_sheet('F-{}'.format(test_case),index=100)
                            set_title_for_sheet('Round',wb['F-{}'.format(test_case)],1,1,1,"daeef3")
                            set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,2,2,"daeef3")
                            set_title_for_sheet('Working PW error info',wb['F-{}'.format(test_case)],1,3,3,"daeef3")
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=3,end_column=3+len(second_name)+len(jitter_name)-2)
                            set_title_for_sheet('Standby PW error info',wb['F-{}'.format(test_case)],1,21,21,"daeef3")
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=21,end_column=21+len(second_name)+len(jitter_name)-2)
                            # --- second row ---
                            for i in range(0,6):
                                set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,3+i,3+i,"daeef3")
                            for i in range(6,len(second_name)):
                                set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,3+len(jitter_name)+i-1,3+len(jitter_name)+i-1,"daeef3")
                            for i in range(0,6):
                                set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,21+i,21+i,"daeef3")
                            for i in range(6,len(second_name)):
                                set_title_for_sheet(second_name[i],wb['F-{}'.format(test_case)],max_row+1,21+len(jitter_name)+i-1,21+len(jitter_name)+i-1,"daeef3")
                            # --- third row ---
                            for i in range(len(jitter_name)):
                                set_title_for_sheet(jitter_name[i],wb['F-{}'.format(test_case)],max_row+2,8+i,8+i,"daeef3")
                            for i in range(len(jitter_name)):
                                set_title_for_sheet(jitter_name[i],wb['F-{}'.format(test_case)],max_row+2,26+i,26+i,"daeef3")
                            # --- merge ---
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=3,start_column=1,end_column=1)
                            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=3,start_column=2,end_column=2)
                            for i in range(3,8):
                                wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                            for i in range(13,26):
                                wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                            for i in range(31,39):
                                wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=3,start_column=i,end_column=i)
                            wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=2,start_column=8,end_column=12)
                            wb['F-{}'.format(test_case)].merge_cells(start_row=2,end_row=2,start_column=26,end_column=30)
                        fail_count = 0
                        for count, fail_item in enumerate(action_result['Working PW error status'],start=1):
                            if action_result['Working PW error status'][count-1]['result'] == 'fail':
                                fail_count += 1
                                index1 = wb['F-{}'.format(test_case)].max_row + 1
                                for ind, (key,val) in enumerate(fail_item.items()):
                                    wb['F-{}'.format(test_case)].cell(index1,1).value = float(round)
                                    wb['F-{}'.format(test_case)].cell(index1,2).value = act
                                    if ind == 0:
                                        wb['F-{}'.format(test_case)].cell(index1,3).value = key
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                        bundle_num = re.search(r'Bundle (\d+)', info_data)
                                        if dut_num:
                                            wb['F-{}'.format(test_case)].cell(index1,4).value = float(dut_num.group(1))
                                        if ctrl_num:
                                            wb['F-{}'.format(test_case)].cell(index1,5).value = float(ctrl_num.group(1))
                                        if bundle_num:
                                            wb['F-{}'.format(test_case)].cell(index1,6).value = float(bundle_num.group(1))
                                        wb['F-{}'.format(test_case)].cell(index1,7).value = (info_data.split())[-1]
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,8).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index1,8).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,9).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index1,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,10).value = info_data
                                        info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                        if info_data1:
                                            result = float(info_data1.group(1))
                                            if result > 12.0 or result < 4.0:
                                                wb['F-{}'.format(test_case)].cell(index1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,11).value = float(info_data)
                                        if float(info_data) < 4.0:
                                            wb['F-{}'.format(test_case)].cell(index1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,12).value = float(info_data)
                                        if float(info_data) > 20.0:
                                            wb['F-{}'.format(test_case)].cell(index1,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,13).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 7:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,14).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            wb['F-{}'.format(test_case)].cell(index1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 8:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,15).value = float(info_data)
                                    elif ind == 9:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,16).value = float(info_data)
                                    elif ind == 10:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,17).value = float(info_data)
                                        if float(info_data) > 0.0:
                                            wb['F-{}'.format(test_case)].cell(index1,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 11:
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,18).value = info_data
                                        if info_data != "up":
                                            wb['F-{}'.format(test_case)].cell(index1,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    if key == "fail_reason":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,19).value = info_data
                                    elif key == "result":
                                        info_data = F"{val}"
                                        wb['F-{}'.format(test_case)].cell(index1,20).value = info_data
                                if action_result['Standby PW error status'][count-1]['result'] == 'fail':
                                    fail_item = action_result['Standby PW error status'][count-1]
                                    for ind, (key,val) in enumerate(fail_item.items()):
                                        if ind == 0:
                                            wb['F-{}'.format(test_case)].cell(index1,21).value = key
                                            info_data = F"{val}"
                                            dut_num = re.search(r'DUT (\d+)', info_data)
                                            ctrl_num = re.search(r'CTRL-(\d+)', info_data)
                                            bundle_num = re.search(r'Bundle (\d+)', info_data)
                                            if dut_num:
                                                wb['F-{}'.format(test_case)].cell(index1,22).value = float(dut_num.group(1))
                                            if ctrl_num:
                                                wb['F-{}'.format(test_case)].cell(index1,23).value = float(ctrl_num.group(1))
                                            if bundle_num:
                                                wb['F-{}'.format(test_case)].cell(index1,24).value = float(bundle_num.group(1))
                                            wb['F-{}'.format(test_case)].cell(index1,25).value = (info_data.split())[-1]
                                        elif ind == 1:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,26).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                wb['F-{}'.format(test_case)].cell(index1,26).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 2:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,27).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                wb['F-{}'.format(test_case)].cell(index1,27).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 3:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,28).value = info_data
                                            info_data1 = re.search(r'(\d+\.\d+)', info_data)
                                            if info_data1:
                                                result = float(info_data1.group(1))
                                                if result > 12.0 or result < 4.0:
                                                    wb['F-{}'.format(test_case)].cell(index1,28).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 4:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,29).value = float(info_data)
                                            if float(info_data) < 4.0:
                                                wb['F-{}'.format(test_case)].cell(index1,29).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 5:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,30).value = float(info_data)
                                            if float(info_data) > 20.0:
                                                wb['F-{}'.format(test_case)].cell(index1,30).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 6:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,31).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                wb['F-{}'.format(test_case)].cell(index1,31).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 7:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,32).value = float(info_data)
                                            if float(info_data) != 0.0:
                                                wb['F-{}'.format(test_case)].cell(index1,32).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 8:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,33).value = float(info_data)
                                        elif ind == 9:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,34).value = float(info_data)
                                        elif ind == 10:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,35).value = float(info_data)
                                            if float(info_data) > 0.0:
                                                wb['F-{}'.format(test_case)].cell(index1,35).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        elif ind == 11:
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,36).value = info_data
                                            if info_data != "up":
                                                wb['F-{}'.format(test_case)].cell(index1,36).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        if key == "fail_reason":
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,37).value = info_data
                                        elif key == "result":
                                            info_data = F"{val}"
                                            wb['F-{}'.format(test_case)].cell(index1,38).value = info_data
                        # --- merge ---
                        wb['F-{}'.format(test_case)].merge_cells(start_row=index1-fail_count+1,end_row=index1,start_column=1,end_column=1)
                        wb['F-{}'.format(test_case)].merge_cells(start_row=index1-fail_count+1,end_row=index1,start_column=2,end_column=2)
                        # --- placed in center ---
                        alignment = Alignment(horizontal='center', vertical='center')
                        for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                            for cell in row:
                                cell.alignment = alignment
                        wb['F-{}'.format(test_case)].freeze_panes = 'C4'
                    # --- placed in center ---
                    alignment = Alignment(horizontal='center', vertical='center')
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.alignment = alignment
                    sheet.freeze_panes = 'E4'
        elif test_case == 'Long Time(RS485 status)':
            slot_second_name = ['Time','DUT','Slot','TX count', "RX count", "No RX count", "Timeout count", 
                           "Alive count","Non-alive count","Oversize count", "Init count", "Bad CMD count",
                            "Bad CMD num","Result"]
            sta_second_name = ['Time','DUT', 'FCS error', 'Slip error', 'Buffer error', 'Result']
            set_title_for_sheet('Round',sheet,max_row,1,1,"daeef3")
            set_title_for_sheet('Action',sheet,max_row,2,1,"daeef3")
            set_title_for_sheet('Pass',sheet,max_row,3,1,"daeef3")
            set_title_for_sheet('Fail',sheet,max_row,4,1,"daeef3")
            set_title_for_sheet('RS485 info by slot',sheet,max_row,5,1,"daeef3")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=5,end_column=5+len(slot_second_name)-1)
            set_title_for_sheet('RS485 info',sheet,max_row,19,1,"daeef3")
            sheet.merge_cells(start_row=max_row,end_row=max_row,start_column=19,end_column=19+len(sta_second_name)-1)
            for i in range(0,len(slot_second_name)):
                set_title_for_sheet(slot_second_name[i],sheet,max_row+1,5+i,5+i,"daeef3")
            for i in range(0,len(sta_second_name)):
                set_title_for_sheet(sta_second_name[i],sheet,max_row+1,19+i,19+i,"daeef3")
            for round, round_result in test_result[test_case].items():
                for rou, action_result in round_result.items():
                    if len(action_result['RS485 error status']) > 0:
                        index = sheet.max_row + 1
                        sheet.cell(index, 1).number_format = '0'
                        sheet.cell(index, 1).value = int(round)
                        sheet.cell(index, 2).value = rou
                        sheet.cell(index, 3).value = action_result['pass']
                        sheet.cell(index, 4).value = action_result['fail']
                        if action_result['fail'] > 0:
                            sheet.cell(index, 4).fill = PatternFill(start_color="e53b31", fill_type = "solid")

                        # --- create fail sheet ---
                            if f'F-{test_case}' not in wb.sheetnames and test_case in ['Long Time(RS485 status)']:
                                wb.create_sheet('F-{}'.format(test_case),index=100)
                                set_title_for_sheet('Round',wb['F-{}'.format(test_case)],1,1,1,"daeef3")
                                set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,2,2,"daeef3")
                                set_title_for_sheet('RS485 error info by slot',wb['F-{}'.format(test_case)],1,3,3,"daeef3")
                                wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=3,end_column=3+len(slot_second_name)-1)
                                set_title_for_sheet('RS485 error info',wb['F-{}'.format(test_case)],1,17,17,"daeef3")
                                wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=17,end_column=17+len(sta_second_name)-1)
                                # --- second row ---
                                for i in range(0,len(slot_second_name)):
                                    set_title_for_sheet(slot_second_name[i],wb['F-{}'.format(test_case)],max_row+1,3+i,3+i,"daeef3")
                                for i in range(0,len(sta_second_name)):
                                    set_title_for_sheet(sta_second_name[i],wb['F-{}'.format(test_case)],max_row+1,17+i,17+i,"daeef3")
                                # --- merge ---
                                for i in range(1,3):
                                    wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                            fail_count = 0
                            for count, fail_item in enumerate(action_result['RS485 error status'],start=1):
                                minus_count = 0
                                if action_result['RS485 error status'][count-1]['result'] == 'fail':
                                    fail_count += 1
                                    index1 = wb['F-{}'.format(test_case)].max_row + 1
                                    for ind, (key,val) in enumerate(fail_item.items()):
                                        if len(fail_item) > 6:
                                            if ind == 0:
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,3).value = key
                                                info_data = F"{val}"
                                                dut_num = re.search(r'DUT (\d+)', info_data)
                                                slot_num = re.search(r'slot (\d+)', info_data)
                                                slot_abc = re.search(r'slot ([a-zA-Z])', info_data)
                                                if dut_num:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,4).value = int(dut_num.group(1))
                                                if slot_num:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,5).value = int(slot_num.group(1))
                                                if slot_abc:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,5).value = slot_abc.group(1)
                                            elif ind == 1:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,6).value = float(info_data)
                                            elif ind == 2:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,7).value = float(info_data) 
                                            elif ind == 3:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,8).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,8).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 4:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,9).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 5:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,10).value = float(info_data)
                                            elif ind == 6:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,11).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 7:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,12).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,12).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 8:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,13).value = float(info_data)
                                            elif ind == 9:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,14).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 10:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,15).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,15).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 11:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1-minus_count,16).value = info_data
                                                if info_data == 'fail':
                                                    wb['F-{}'.format(test_case)].cell(index1-minus_count,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                        else:
                                            if ind == 0:
                                                wb['F-{}'.format(test_case)].cell(index1+minus_count,17).value = key
                                                info_data = F"{val}"
                                                dut_num = re.search(r'DUT (\d+)', info_data)
                                                if dut_num:
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,18).value = int(dut_num.group(1))
                                            elif ind == 1:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1+minus_count,19).value = float(info_data)
                                            elif ind == 2:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1+minus_count,20).value = float(info_data) 
                                            elif ind == 3:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1+minus_count,21).value = float(info_data)
                                                if float(info_data) != 0.0:
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,21).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                            elif ind == 4:
                                                info_data = F"{val}"
                                                wb['F-{}'.format(test_case)].cell(index1+minus_count,22).value = info_data
                                                if info_data == 'fail':
                                                    wb['F-{}'.format(test_case)].cell(index1+minus_count,22).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    if len(fail_item) <= 6:
                                        minus_count += 1
                                    wb['F-{}'.format(test_case)].cell(index1,1).value = int(round)
                                    wb['F-{}'.format(test_case)].cell(index1,2).value = rou
                            # --- merge ---
                            # for i in range(1,3):
                            #     wb['F-{}'.format(test_case)].merge_cells(start_row=index1-fail_count+1,end_row=index1,start_column=i,end_column=i)
                            # --- placed in center ---
                            alignment = Alignment(horizontal='center', vertical='center')
                            for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
                                for cell in row:
                                    cell.alignment = alignment
                            wb['F-{}'.format(test_case)].freeze_panes = 'C3'

                    if len(action_result['RS485 error status']) > 0:
                        info_data = ''
                        minus_count = 0
                        for count, fail_item in enumerate(action_result['RS485 error status'],start=1):
                            for ind, (key,val) in enumerate(fail_item.items()):
                                if len(fail_item) > 6:
                                    if ind == 0:
                                        sheet.cell(index+count-minus_count-1,5).value = key
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        slot_num = re.search(r'slot (\d+)', info_data)
                                        slot_abc = re.search(r'slot ([a-zA-Z])', info_data)
                                        if dut_num:
                                            sheet.cell(index+count-minus_count-1,6).value = int(dut_num.group(1))
                                        if slot_num:
                                            sheet.cell(index+count-minus_count-1,7).value = int(slot_num.group(1))
                                        if slot_abc:
                                            sheet.cell(index+count-minus_count-1,7).value = slot_abc.group(1)
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,8).value = float(info_data)
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,9).value = float(info_data) 
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,10).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+count-minus_count-1,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,11).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+count-minus_count-1,11).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 5:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,12).value = float(info_data)
                                    elif ind == 6:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,13).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+count-minus_count-1,13).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 7:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,14).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+count-minus_count-1,14).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 8:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,15).value = float(info_data)
                                    elif ind == 9:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,16).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+count-minus_count-1,16).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 10:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,17).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+count-minus_count-1,17).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 11:
                                        info_data = F"{val}"
                                        sheet.cell(index+count-minus_count-1,18).value = info_data
                                        if info_data == 'fail':
                                            sheet.cell(index+count-minus_count-1,18).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                else:
                                    if ind == 0:
                                        sheet.cell(index+minus_count,19).value = key
                                        info_data = F"{val}"
                                        dut_num = re.search(r'DUT (\d+)', info_data)
                                        if dut_num:
                                            sheet.cell(index+minus_count,20).value = int(dut_num.group(1))
                                    elif ind == 1:
                                        info_data = F"{val}"
                                        sheet.cell(index+minus_count,21).value = float(info_data)
                                    elif ind == 2:
                                        info_data = F"{val}"
                                        sheet.cell(index+minus_count,22).value = float(info_data) 
                                    elif ind == 3:
                                        info_data = F"{val}"
                                        sheet.cell(index+minus_count,23).value = float(info_data)
                                        if float(info_data) != 0.0:
                                            sheet.cell(index+minus_count,23).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                    elif ind == 4:
                                        info_data = F"{val}"
                                        sheet.cell(index+minus_count,24).value = info_data
                                        if info_data == 'fail':
                                            sheet.cell(index+minus_count,24).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            if len(fail_item) <= 6:
                                minus_count += 1
                        # --- merge ---
                        for i in range(1,5):
                            sheet.merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
                            sheet.merge_cells(start_row=index,end_row=index+count-minus_count-1,start_column=i,end_column=i)
                    # --- placed in center ---
                    alignment = Alignment(horizontal='center', vertical='center')
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.alignment = alignment
                    sheet.freeze_panes = 'E3'
        elif (test_case == 'Bert'or test_case == '50ms') and len(list(test_result[test_case]["Items"].keys())) > 0 :
            if test_case == '50ms':
                # card_title = ['Card type','Frame type','Action', 'Pass', 'Fail', 'Pass Rate', 'Fail info']
                merge_row_title_for_sheet("Card type",sheet,1,1,2,'5cb800')
                merge_row_title_for_sheet("Frame type",sheet,1,2,2,'5cb800')
                merge_row_title_for_sheet("Action",sheet,1,3,2,'5cb800')
                set_title_for_sheet('PW setting',sheet,1,4,5,"5cb800")
                set_title_for_sheet('NumOfFrame',sheet,2,4,4,"5cb800")
                # Should use Jitter depth from PW status, can deleye jitter delay and size display
                set_title_for_sheet('JitterDepth',sheet,2,5,5,"5cb800")
                set_title_for_sheet('Test result',sheet,1,6,12,"5cb800")
                set_title_for_sheet('Port A Result',sheet,2,6,6,"5cb800")
                set_title_for_sheet('Port A switch time(us)',sheet,2,7,7,"5cb800")
                set_title_for_sheet('Port A delay time(us)',sheet,2,8,8,"5cb800")
                
                set_title_for_sheet('Port B Result',sheet,2,9,9,"5cb800")
                set_title_for_sheet('Port B switch time(us)',sheet,2,10,10,"5cb800")
                set_title_for_sheet('Port B delay time(us)',sheet,2,11,11,"5cb800")
                set_title_for_sheet('Timestamp',sheet,2,12,12,"5cb800")

                start_row = 3
            else:
                for x, item in enumerate(card_title):
                    sheet.cell(max_row,x+1).value = item
                    sheet.cell(max_row,x+1).font = Font(bold=True)
                    sheet.cell(max_row,x+1).fill = PatternFill("solid", start_color="5cb800")
                start_row = 2
            row_minus = 0
            for row_index, item in enumerate(list(sorted(test_result[test_case]["Items"].keys())),start=start_row):
                pass_cnt = test_result[test_case]["Items"][item]['pass']
                fail_cnt = test_result[test_case]["Items"][item]['fail']
                if test_case == 'Bert':
                    fail_items = test_result[test_case]["Items"][item]['fail_item']
                elif test_case == '50ms':
                    pw_setting = test_result[test_case]["Items"][item]['pw_setting']
                    switch_test_result = test_result[test_case]["Items"][item]['test_result']
                if pass_cnt + fail_cnt == 0:
                    continue
                elif pass_cnt + fail_cnt > 0:
                    if test_case == 'Bert':
                        card_info_array = item.split('-')
                        # print(card_info_array)
                        sheet.cell(row_index,1).value = card_info_array[0]
                        sheet.cell(row_index,2).value = card_info_array[1]
                        sheet.cell(row_index,3).value = card_info_array[2]
                        sheet.cell(row_index,4).number_format = '0'
                        sheet.cell(row_index,5).number_format = '0'
                        sheet.cell(row_index,6).number_format = '0'
                        sheet.cell(row_index,4).value = pass_cnt
                        sheet.cell(row_index,5).value = fail_cnt
                        if pass_cnt != 0:
                            sheet.cell(row_index,6).value = "{0:.1%}".format(pass_cnt / (pass_cnt+fail_cnt))
                        else:
                            sheet.cell(row_index,6).value = 0
                        if len(fail_items) > 0:
                            append_data = ""
                            for fail_item in fail_items:
                                if type(fail_item) == dict:
                                    for key in fail_item.keys():
                                        if key == "Fireberd error":
                                            append_data += fail_item[key]+" "
                                        else:
                                            append_data += key+":"+fail_item[key]+" "
                                        # print("======"+append_data+"======")
                                    append_data += '\n'
                                elif type(fail_item) == str:
                                    append_data += fail_item
                            sheet.cell(row_index,7).alignment = Alignment(wrap_text=True)
                            sheet.cell(row_index,7).value = append_data
                            create_fail_info_sheet(wb,test_case,card_info_array[0],card_info_array[1],card_info_array[2],append_data)
                    elif test_case == '50ms':
                        row_index = sheet.max_row + 1
                        for i in range((pass_cnt + fail_cnt)):
                            # print(i)
                            card_info_array = item.split('-')
                            # print(card_info_array)
                            sheet.cell(row_index,1).value = card_info_array[0]
                            sheet.cell(row_index,2).value = card_info_array[1]
                            sheet.cell(row_index,3).value = card_info_array[2]
                            sheet.cell(row_index,4).number_format = '0'
                            sheet.cell(row_index,5).number_format = '0'
                            sheet.cell(row_index,4).value = pw_setting[i][0]['NumofFrame']
                            sheet.cell(row_index,5).value = pw_setting[i][0]['JitterDepth']
                            # if pw_setting[i] == []:
                            #     sheet.cell(row_index,4).value = pw_setting[i-1][0]['NumofFrame']
                            #     sheet.cell(row_index,5).value = pw_setting[i-1][0]['JitterDepth']
                            #     sheet.cell(row_index,6).value = pw_setting[i-1][0]['JitterSize']
                            # else:
                            #     sheet.cell(row_index,4).value = pw_setting[i][0]['NumofFrame']
                            #     sheet.cell(row_index,5).value = pw_setting[i][0]['JitterDelay']
                            #     sheet.cell(row_index,6).value = pw_setting[i][0]['JitterSize']
                            sheet.cell(row_index,6).value = switch_test_result[i]['Port A Result']
                            if switch_test_result[i]['Port A Result'] == 'Fail':
                                sheet.cell(row_index,6).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            sheet.cell(row_index,7).value = switch_test_result[i]['Port A SwitchTime']
                            if float(switch_test_result[i]['Port A SwitchTime']) > 50000:
                                sheet.cell(row_index,7).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            sheet.cell(row_index,8).value = switch_test_result[i]['Port A Delaytime']
                            sheet.cell(row_index,9).value = switch_test_result[i]['Port B Result']
                            sheet.cell(row_index,10).value = switch_test_result[i]['Port B SwitchTime']
                            if switch_test_result[i]['Port B SwitchTime'] == 'los':
                                sheet.cell(row_index,9).value = "n/a"
                                sheet.cell(row_index,10).value = "n/a"
                            else:
                                if float(switch_test_result[i]['Port B SwitchTime']) > 50000:
                                    sheet.cell(row_index,10).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                                if switch_test_result[i]['Port B Result'] == 'Fail':
                                    sheet.cell(row_index,9).fill = PatternFill(start_color="e53b31", fill_type = "solid")
                            sheet.cell(row_index,11).value = switch_test_result[i]['Port B Delaytime']
                            sheet.cell(row_index,12).value = switch_test_result[i]['TimeStamp']
                            row_index += 1

        # --- RS232 stutas ---
        elif (test_case == 'RS232 Status'):
            RS232_title = ['Rate', 'Async', 'Pass', 'Fail']
            for x, item in enumerate(RS232_title):
                sheet.cell(max_row,x+1).value = item
                sheet.cell(max_row,x+1).font = Font(bold=True)
                sheet.cell(max_row,x+1).fill = PatternFill("solid", start_color="5cb800")
                sheet.cell(max_row,x+1).alignment = Alignment(horizontal='center', vertical='center') 
            max_row = sheet.max_row + 1
            for row_index, item in enumerate(list(sorted(test_result[test_case]["Items"].keys())),start=2):
                pass_cnt = test_result[test_case]["Items"][item]['pass']
                fail_cnt = test_result[test_case]["Items"][item]['fail']
                if pass_cnt + fail_cnt > 0:
                    sheet.cell(row_index,1).value = item.split('_')[0]
                    sheet.cell(row_index,2).value = int(item.split('_')[1])
                    sheet.cell(row_index,3).value = pass_cnt
                    sheet.cell(row_index,4).value = fail_cnt
        # --- RS232 stutas ---
        # ------ voice test for cisco-------- 
        elif (test_case == 'Voice Bert' or test_case == 'Tone' or test_case == 'Ring'):
            for x, item1 in enumerate(cisco_voice_title):
                sheet.cell(max_row,x+1).value = item1
                sheet.cell(max_row,x+1).font = Font(bold=True)
                sheet.cell(max_row,x+1).fill = PatternFill("solid", start_color="5cb800")
                sheet.cell(max_row,x+1).alignment = Alignment(horizontal='center', vertical='center') 
            max_row = sheet.max_row + 1
            row_cnt = 1
            for row_index, item in enumerate(list(sorted(test_result[test_case]["Items"].keys())),start=2):
                for row_index2, item2 in enumerate(list(sorted(test_result[test_case]["Items"][item].keys())),start=2):
                    row_cnt += 1
                    pass_cnt = test_result[test_case]["Items"][item][item2]['pass']
                    fail_cnt = test_result[test_case]["Items"][item][item2]['fail']
                    fail_items = test_result[test_case]["Items"][item][item2]['fail_item']
                    if pass_cnt + fail_cnt > 0:
                        round_info = item.split('-')[0]
                        action_info = item.split('-')[1]
                        stage_info = item.split('-')[2]
                        merge_row_title_for_sheet(round_info,sheet,4*(row_index-1)-2,1,4*(row_index-1)+1,'FFFFFF')
                        merge_row_title_for_sheet(action_info,sheet,4*(row_index-1)-2,2,4*(row_index-1)+1,'FFFFFF')
                        merge_row_title_for_sheet(stage_info,sheet,4*(row_index-1)-2,3,4*(row_index-1)+1,'FFFFFF')               
                        card_info_array = item2.split('-')
                        au_ohm = item2.split('_')
                        sheet.cell(row_cnt,4).value = au_ohm[0]
                        sheet.cell(row_cnt,5).value = au_ohm[1]

                        sheet.cell(row_cnt,6).number_format = '0'
                        sheet.cell(row_cnt,7).number_format = '0'

                        sheet.cell(row_cnt,6).value = pass_cnt
                        sheet.cell(row_cnt,7).value = fail_cnt

                        for a in range(1,9):
                            sheet.cell(row_cnt,a).alignment = Alignment(horizontal='center', vertical='center')                   

                        if pass_cnt != 0:
                            sheet.cell(row_cnt,8).value = "{0:.1%}".format(pass_cnt / (pass_cnt+fail_cnt))
                        else:
                            sheet.cell(row_cnt,8).value = 0
                        
                        if len(fail_items) > 0:
                            append_data = ""
                            for fail_item in fail_items:
                                if type(fail_item) == dict:
                                    for key in fail_item.keys():
                                        if key == "Fireberd error":
                                            append_data += fail_item[key]+" "
                                        else:
                                            append_data += key+":"+fail_item[key]+" "
                                        # print("======"+append_data+"======")
                                    append_data += '\n'
                                elif type(fail_item) == str:
                                    append_data += fail_item.rstrip().strip() + "\n"

                            append_data = append_data.rstrip()
                            append_data = append_data.strip()                                
                            sheet.cell(row_cnt,9).alignment = Alignment(wrap_text=True)
                            sheet.cell(row_cnt,9).value = append_data
                            if (('F-{}'.format(test_case)) not in wb.sheetnames) and (test_case == 'Voice Bert' or test_case == 'Tone' or test_case == 'Ring'):
                                wb.create_sheet('F-{}'.format(test_case),index=100)
                                set_title_for_sheet('Test case',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
                                set_title_for_sheet('Round',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
                                set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
                                set_title_for_sheet('Stage',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
                                set_title_for_sheet('au-law',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
                                set_title_for_sheet('Impedance',wb['F-{}'.format(test_case)],1,6,6,"5cb800")
                                set_title_for_sheet('Fail info',wb['F-{}'.format(test_case)],1,7,7,"5cb800")
                                if not hasattr(create_voice_fail_info_sheet, 'call_counter'):
                                    create_voice_fail_info_sheet.call_counter = {}
                            create_voice_fail_info_sheet.call_counter[test_case] = create_voice_fail_info_sheet.call_counter.get(test_case, 2)
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],1).value = test_case
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],2).value = round_info
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],3).value = action_info
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],4).value = stage_info
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],5).value = au_ohm[0]
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],6).value = au_ohm[1]
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],7).value = append_data
                            wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],7).alignment = Alignment(wrap_text=True)
                            create_voice_fail_info_sheet.call_counter[test_case] += 1
                            wb['F-{}'.format(test_case)].freeze_panes = 'A2'
                            # --- placed in center ---
                            alignment = Alignment(horizontal='center', vertical='center')
                            for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=4):
                                for cell in row:
                                    cell.alignment = alignment
            sheet.freeze_panes = 'A2'
        # ------ voice test for cisco-------- 
        print("Done!")

    max_row = s1.max_row
    for x, item in enumerate(summary_title):
        # s1.cell(max_row,x+1).value = item
        set_title_for_sheet(item,s1,max_row,x+1,1,"5cb800")
        s1.cell(max_row,x+1).font= Font(name="Calibri")
        # s1.cell(max_row,x+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    max_row = s1.max_row + 1

    pcnt = 0
    fcnt = 0
    wcnt = 0

    test_category = {
        'CTRL Bert': {'warmreset':{}, 'coldreset(others)':{}, 'Ctrl1Eth1DisEn':{}, 'powereset':{}, 'systemreset':{}, 'Check setup status':{}, 'coldreset':{}}, 
        'IP6704':{'50ms':{}},
        'PW' : {'PW Status':{}, 'PW Statist':{}}, 
        'RS485': {'RS485 Status':{}}, 
        'DPLL' : {'DPLL Status':{}}, 
        'Card' : {'Card port status':{}}, 
        'QE1 Bert' : {'Long Time Bert':{}},
        'TSI in out' : {'Basic TSI in out(Signalling)':{},'Basic TSI in out(Data)':{},'Cross TSI in out(Signalling)':{},'Cross TSI in out(Data)':{},'TSI in out(AU_Law)':{}}, 
        'Voice' : {'Tone':{},'Ring':{},'Voice Bert':{},'PW status(Voice test)':{},'RS485 status(Voice test)':{}}, 
        'Bundle Test': {'Bundle test(bert)':{}, 'Bundle test(PW status)':{}, 'Bundle test(RS485 status)':{}},
        'Long Time': {'Long Time(bert)':{}, 'Long Time(PW status)':{}, 'Long Time(RS485 status)':{}},
        'MIB Scan' : {'MIB Scan':{}}, 
        'Firmware' : {'Firmware test':{}}
        }
    category_exist = set()
    for x, item in enumerate(list(test_summary.keys())):
        if test_summary[item]['pass'] != 0 or test_summary[item]['fail'] != 0 or test_summary[item]['warning'] != 0:
            for x_s, item_s in enumerate(list(test_category.keys())):
                if item in test_category[item_s]:
                    test_category[item_s][item].update(test_summary[item])
                    category_exist.add(item_s)

    for x, item in enumerate(list(test_category.keys())):
        if item in category_exist:
            max_row = s1.max_row + 1
            row_ind = max_row
            for x_s, item_s in enumerate(list(test_category[item].keys())):
                if test_category[item][item_s] != {}:
                    if test_category[item][item_s]['pass'] != 0 or test_category[item][item_s]['fail'] !=0 or test_category[item][item_s]['warning'] !=0 or \
                    item_s == 'MIB Scan' or item_s == 'Firmware test' or item_s == 'Alarm Trap check' or item_s =='MIB Diff':
                        s1.cell(max_row,2).value = item_s
                        # print(item)
                        link = F"{ntpath.basename(excel_path)}#'{test_category[item][item_s]['sheet link']}'!A1"
                        # print(link)
                        s1.cell(max_row,2).hyperlink = link
                        s1.cell(max_row,2).font= Font(name="Calibri",underline='single',color='0563C1')
                        # s1.cell(max_row,1).style= "Hyperlink"
                        s1.cell(max_row,3).number_format = '0'
                        s1.cell(max_row,3).value = test_category[item][item_s]['pass']
                        s1.cell(max_row,3).font= Font(name="Calibri")
                        s1.cell(max_row,4).number_format = '0'
                        s1.cell(max_row,4).value = test_category[item][item_s]['fail']
                        s1.cell(max_row,4).font= Font(name="Calibri")
                        s1.cell(max_row,5).number_format = '0'
                        s1.cell(max_row,5).value = test_category[item][item_s]['warning']
                        s1.cell(max_row,5).font= Font(name="Calibri")
                        # if item_s == 'RS485 Status':
                        #     s1.cell(max_row,4).number_format = '0'
                        #     s1.cell(max_row,4).value = ''
                        #     s1.cell(max_row,4).font= Font(name="Calibri")
                        #     s1.cell(max_row,5).number_format = '0'
                        #     s1.cell(max_row,5).value = test_category[item][item_s]['fail']
                        #     s1.cell(max_row,5).font= Font(name="Calibri")
            # --- 超連結 --- #
                        if s1.cell(max_row,4).value != 0:
                            link_fail = F"{ntpath.basename(excel_path)}#'{'F-{}'.format(test_category[item][item_s]['sheet link'])}'!A1"
                            s1.cell(max_row,4).hyperlink = link_fail
                            s1.cell(max_row,4).font= Font(name="Calibri",underline='single',color='0563C1')
            # --- 超連結 --- #
                        if s1.cell(max_row,5).value != 0:
                            link_fail = F"{ntpath.basename(excel_path)}#'{'F-{}'.format(test_category[item][item_s]['sheet link'])}'!A1"
                            s1.cell(max_row,5).hyperlink = link_fail
                            s1.cell(max_row,5).font= Font(name="Calibri",underline='single',color='0563C1')
                        pcnt += test_category[item][item_s]['pass']
                        fcnt += test_category[item][item_s]['fail']
                        wcnt += test_category[item][item_s]['warning']
                        max_row += 1
            merge_row_title_for_sheet(item,s1,row_ind,1,s1.max_row,"FFFFFF")
            s1.cell(row_ind,1).font= Font(name="Calibri")
# --- 增加DUT在Summary sheet --- #
    max_column = s1.max_column + 1

    Summary_for_DUT_first_row = max_row
    Summary_for_DUT_row = max_row + 1
  
    set_title_for_sheet('Category',s1,Summary_for_DUT_row,1,1,"5cb800")
    set_title_for_sheet('Item',s1,Summary_for_DUT_row,2,2,"5cb800")
    merge_row_title_for_sheet('Info',s1,Summary_for_DUT_row+1,1,Summary_for_DUT_row+1,"FFFFFF")
    merge_row_title_for_sheet('DUTs',s1,Summary_for_DUT_row+1,2,Summary_for_DUT_row+1,"FFFFFF")
    s1.cell(row=Summary_for_DUT_row+1, column=2).hyperlink = f"{ntpath.basename(excel_path)}#'{wb.sheetnames[1]}'!A1"

    thin = Side(border_style="thin", color="000000")
    for i in range(Summary_for_DUT_first_row,Summary_for_DUT_row+2):
        for j in range(1,3):
            s1.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            s1.cell(i,j).font= Font(name="Calibri")
    s1.cell(Summary_for_DUT_row+1,2).font= Font(name="Calibri",underline='single',color='0563C1')
# --- 增加DUT在Summary sheet --- #    
    
    thin = Side(border_style="thin", color="000000")
    for i in range(1,max_row):
        for j in range(1,max_column):
            s1.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
# --- 每個sheet超連結 ---
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        current_sheet = wb[sheet_name]
        current_sheet.cell(row=1, column=1).hyperlink = f"{ntpath.basename(excel_path)}#'{wb.sheetnames[0]}'!A1"
# --- 每個sheet超連結 ---

    # data  = [pcnt,fcnt]
    # label = 'Pass', 'Fail'
    # if pcnt+fcnt > 0:
    #     create_pie_chart_from_data_label(data,label,'Rate',F'./Summary/summary_pie_{title}.png')

    max_row = s4.max_row
    for x, item in enumerate(list(dut_info.keys()),start=1):
        row = 1
        start_column = x*2-1
        end_column = x*2
        set_title_for_sheet(F"DUT-{item} {dut_info[item]['Chassis']}",s4,row,start_column,end_column,"ffc000")
        # s4.cell(row,s4.max_column+1).value = ""
        s4.cell(row+1,start_column).value = F"Chassis"
        s4.cell(row+1,end_column).value = F"{dut_info[item]['Chassis']}"
        s4.cell(row+2,start_column).value = F"Slots"
        s4.cell(row+2,end_column).value = F"{dut_info[item]['Total slots']}"
        row += 3
        # print(dut_info[item]['Slot Info'])
        for slot_item in dut_info[item]['Slot Info']:
            for val in list(slot_item.keys()):
                if 'Name' == val:
                    set_title_for_sheet(F"{slot_item[val]}",s4,row,start_column,end_column,"b8cce4")
                else:
                    s4.cell(row,start_column).value = val
                    s4.cell(row,end_column).value = slot_item[val]
                row+=1
    #copy_mr_sheet(wb,"D:\Jason-Huang\RPM20230511.xlsx")
    print('{:-<30}'.format("Formatting"),end="")
    for sheet in wb.sheetnames:
        if sheet != 'Summary':
            # print(sheet+',')
            change_font_for_sheet(wb[sheet],font='Calibri')
        adjust_column_width(wb[sheet])
        if sheet != 'Summary':
            add_border_in_excel(wb[sheet])
    print('Done!')


    wb.save(excel_path)
    # if 'Bert' in wb.sheetnames:
    #     create_pie_chart_from_sheet(excel_path, "Bert",card_title,F"./Summary/Bert_test_pie_{title}.png")


    # data = {
    #     'Time':list(test_summary['coldreset']['50ms sort data'].keys()),
    #     'Count':list(test_summary['coldreset']['50ms sort data'].values())
    # }
    # print(data)
    # create_pie_chart_from_data(data,'Time','Time','Count','./Summary/time_pie.png')
    # Html usage

    # add_style_to_html('./Summary/Test cases')
    # add_test_description_to_html('./Summary/Test cases','./Summary/test_description')
    # append_img_to_html('./Summary/Summary.html','./summary_pie.png')
    # append_img_to_html('./Summary/Action.html','./action_pie.png')
    # append_img_to_html('./Summary/Case.html','./test_description/autest.png')

    # if pcnt+fcnt > 0:
    #     append_img_to_excel(s1,'F2',F'./Summary/summary_pie_{title}.png')
    # if "Bert" in wb.sheetnames:
    #     append_img_to_excel(wb["Bert"],'H2',F'./Summary/Bert_test_pie_{title}.png')
    # if "Bundle test" in wb.sheetnames:
    #     append_img_to_excel(wb["Bundle test"],'J3','./Summary/test_description/Bundle64.jpg')
    # if "PDH ring" in wb.sheetnames:
    #     append_img_to_excel(wb["PDH ring"],'A8','./Summary/test_description/PDH ring.jpg')
    # if "Long Time" in wb.sheetnames:
    #     append_img_to_excel(wb["Long Time"],'I2','./Summary/test_description/LongTime.jpg')
    # wb.save(excel_path)

def get_card_config_action(info):
    if "init" in info:
        card = ""
        config_type = ""
        card_action = info
    elif "after action" in info or "before action" in info or "Reget action" in info :
        if "after action" in info:
            action = "After"
        elif "before action" in info:
            action = "Before"
        elif "Reget action" in info:
            action = "Reget after"
        card_info = re.sub(r"(record after action-|record before action-|Reget action-)","", info)
        # print(card_info)
        card = card_info.split('-')[0]
        config_type = card_info.split('-')[1]
        card_action = action + " " +card_info.split('-')[2]
    elif "Check setup status" in info:
        card_info = re.sub(r"- ","", info)
        card = card_info.split('-')[0]
        config_type = card_info.split('-')[1]
        card_action = card_info.split('-')[2]
    else:
        card = info.split('-')[0]
        config_type = info.split('-')[1]
        card_action = info.split('-')[2]
    return card,config_type,card_action

def set_title_for_sheet(name,sheet,row,col,end_col,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_col-col >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=row,start_column=col,end_column=end_col)

def merge_row_title_for_sheet(name,sheet,row,col,end_row,color,merge=True):
    sheet.cell(row,col).value = name
    sheet.cell(row,col).font = Font(bold=True)
    sheet.cell(row,col).alignment = Alignment(horizontal='center',vertical='center')
    sheet.cell(row,col).fill = PatternFill("solid", start_color=color)
    if end_row-row >= 1 and merge:
        sheet.merge_cells(start_row=row,end_row=end_row,start_column=col,end_column=col)

def add_border_in_excel(sheet):
    thin = Side(border_style="thin", color="000000")
    max_row = sheet.max_row
    max_column  = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            # for merged_cell in sheet.merged_cells.ranges:
            #     if sheet.cell(i,j).coordinate in merged_cell:
            #         sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet.cell(i,j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
def change_font_for_sheet(sheet, font):
    # print('{:-<30}'.format("Change font"),end="")
    # DEFAULT_FONT.name = "Calibri"
    max_row = sheet.max_row
    max_column  = sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            # for merged_cell in sheet.merged_cells.ranges:
            #     if sheet.cell(i,j).coordinate in merged_cell:
            #         sheet.cell(i,j).font = Font(name=font)
            sheet.cell(i,j).font = Font(name=font)
    # print('Done!')
def append_img_to_excel(sheet,img_location,img_path):
    img = openpyxl.drawing.image.Image(img_path)
    img.anchor = img_location
    sheet.add_image(img)


def adjust_column_width(sheet):
    new_column_length = 0
    # print('{:-<30}'.format("Adjust width"),end="")
    for column_cells in sheet.columns:
        # for cell in column_cells:
        #     if str(cell.value).find('\n') != -1:
        #         new_column_length = str(cell.value).find('\n')
        # if new_column_length == 0 :
        break_line_count = max(str(cell.value).count('\n') for cell in column_cells)
        # print(break_line_count)

        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        if break_line_count > 0 :
            new_column_length /= break_line_count
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            sheet.column_dimensions[new_column_letter].width = new_column_length*1.23
    # print('Done!')
def copy_mr_sheet(wb,mr_excel_path):
    wb_mr = openpyxl.load_workbook(filename=mr_excel_path)
    ws_mr = wb_mr['MR']
    ws_summary = wb.create_sheet(ws_mr.title,1)
    for row in ws_mr:
        for cell in row:
            if cell.row == 1:
                ws_summary[cell.coordinate].fill = PatternFill("solid", start_color='9bc2e6')
            if cell.column == 3 and cell.row > 1:
                ws_summary[cell.coordinate].number_format = 'yyyy/mm/dd'
            ws_summary[cell.coordinate].value = cell.value

# ---------- fail info sheet ---------

def create_fail_info_sheet(wb,test_case,card,config_type,card_action,append_data):
    if f'F-{test_case}' not in wb.sheetnames and test_case not in ['Voice Bert', 'Tone', 'Ring']:
        wb.create_sheet('F-{}'.format(test_case),index=100)
        set_title_for_sheet('Test Item',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
        set_title_for_sheet('Card',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
        set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
        set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
        set_title_for_sheet('Fail info',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
        
        wb['F-{}'.format(test_case)].cell(2,1).value = test_case
        wb['F-{}'.format(test_case)].cell(2,2).value = card
        wb['F-{}'.format(test_case)].cell(2,3).value = config_type
        wb['F-{}'.format(test_case)].cell(2,4).value = card_action
        wb['F-{}'.format(test_case)].cell(2,5).value = append_data
        wb['F-{}'.format(test_case)].cell(2,5).alignment = Alignment(wrap_text=True)
    elif (f'F-{test_case}' in wb.sheetnames) and (test_case != 'Voice Bert' or test_case != 'Tone' or test_case != 'Ring'):
        if not hasattr(create_fail_info_sheet, 'call_counter'):
            create_fail_info_sheet.call_counter = {}
        create_fail_info_sheet.call_counter[test_case] = create_fail_info_sheet.call_counter.get(test_case, 3)
        wb['F-{}'.format(test_case)].cell(create_fail_info_sheet.call_counter[test_case],1).value = test_case
        wb['F-{}'.format(test_case)].cell(create_fail_info_sheet.call_counter[test_case],2).value = card
        wb['F-{}'.format(test_case)].cell(create_fail_info_sheet.call_counter[test_case],3).value = config_type
        wb['F-{}'.format(test_case)].cell(create_fail_info_sheet.call_counter[test_case],4).value = card_action
        wb['F-{}'.format(test_case)].cell(create_fail_info_sheet.call_counter[test_case],5).value = append_data
        wb['F-{}'.format(test_case)].cell(create_fail_info_sheet.call_counter[test_case],5).alignment = Alignment(wrap_text=True)
        create_fail_info_sheet.call_counter[test_case] += 1

def create_voice_fail_info_sheet(wb,test_case,card,config_type,card_action,au,ohm,append_data):
    if (('Fail-{}'.format(test_case)) not in wb.sheetnames) and (test_case == 'Voice Bert' or test_case == 'Tone' or test_case == 'Ring'):
        wb.create_sheet('F-{}'.format(test_case),index=100)
        set_title_for_sheet('Test Item',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
        set_title_for_sheet('Card',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
        set_title_for_sheet('Config type',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
        set_title_for_sheet('Action',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
        set_title_for_sheet('au-law',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
        set_title_for_sheet('Impedance',wb['F-{}'.format(test_case)],1,6,6,"5cb800")
        set_title_for_sheet('Fail info',wb['F-{}'.format(test_case)],1,7,7,"5cb800")
        wb['F-{}'.format(test_case)].cell(2,1).value = test_case
        wb['F-{}'.format(test_case)].cell(2,2).value = card
        wb['F-{}'.format(test_case)].cell(2,3).value = config_type
        wb['F-{}'.format(test_case)].cell(2,4).value = card_action
        wb['F-{}'.format(test_case)].cell(2,5).value = au
        wb['F-{}'.format(test_case)].cell(2,6).value = ohm
        wb['F-{}'.format(test_case)].cell(2,7).value = append_data
        wb['F-{}'.format(test_case)].cell(2,7).alignment = Alignment(wrap_text=True)
    elif (f'F-{test_case}' in wb.sheetnames) and (test_case == 'Voice Bert' or test_case == 'Tone' or test_case == 'Ring'):
        if not hasattr(create_voice_fail_info_sheet, 'call_counter'):
            create_voice_fail_info_sheet.call_counter = {}
        create_voice_fail_info_sheet.call_counter[test_case] = create_voice_fail_info_sheet.call_counter.get(test_case, 3)
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],1).value = test_case
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],2).value = card
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],3).value = config_type
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],4).value = card_action
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],5).value = au
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],6).value = ohm
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],7).value = append_data
        wb['F-{}'.format(test_case)].cell(create_voice_fail_info_sheet.call_counter[test_case],7).alignment = Alignment(wrap_text=True)
        create_voice_fail_info_sheet.call_counter[test_case] += 1

def create_bundlePW_fail_info_sheet(wb,test_case,vlan_mode,bundle_ids,action,Working_fail_data,Standby_fail_data):
    if f'F-{test_case}' not in wb.sheetnames and test_case in ['Bundle test(PW status)']:
        wb.create_sheet('F-{}'.format(test_case),index=100)
        second_name = ['DUT','CTRL','Primary/Secondary', "Jitter buffer", "L bit count", "R bit count", 
                       "TX Good count","RX Good count","RX lost count", "Link", "Fail reason","Result"]
        jitter_name = ['Under run','Over run','Depth','MIN','MAX']
        set_title_for_sheet('Bundle mode',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
        set_title_for_sheet('Bundle IDs',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
        set_title_for_sheet('Check time',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
        set_title_for_sheet('Working PW error info',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=4,end_column=4+len(second_name)+len(jitter_name)-2)
        set_title_for_sheet('Standby PW error info',wb['F-{}'.format(test_case)],1,20,20,"5cb800")
        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=20,end_column=20+len(second_name)+len(jitter_name)-2)
        
        wb['F-{}'.format(test_case)].cell(2,1).value = test_case
        wb['F-{}'.format(test_case)].cell(2,2).value = vlan_mode
        wb['F-{}'.format(test_case)].cell(2,3).value = bundle_ids
        wb['F-{}'.format(test_case)].cell(2,4).value = action
        wb['F-{}'.format(test_case)].cell(2,5).value = Working_fail_data
        wb['F-{}'.format(test_case)].cell(2,5).alignment = Alignment(wrap_text=True)
        wb['F-{}'.format(test_case)].cell(2,6).value = Standby_fail_data
        wb['F-{}'.format(test_case)].cell(2,6).alignment = Alignment(wrap_text=True)
    elif (f'F-{test_case}' in wb.sheetnames) and test_case in ['Bundle test(PW status)']:
        if not hasattr(create_bundlePW_fail_info_sheet, 'call_counter'):
            create_bundlePW_fail_info_sheet.call_counter = {}
        create_bundlePW_fail_info_sheet.call_counter[test_case] = create_bundlePW_fail_info_sheet.call_counter.get(test_case, 3)
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],1).value = test_case
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],2).value = vlan_mode
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],3).value = bundle_ids
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],4).value = action
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],5).value = Working_fail_data
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],5).alignment = Alignment(wrap_text=True)
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],6).value = Standby_fail_data
        wb['F-{}'.format(test_case)].cell(create_bundlePW_fail_info_sheet.call_counter[test_case],6).alignment = Alignment(wrap_text=True)
        create_bundlePW_fail_info_sheet.call_counter[test_case] += 1

def create_bundleRS485_fail_info_sheet(wb,test_case,vlan_mode,bundle_ids,action,RS485_fail_data):
    if f'F-{test_case}' not in wb.sheetnames and test_case in ['Bundle test(RS485 status)']:
        wb.create_sheet('F-{}'.format(test_case),index=100)
        set_title_for_sheet('Test Item',wb['F-{}'.format(test_case)],1,1,1,"5cb800")
        set_title_for_sheet('Bundle mode',wb['F-{}'.format(test_case)],1,2,2,"5cb800")
        set_title_for_sheet('Bundle IDs',wb['F-{}'.format(test_case)],1,3,3,"5cb800")
        set_title_for_sheet('Check time',wb['F-{}'.format(test_case)],1,4,4,"5cb800")
        set_title_for_sheet('RS485 error info',wb['F-{}'.format(test_case)],1,5,5,"5cb800")
        
        wb['F-{}'.format(test_case)].cell(2,1).value = test_case
        wb['F-{}'.format(test_case)].cell(2,2).value = vlan_mode
        wb['F-{}'.format(test_case)].cell(2,3).value = bundle_ids
        wb['F-{}'.format(test_case)].cell(2,4).value = action
        wb['F-{}'.format(test_case)].cell(2,5).value = RS485_fail_data
        wb['F-{}'.format(test_case)].cell(2,5).alignment = Alignment(wrap_text=True)

    elif (f'F-{test_case}' in wb.sheetnames) and test_case in ['Bundle test(RS485 status)']:
        if not hasattr(create_bundleRS485_fail_info_sheet, 'call_counter'):
            create_bundleRS485_fail_info_sheet.call_counter = {}
        create_bundleRS485_fail_info_sheet.call_counter[test_case] = create_bundleRS485_fail_info_sheet.call_counter.get(test_case, 3)
        wb['F-{}'.format(test_case)].cell(create_bundleRS485_fail_info_sheet.call_counter[test_case],1).value = test_case
        wb['F-{}'.format(test_case)].cell(create_bundleRS485_fail_info_sheet.call_counter[test_case],2).value = vlan_mode
        wb['F-{}'.format(test_case)].cell(create_bundleRS485_fail_info_sheet.call_counter[test_case],3).value = bundle_ids
        wb['F-{}'.format(test_case)].cell(create_bundleRS485_fail_info_sheet.call_counter[test_case],4).value = action
        wb['F-{}'.format(test_case)].cell(create_bundleRS485_fail_info_sheet.call_counter[test_case],5).value = RS485_fail_data
        wb['F-{}'.format(test_case)].cell(create_bundleRS485_fail_info_sheet.call_counter[test_case],5).alignment = Alignment(wrap_text=True)
        create_bundleRS485_fail_info_sheet.call_counter[test_case] += 1

def create_longtimebert_fail_info_sheet(wb,test_case,round,bert_type,bert_fail_data):
    if f'F-{test_case}' not in wb.sheetnames and test_case in ['Long Time(bert)']:
        wb.create_sheet('F-{}'.format(test_case),index=100)
        set_title_for_sheet('Test Item',wb['F-{}'.format(test_case)],1,1,1,"daeef3")
        set_title_for_sheet('Round',wb['F-{}'.format(test_case)],1,2,1,"daeef3")
        set_title_for_sheet('Bert type',wb['F-{}'.format(test_case)],1,3,1,"daeef3")
        for i in range(1,4):
            wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=2,start_column=i,end_column=i)
        set_title_for_sheet('Fail item',wb['F-{}'.format(test_case)],1,4,1,"daeef3")
        wb['F-{}'.format(test_case)].merge_cells(start_row=1,end_row=1,start_column=4,end_column=7)
        set_title_for_sheet('Time',wb['F-{}'.format(test_case)],2,4,1,"daeef3")
        set_title_for_sheet('Bit Error',wb['F-{}'.format(test_case)],2,5,1,"daeef3")
        set_title_for_sheet('Error ID',wb['F-{}'.format(test_case)],2,6,1,"daeef3")
        set_title_for_sheet('Elapsed Time',wb['F-{}'.format(test_case)],2,7,1,"daeef3")
        wb['F-{}'.format(test_case)].cell(3,1).value = test_case
        wb['F-{}'.format(test_case)].cell(3,2).value = int(round)
        wb['F-{}'.format(test_case)].cell(3,3).value = bert_type
        match = re.search(r'Time: (.*?) Bit Error: (\d+) Error ID: (\d+) Elapsed Time: (\d+)', bert_fail_data)
        if match:
            wb['F-{}'.format(test_case)].cell(3,4).value = match.group(1)
            wb['F-{}'.format(test_case)].cell(3,5).value = int(match.group(2))
            wb['F-{}'.format(test_case)].cell(3,6).value = int(match.group(3))
            wb['F-{}'.format(test_case)].cell(3,7).value = int(match.group(4))
    elif (f'F-{test_case}' in wb.sheetnames) and test_case in ['Long Time(bert)']:
        if not hasattr(create_longtimebert_fail_info_sheet, 'call_counter'):
            create_longtimebert_fail_info_sheet.call_counter = {}
        create_longtimebert_fail_info_sheet.call_counter[test_case] = create_longtimebert_fail_info_sheet.call_counter.get(test_case, 4)
        wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],1).value = test_case
        wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],2).value = int(round)
        wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],3).value = bert_type
        match = re.search(r'Time: (.*?) Bit Error: (\d+) Error ID: (\d+) Elapsed Time: (\d+)', bert_fail_data)
        if match:
            wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],4).value = match.group(1)
            wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],5).value = int(match.group(2))
            wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],6).value = int(match.group(3))
            wb['F-{}'.format(test_case)].cell(create_longtimebert_fail_info_sheet.call_counter[test_case],7).value = int(match.group(4))
        create_longtimebert_fail_info_sheet.call_counter[test_case] += 1
    # --- placed in center ---
    alignment = Alignment(horizontal='center', vertical='center')
    for row in wb['F-{}'.format(test_case)].iter_rows(min_row=1, max_row=wb['F-{}'.format(test_case)].max_row, min_col=1, max_col=wb['F-{}'.format(test_case)].max_column):
        for cell in row:
            cell.alignment = alignment
    wb['F-{}'.format(test_case)].freeze_panes = 'A3'